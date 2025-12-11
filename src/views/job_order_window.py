import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import filedialog
from src.utils.helpers import format_ton, setup_window_restore_behavior

class JobOrderWindow:
    def __init__(self, parent, db):
        self.db = db
        self.window = tk.Toplevel(parent)
        self.window.title("Job Order Management")

        # Setup window restore behavior (fix minimize/restore issue)
        setup_window_restore_behavior(self.window)

        # Set minimum window size
        self.window.minsize(800, 600)
        
        # Get screen dimensions for responsive sizing
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Set window size based on screen size (80% of screen)
        window_width = int(screen_width * 0.8)
        window_height = int(screen_height * 0.8)
        
        # Center window on screen
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Allow window to be resizable
        self.window.resizable(True, True)
        
        # Selected JOA
        self.selected_joa = None
        self.current_data = []
        
        self.setup_ui()
        self.load_joa_list()
    
    def parse_party(self, party_text):
        import re

        if not party_text or party_text == '-':
            return (0, 0, 0, '-')

        count_20 = 0
        count_21 = 0
        count_40 = 0

        # Pattern untuk container type yang exact match (case insensitive)
        party_upper = party_text.strip().upper()

        # Deteksi single container type
        if party_upper == "20'" or party_upper == "20":
            count_20 = 1
        elif party_upper == "21'" or party_upper == "21":
            count_21 = 1
        elif party_upper in ("40'HC", "40HC", "40'", "40"):
            count_40 = 1

        # Pattern untuk "3X40", "5x40", "3x40'", dll
        matches_40 = re.findall(r"(\d+)\s*[xX]\s*40['\"]?", party_text, re.IGNORECASE)
        if matches_40:
            count_40 = sum(int(m) for m in matches_40)

        # Pattern untuk "2X20", "4x20", "2x20'", dll
        matches_20 = re.findall(r"(\d+)\s*[xX]\s*20['\"]?", party_text)
        if matches_20:
            count_20 = sum(int(m) for m in matches_20)

        # Pattern untuk "2X21", "4x21", "2x21'", dll
        matches_21 = re.findall(r"(\d+)\s*[xX]\s*21['\"]?", party_text)
        if matches_21:
            count_21 = sum(int(m) for m in matches_21)

        # Build display text
        parts = []
        if count_20 > 0:
            parts.append(f"{count_20} X 20'")
        if count_21 > 0:
            parts.append(f"{count_21} X 21'")
        if count_40 > 0:
            parts.append(f"{count_40} X 40'HC")

        if parts:
            total = count_20 + count_21 + count_40
            # Format lengkap dengan total untuk display di header
            display_full = " + ".join(parts) + f" (Total: {total} container{'s' if total > 1 else ''})"
            # Format singkat untuk kolom UNIT (tanpa total)
            display_short = " + ".join(parts)
        else:
            display_full = party_text  # Fallback ke text asli jika tidak bisa di-parse
            display_short = party_text

        return (count_20, count_21, count_40, display_full, display_short)
    
    def setup_ui(self):
        """Setup user interface"""
        # Main container
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights for responsive layout
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)

        # Responsive minsize based on window width
        screen_width = self.window.winfo_screenwidth()
        left_minsize = max(200, int(screen_width * 0.15))
        right_minsize = max(400, int(screen_width * 0.30))

        main_frame.columnconfigure(0, weight=1, minsize=left_minsize)  # Left panel responsif
        main_frame.columnconfigure(1, weight=3, minsize=right_minsize)  # Right panel responsif
        main_frame.rowconfigure(0, weight=1)
        
        # ===== LEFT PANEL - JOA LIST =====
        left_frame = ttk.LabelFrame(main_frame, text="Daftar Job Order Account (JOA)", padding="10")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        # Search frame
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Cari JOA:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_joa_list())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # JOA Listbox with scrollbar
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.joa_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                                       font=('Arial', 10))
        self.joa_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.joa_listbox.yview)
        
        self.joa_listbox.bind('<<ListboxSelect>>', self.on_joa_select)
        
        # Buttons frame
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="Refresh", command=self.load_joa_list).pack(fill=tk.X, pady=2)
        
        # ===== RIGHT PANEL - DETAILS =====
        right_frame = ttk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(1, weight=1)
        
        # Info frame
        info_frame = ttk.LabelFrame(right_frame, text="Informasi Job Order", padding="10")
        info_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # JOA Info
        info_grid = ttk.Frame(info_frame)
        info_grid.pack(fill=tk.X)
        info_grid.columnconfigure(1, weight=1)  # Kolom value bisa expand
        info_grid.columnconfigure(3, weight=1)  # Kolom value bisa expand
        
        ttk.Label(info_grid, text="No. JOA:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=5)
        self.joa_label = ttk.Label(info_grid, text="-", font=('Arial', 10))
        self.joa_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        
        ttk.Label(info_grid, text="Feeder:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky=tk.W, padx=5)
        self.feeder_label = ttk.Label(info_grid, text="-", font=('Arial', 10))
        self.feeder_label.grid(row=1, column=1, sticky=tk.W, padx=5)
        
        ttk.Label(info_grid, text="Destination:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky=tk.W, padx=5)
        self.destination_label = ttk.Label(info_grid, text="-", font=('Arial', 10))
        self.destination_label.grid(row=1, column=3, sticky=tk.W, padx=5)
        
        # Tables Container - Split into TOP (Sales) and BOTTOM (Purchase)
        tables_container = ttk.Frame(right_frame)
        tables_container.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tables_container.columnconfigure(0, weight=1)
        tables_container.rowconfigure(0, weight=1)  # Sales Invoice
        tables_container.rowconfigure(1, weight=1)  # Purchase Invoice
        
        # ===== TOP TABLE - SALES INVOICE =====
        sales_main_frame = ttk.Frame(tables_container)
        sales_main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 5))
        sales_main_frame.columnconfigure(0, weight=1)
        sales_main_frame.rowconfigure(1, weight=1)
        
        # Header Frame with Document Info
        sales_header_frame = ttk.LabelFrame(sales_main_frame, text="JOB ORDER ACCOUNT - SALES INVOICE", padding="10")
        sales_header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        sales_header_frame.columnconfigure(1, weight=1)
        sales_header_frame.columnconfigure(3, weight=1)
        
        # Left side info
        ttk.Label(sales_header_frame, text="SHIPPING LINE:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.shipping_line_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.shipping_line_label.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(sales_header_frame, text="NAMA KAPAL:", font=('Arial', 9, 'bold')).grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.nama_kapal_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.nama_kapal_label.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(sales_header_frame, text="PARTY:", font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.party_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.party_label.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(sales_header_frame, text="NO. JOA:", font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.no_joa_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.no_joa_label.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Right side info
        ttk.Label(sales_header_frame, text="ETD:", font=('Arial', 9, 'bold')).grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        self.etd_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.etd_label.grid(row=0, column=3, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(sales_header_frame, text="ETA:", font=('Arial', 9, 'bold')).grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)
        self.eta_label = ttk.Label(sales_header_frame, text="-", font=('Arial', 9))
        self.eta_label.grid(row=1, column=3, sticky=tk.W, padx=5, pady=2)
        
        # Sales Data Table
        sales_table_frame = ttk.Frame(sales_main_frame)
        sales_table_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sales_table_frame.columnconfigure(0, weight=1)
        sales_table_frame.rowconfigure(0, weight=1)

        # Scrollbars untuk Sales
        sales_scroll_y = ttk.Scrollbar(sales_table_frame, orient=tk.VERTICAL)
        sales_scroll_x = ttk.Scrollbar(sales_table_frame, orient=tk.HORIZONTAL)

        # ✅ Sales Treeview TANPA kolom INVOICE (responsive height)
        sales_columns = ('customer', 'rp', 'kubikasi', 'tonase')
        self.sales_tree = ttk.Treeview(sales_table_frame, columns=sales_columns, show='headings',
                                       yscrollcommand=sales_scroll_y.set,
                                       xscrollcommand=sales_scroll_x.set)
        
        sales_scroll_y.config(command=self.sales_tree.yview)
        sales_scroll_x.config(command=self.sales_tree.xview)
        
        # Configure Sales columns (responsive widths)
        self.sales_tree.heading('customer', text='CUSTOMER')
        self.sales_tree.heading('rp', text='Rp')
        self.sales_tree.heading('kubikasi', text='KUBIKASI')
        self.sales_tree.heading('tonase', text='TONASE')

        # Responsive column widths based on screen
        col_base_width = max(80, int(screen_width * 0.08))
        self.sales_tree.column('customer', width=int(col_base_width * 2.5), anchor=tk.W, minwidth=150)
        self.sales_tree.column('rp', width=int(col_base_width * 1.5), anchor=tk.E, minwidth=100)
        self.sales_tree.column('kubikasi', width=col_base_width, anchor=tk.E, minwidth=80)
        self.sales_tree.column('tonase', width=col_base_width, anchor=tk.E, minwidth=80)
        
        # Grid layout untuk sales
        self.sales_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sales_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        sales_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Sales Total Frame
        sales_total_frame = ttk.Frame(sales_main_frame)
        sales_total_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Sales Grand Total (kanan)
        sales_grand_frame = ttk.Frame(sales_total_frame)
        sales_grand_frame.pack(side=tk.RIGHT, padx=10)
        ttk.Label(sales_grand_frame, text="SALES TOTAL:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.sales_total_label = ttk.Label(sales_grand_frame, text="Rp 0", font=('Arial', 10, 'bold'), foreground='blue')
        self.sales_total_label.pack(side=tk.LEFT)
        
        # ===== BOTTOM TABLE - PURCHASE INVOICE =====
        purchase_main_frame = ttk.Frame(tables_container)
        purchase_main_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(5, 0))
        purchase_main_frame.columnconfigure(0, weight=1)
        purchase_main_frame.columnconfigure(1, weight=1)
        purchase_main_frame.rowconfigure(1, weight=1)
        
        # Main header for Purchase
        purchase_title = ttk.Label(purchase_main_frame, text="PURCHASE INVOICE", 
                                   font=('Arial', 10, 'bold'))
        purchase_title.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # ===== LEFT SIDE - BIAYA POL =====
        pol_frame = ttk.LabelFrame(purchase_main_frame, text="BIAYA POL", padding="5")
        pol_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        pol_frame.columnconfigure(0, weight=1)
        pol_frame.rowconfigure(0, weight=1)
        
        # POL Scrollbars
        pol_scroll_y = ttk.Scrollbar(pol_frame, orient=tk.VERTICAL)
        pol_scroll_x = ttk.Scrollbar(pol_frame, orient=tk.HORIZONTAL)
        
        # POL Treeview (responsive height)
        pol_columns = ('keterangan', 'unit', 'biaya_pol', 'total')
        self.pol_tree = ttk.Treeview(pol_frame, columns=pol_columns, show='headings',
                                     yscrollcommand=pol_scroll_y.set,
                                     xscrollcommand=pol_scroll_x.set)
        
        pol_scroll_y.config(command=self.pol_tree.yview)
        pol_scroll_x.config(command=self.pol_tree.xview)
        
        # Configure POL columns (responsive widths)
        self.pol_tree.heading('keterangan', text='KETERANGAN')
        self.pol_tree.heading('unit', text='UNIT')
        self.pol_tree.heading('biaya_pol', text='BIAYA POL')
        self.pol_tree.heading('total', text='TOTAL')

        self.pol_tree.column('keterangan', width=int(col_base_width * 1.8), anchor=tk.W, minwidth=120)
        self.pol_tree.column('unit', width=int(col_base_width * 0.8), anchor=tk.CENTER, minwidth=60)
        self.pol_tree.column('biaya_pol', width=int(col_base_width * 1.2), anchor=tk.E, minwidth=90)
        self.pol_tree.column('total', width=int(col_base_width * 1.2), anchor=tk.E, minwidth=90)
        
        # Grid layout untuk POL
        self.pol_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        pol_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        pol_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # ===== RIGHT SIDE - BIAYA POD =====
        pod_frame = ttk.LabelFrame(purchase_main_frame, text="BIAYA POD", padding="5")
        pod_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))
        pod_frame.columnconfigure(0, weight=1)
        pod_frame.rowconfigure(0, weight=1)
        
        # POD Scrollbars
        pod_scroll_y = ttk.Scrollbar(pod_frame, orient=tk.VERTICAL)
        pod_scroll_x = ttk.Scrollbar(pod_frame, orient=tk.HORIZONTAL)
        
        # POD Treeview (responsive height)
        pod_columns = ('keterangan', 'unit', 'biaya_pod', 'total')
        self.pod_tree = ttk.Treeview(pod_frame, columns=pod_columns, show='headings',
                                     yscrollcommand=pod_scroll_y.set,
                                     xscrollcommand=pod_scroll_x.set)
        
        pod_scroll_y.config(command=self.pod_tree.yview)
        pod_scroll_x.config(command=self.pod_tree.xview)
        
        # Configure POD columns (responsive widths)
        self.pod_tree.heading('keterangan', text='KETERANGAN')
        self.pod_tree.heading('unit', text='UNIT')
        self.pod_tree.heading('biaya_pod', text='BIAYA POD')
        self.pod_tree.heading('total', text='TOTAL')

        self.pod_tree.column('keterangan', width=int(col_base_width * 1.8), anchor=tk.W, minwidth=120)
        self.pod_tree.column('unit', width=int(col_base_width * 0.8), anchor=tk.CENTER, minwidth=60)
        self.pod_tree.column('biaya_pod', width=int(col_base_width * 1.2), anchor=tk.E, minwidth=90)
        self.pod_tree.column('total', width=int(col_base_width * 1.2), anchor=tk.E, minwidth=90)
        
        # Grid layout untuk POD
        self.pod_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        pod_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        pod_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Purchase Total Frame
        purchase_total_frame = ttk.Frame(purchase_main_frame)
        purchase_total_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # POL Total (kiri)
        pol_total_frame = ttk.Frame(purchase_total_frame)
        pol_total_frame.pack(side=tk.LEFT, padx=10)
        ttk.Label(pol_total_frame, text="TOTAL POL:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.pol_purchase_total_label = ttk.Label(pol_total_frame, text="Rp 0", font=('Arial', 9, 'bold'), foreground='green')
        self.pol_purchase_total_label.pack(side=tk.LEFT)
        
        # POD Total (tengah)
        pod_total_frame = ttk.Frame(purchase_total_frame)
        pod_total_frame.pack(side=tk.LEFT, padx=10)
        ttk.Label(pod_total_frame, text="TOTAL POD:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.pod_purchase_total_label = ttk.Label(pod_total_frame, text="Rp 0", font=('Arial', 9, 'bold'), foreground='orange')
        self.pod_purchase_total_label.pack(side=tk.LEFT)
        
        # Grand Total (kanan)
        purchase_grand_frame = ttk.Frame(purchase_total_frame)
        purchase_grand_frame.pack(side=tk.RIGHT, padx=10)
        ttk.Label(purchase_grand_frame, text="PURCHASE TOTAL:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.purchase_total_label = ttk.Label(purchase_grand_frame, text="Rp 0", font=('Arial', 10, 'bold'), foreground='red')
        self.purchase_total_label.pack(side=tk.LEFT)
        
        # ===== BOTTOM BUTTONS =====
        buttons_frame = ttk.Frame(right_frame)
        buttons_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))

        ttk.Button(buttons_frame, text="📄 Export to Excel",
                  command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_joa_list(self):
        """Load list of JOAs"""
        try:
            query = """
                SELECT DISTINCT c.ref_joa, c.container_id, c.container, c.kapal_id, c.party
                FROM containers c
                WHERE c.ref_joa IS NOT NULL AND c.ref_joa != ''
                ORDER BY c.ref_joa
            """
            
            containers = self.db.execute(query)
            
            # Group containers by JOA
            self.joa_data = {}
            for row in containers:
                joa = row['ref_joa']
                if joa not in self.joa_data:
                    self.joa_data[joa] = {
                        'containers': [],
                        'container': None,
                        'feeder': None,
                        'destination': None,
                        'etd_sub': None,
                        'kapal_id': None,
                        'party': None,
                        'shipping_line': None
                    }
                
                container_id = row['container_id']
                container = row['container']
                kapal_id = row['kapal_id'] if row['kapal_id'] else None
                party = row['party'] if row['party'] else None
                
                self.joa_data[joa]['containers'].append(container)
                
                if self.joa_data[joa]['kapal_id'] is None:
                    self.joa_data[joa]['kapal_id'] = kapal_id
                if self.joa_data[joa]['party'] is None:
                    self.joa_data[joa]['party'] = party
                
                if kapal_id:
                    try:
                        kapal_query = """
                            SELECT feeder, destination, etd_sub, cls, open, full, shipping_line
                            FROM kapals
                            WHERE kapal_id = ?
                        """

                        kapal_result = self.db.execute_one(kapal_query, (kapal_id,))

                        if kapal_result:
                            if self.joa_data[joa]['feeder'] is None:
                                self.joa_data[joa]['feeder'] = kapal_result['feeder']
                                self.joa_data[joa]['destination'] = kapal_result['destination']
                                self.joa_data[joa]['etd_sub'] = kapal_result['etd_sub']
                                self.joa_data[joa]['cls'] = kapal_result['cls']
                                self.joa_data[joa]['open'] = kapal_result['open']
                                self.joa_data[joa]['full'] = kapal_result['full']
                                self.joa_data[joa]['shipping_line'] = kapal_result['shipping_line']
                    except Exception as e:
                        print(f"ERROR getting kapal info: {e}")
                
                if self.joa_data[joa]['container'] is None:
                    self.joa_data[joa]['container'] = container
            
            # Populate listbox
            self.joa_listbox.delete(0, tk.END)
            for joa in sorted(self.joa_data.keys()):
                self.joa_listbox.insert(tk.END, joa)
            
            print(f"Loaded {len(self.joa_data)} JOAs")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal memuat daftar JOA:\n{str(e)}")
    
    def filter_joa_list(self):
        """Filter JOA list based on search"""
        search_text = self.search_var.get().lower()
        
        self.joa_listbox.delete(0, tk.END)
        for joa in sorted(self.joa_data.keys()):
            if search_text in joa.lower():
                self.joa_listbox.insert(tk.END, joa)
    
    def on_joa_select(self, event):
        """Handle JOA selection"""
        selection = self.joa_listbox.curselection()
        if selection:
            index = selection[0]
            self.selected_joa = self.joa_listbox.get(index)
            self.load_joa_details(self.selected_joa)
    
    def load_joa_details(self, joa):
        """Load details for selected JOA (Sales + Purchase)"""
        try:
            print(f"\n===== LOADING JOA: {joa} =====")

            # Get JOA info
            joa_info = self.joa_data[joa]

            # Update header info
            self.joa_label.config(text=joa)
            containers_text = ", ".join(joa_info['containers'])
            self.feeder_label.config(text=joa_info.get('feeder', '-') or '-')
            self.destination_label.config(text=joa_info.get('destination', '-') or '-')

            # Ambil semua containers untuk JOA ini
            query = """
                SELECT container_id, container, party
                FROM containers 
                WHERE ref_joa = ?
            """
            containers = self.db.execute(query, (joa,))
            container_ids = [c['container_id'] for c in containers]

            # Hitung jumlah container per tipe (20', 21', 40'HC)
            total_20, total_21, total_40 = 0, 0, 0
            for cont in containers:
                p = (cont['party'] or '').strip().upper()
                # Prioritas: 40'HC > 21' > 20'
                if '40' in p:
                    total_40 += 1
                elif '21' in p:
                    total_21 += 1
                elif '20' in p:
                    total_20 += 1

            # Buat string party
            party_parts = []
            if total_20 > 0:
                party_parts.append(f"{total_20} X 20'")
            if total_21 > 0:
                party_parts.append(f"{total_21} X 21'")
            if total_40 > 0:
                party_parts.append(f"{total_40} X 40'HC")
            party_display = " + ".join(party_parts) if party_parts else "-"

            joa_info['party_count_20'] = total_20
            joa_info['party_count_21'] = total_21
            joa_info['party_count_40'] = total_40
            joa_info['party'] = party_display
            print(f"Calculated Party: {party_display} (20'={total_20}, 21'={total_21}, 40'HC={total_40})")

            # Update header
            self.shipping_line_label.config(text=joa_info.get('shipping_line', '-') or '-')
            self.nama_kapal_label.config(text=joa_info.get('feeder', '-') or '-')
            self.party_label.config(text=party_display)
            self.no_joa_label.config(text=joa)
            self.etd_label.config(text=joa_info.get('etd_sub', '-') or '-')
            self.eta_label.config(text=joa_info.get('eta', '-') or '-')

            # Bersihkan treeview lama
            for tree in (self.sales_tree, self.pol_tree, self.pod_tree):
                for item in tree.get_children():
                    tree.delete(item)

            if not container_ids:
                print("No containers found!")
                return

            # ============================================================
            # SALES INVOICE
            # ============================================================
            sales_total, pol_total, pod_total = 0, 0, 0
            placeholders = ','.join(['?' for _ in container_ids])

            # Cari distinct customer (penerima)
            customer_query = f"""
                SELECT DISTINCT
                    b.penerima as customer_id,
                    COALESCE(c.nama_customer, b.penerima) as customer_name
                FROM barang b
                JOIN detail_container dc ON b.barang_id = dc.barang_id
                JOIN customers c ON c.customer_id = b.penerima
                WHERE dc.container_id IN ({placeholders})
                ORDER BY customer_name
            """
            customers = self.db.execute(customer_query, tuple(container_ids))
            print(f"Found {len(customers)} customers")

            for customer in customers:
                customer_id = customer['customer_id']
                customer_name = customer['customer_name']

                # Ambil barang untuk customer ini
                barang_query = f"""
                    SELECT
                        b.barang_id,
                        b.nama_barang,
                        b.m3_barang,
                        b.ton_barang,
                        dc.door_type,
                        dc.colli_amount,
                        dc.total_harga
                    FROM barang b
                    JOIN detail_container dc ON b.barang_id = dc.barang_id
                    WHERE b.penerima = ? AND dc.container_id IN ({placeholders})
                """
                barangs = self.db.execute(barang_query, (customer_id, *container_ids))

                if not barangs:
                    continue

                total_m3, total_ton, total_colli, total_invoice = 0, 0, 0, 0
                for b in barangs:
                    total_colli += b['colli_amount']
                    if b['m3_barang']:
                        total_m3 += b['m3_barang'] * b['colli_amount']
                    if b['ton_barang']:
                        total_ton += b['ton_barang'] * b['colli_amount']
                    total_invoice += b['total_harga'] if b['total_harga'] else 0

                # Door type → tentukan POL / POD
                door_type = barangs[0]['door_type'] if barangs[0]['door_type'] else 'PP'
                if door_type == 'PP':
                    pol_total += total_invoice
                else:
                    pod_total += total_invoice

                sales_total += total_invoice

                # ✅ Insert ke sales_tree TANPA INVOICE
                self.sales_tree.insert('', 'end', values=(
                    customer_name,
                    f"Rp {total_invoice:,.0f}",
                    f"{total_m3:.4f}" if total_m3 else "-",
                    format_ton(total_ton) if total_ton else "-"
                ))

            # Update sales labels
            self.sales_total_label.config(text=f"Rp {sales_total:,.0f}")

            # ============================================================
            # PURCHASE INVOICE - LOGIKA POL vs POD BERDASARKAN LOKASI
            # ============================================================
            purchase_pol_total, purchase_pod_total, purchase_total = 0, 0, 0
            
            delivery_cost_query = f"""
                SELECT 
                    cdc.description as cost_type,
                    cdc.delivery as delivery_type,
                    cdc.cost as cost_amount,
                    cdc.cost_description as notes,
                    cont.party as party
                FROM container_delivery_costs cdc
                JOIN containers cont ON cont.container_id = cdc.container_id
                WHERE cdc.container_id IN ({placeholders})
            """
            delivery_costs = self.db.execute(delivery_cost_query, tuple(container_ids))

            # ✅ LOGIKA: Pakai kolom delivery untuk tentukan POL vs POD
            # POL = Lokasi Surabaya (atau kosong)
            # POD = Lokasi selain Surabaya
            
            cost_summary_pol, cost_summary_pod = {}, {}
            
            for cost in delivery_costs:
                cost_type = cost['cost_type'] or ''
                delivery_type = cost['delivery_type'] or ''
                amount = cost['cost_amount'] or 0
                notes = cost['notes'] or ''
                party = cost['party'] or ''

                # Deteksi ukuran container
                size = '40' if '40' in party else '21' if '21' in party else '20' if '20' in party else None
                
                # ✅ PERBAIKAN LOGIKA:
                # Jika delivery_type kosong atau mengandung "Surabaya" → POL
                # Jika delivery_type mengandung nama kota lain → POD
                delivery_lower = delivery_type.lower()
                
                is_pol = (
                    not delivery_type or  # Kosong = default Surabaya = POL
                    'surabaya' in delivery_lower or 
                    'sby' in delivery_lower or
                    'sub' in delivery_lower
                )
                
                print(f"DEBUG Cost: '{cost_type}' | Delivery: '{delivery_type}' | Lokasi: {'POL (Surabaya)' if is_pol else 'POD (Destination)'}")
                
                key = (cost_type, size)
                target = cost_summary_pol if is_pol else cost_summary_pod
                
                if key not in target:
                    target[key] = {'notes': notes, 'size': size, 'unit_cost': amount}

            # POL costs
            for (cost_type, size), data in cost_summary_pol.items():
                # Determine quantity based on container size/type
                qty_map = {'20': total_20, '21': total_21, '40': total_40}
                qty = qty_map.get(size, 1)
                unit = f"{qty}x {size}'" if size else "1 invoice"
                total = data['unit_cost'] * qty
                ket = f"{cost_type} {size}" if size else cost_type

                purchase_pol_total += total
                purchase_total += total
                self.pol_tree.insert('', 'end', values=(
                    ket,
                    unit,
                    f"Rp {data['unit_cost']:,.0f}",
                    f"Rp {total:,.0f}"
                ))

            # POD costs
            for (cost_type, size), data in cost_summary_pod.items():
                # Determine quantity based on container size/type
                qty_map = {'20': total_20, '21': total_21, '40': total_40}
                qty = qty_map.get(size, 1)
                unit = f"{qty}x {size}'" if size else "1 invoice"
                total = data['unit_cost'] * qty
                ket = f"{cost_type} {size}" if size else cost_type

                purchase_pod_total += total
                purchase_total += total
                self.pod_tree.insert('', 'end', values=(
                    ket,
                    unit,
                    f"Rp {data['unit_cost']:,.0f}",
                    f"Rp {total:,.0f}"
                ))

            # Update purchase labels
            self.pol_purchase_total_label.config(text=f"Rp {purchase_pol_total:,.0f}")
            self.pod_purchase_total_label.config(text=f"Rp {purchase_pod_total:,.0f}")
            self.purchase_total_label.config(text=f"Rp {purchase_total:,.0f}")

            print(f"\n{'='*60}")
            print(f"SUMMARY:")
            print(f"{'='*60}")
            print(f"Sales Total        : Rp {sales_total:,.0f}")
            print(f"  - POL            : Rp {pol_total:,.0f}")
            print(f"  - POD            : Rp {pod_total:,.0f}")
            print(f"Purchase POL Total : Rp {purchase_pol_total:,.0f}")
            print(f"Purchase POD Total : Rp {purchase_pod_total:,.0f}")
            print(f"Purchase Total     : Rp {purchase_total:,.0f}")
            print(f"{'='*60}\n")

        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"ERROR in load_joa_details: {str(e)}")
            messagebox.showerror("Error", f"Gagal memuat detail JOA:\n{str(e)}")

    def refresh_current_joa(self):
        """Refresh current JOA data"""
        if self.selected_joa:
            self.load_joa_details(self.selected_joa)
    
    def export_to_excel(self):
        """Export current JOA to Excel"""
        if not self.selected_joa:
            messagebox.showwarning("Warning", "Pilih JOA terlebih dahulu!")
            return
        
        try:
            import time
            import os
            import platform
            import subprocess
            
            # Generate filename dengan epoch timestamp
            epoch = int(time.time())
            default_filename = f"JOA_{epoch}.xlsx"
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename
            )
            
            if not filename:
                return
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Job Order Account"
            
            # Get JOA info
            joa_info = {
                'joa': self.joa_label.cget('text'),
                'feeder': self.feeder_label.cget('text'),
                'destination': self.destination_label.cget('text'),
                'shipping_line': self.shipping_line_label.cget('text'),
                'etd_sub': self.etd_label.cget('text'),
                'eta_sub': self.eta_label.cget('text'),
                'party_display': self.party_label.cget('text'),
            }
            
            # Create side-by-side layout
            self._create_side_by_side_from_ui(ws, joa_info)
            
            # Save
            wb.save(filename)
            
            # Tanya user apakah mau buka file
            response = messagebox.askyesno(
                "Export Berhasil", 
                f"File berhasil disimpan:\n{filename}\n\nApakah Anda ingin membuka file sekarang?"
            )
            
            if response:
                try:
                    if platform.system() == 'Windows':
                        os.startfile(filename)
                    elif platform.system() == 'Darwin':  # macOS
                        subprocess.run(['open', filename])
                    else:  # Linux
                        subprocess.run(['xdg-open', filename])
                except Exception as e:
                    print(f"Error opening file: {e}")
                    messagebox.showinfo("Info", "File berhasil disimpan, tapi gagal membuka otomatis.")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal mengekspor ke Excel:\n{str(e)}")

    def _apply_border_to_range(self, ws, start_row, start_col, end_row, end_col, border):
        """Helper function to apply border to a range of cells"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border

    def _create_side_by_side_from_ui(self, ws, joa_info):
        """Create Sales and Purchase Invoice side by side"""
        # Styling
        header_font = Font(bold=True, size=10)
        title_font = Font(bold=True, size=12)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== LEFT SIDE: SALES INVOICE ====================
        
        # Title Sales
        ws.merge_cells('A1:D1')
        title_sales = ws['A1']
        title_sales.value = "JOB ORDER ACCOUNT - SALES INVOICE"
        title_sales.font = title_font
        title_sales.alignment = Alignment(horizontal='center', vertical='center')
        self._apply_border_to_range(ws, 1, 1, 1, 4, border_thin)
        ws.row_dimensions[1].height = 25
        
        # Sales Document Info
        row = 3
        ws[f'A{row}'] = "SHIPPING LINE:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = joa_info.get('shipping_line', '-')
        ws[f'C{row}'] = "ETD:"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = joa_info.get('etd_sub', '-')
        
        row += 1
        ws[f'A{row}'] = "NAMA KAPAL:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = joa_info.get('feeder', '-')
        ws[f'C{row}'] = "ETA:"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = joa_info.get('eta_sub', '-')
        
        row += 1
        ws[f'A{row}'] = "PARTY:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = joa_info.get('party_display', '-')
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "NO. JOA:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = joa_info.get('joa', '-')
        ws.merge_cells(f'B{row}:D{row}')
        
        # ✅ Sales Table Headers TANPA INVOICE
        row += 2
        headers_sales = ['CUSTOMER', 'Rp', 'KUBIKASI', 'TONASE']
        for col_idx, header in enumerate(headers_sales, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ambil Sales Data dari Treeview
        row += 1
        sales_start_row = row
        
        for item in self.sales_tree.get_children():
            values = self.sales_tree.item(item)['values']
            
            # Customer name
            cell = ws.cell(row=row, column=1, value=values[0])
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Rp
            rp_str = str(values[1]).replace('Rp', '').replace(',', '').strip()
            try:
                rp_val = float(rp_str)
            except:
                rp_val = 0
            
            cell = ws.cell(row=row, column=2, value=rp_val)
            cell.border = border_thin
            cell.number_format = '"Rp "#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Kubikasi
            kubikasi_str = str(values[2]).replace('-', '0')
            try:
                kubikasi = float(kubikasi_str)
            except:
                kubikasi = 0
            
            cell = ws.cell(row=row, column=3, value=kubikasi)
            cell.border = border_thin
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Tonase
            tonase_str = str(values[3]).replace('-', '0')
            try:
                tonase = float(tonase_str)
            except:
                tonase = 0
            
            cell = ws.cell(row=row, column=4, value=tonase)
            cell.border = border_thin
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        sales_end_row = row
        
        # ==================== RIGHT SIDE: PURCHASE INVOICE ====================
        
        # Title Purchase
        ws.merge_cells('F1:I1')
        title_purchase = ws['F1']
        title_purchase.value = "JOB ORDER ACCOUNT - PURCHASE INVOICE"
        title_purchase.font = title_font
        title_purchase.alignment = Alignment(horizontal='center', vertical='center')
        self._apply_border_to_range(ws, 1, 6, 1, 9, border_thin)
        
        # Purchase Document Info
        row = 3
        ws[f'F{row}'] = "SHIPPING LINE:"
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = joa_info.get('shipping_line', '-')
        ws[f'H{row}'] = "ETD:"
        ws[f'H{row}'].font = Font(bold=True)
        ws[f'I{row}'] = joa_info.get('etd_sub', '-')
        
        row += 1
        ws[f'F{row}'] = "NAMA KAPAL:"
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = joa_info.get('feeder', '-')
        ws[f'H{row}'] = "ETA:"
        ws[f'H{row}'].font = Font(bold=True)
        ws[f'I{row}'] = joa_info.get('eta_sub', '-')
        
        row += 1
        ws[f'F{row}'] = "PARTY:"
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = joa_info.get('party_display', '-')
        ws.merge_cells(f'G{row}:I{row}')
        
        row += 1
        ws[f'F{row}'] = "NO. JOA:"
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'G{row}'] = joa_info.get('joa', '-')
        ws.merge_cells(f'G{row}:I{row}')
        
        # ===== BIAYA POL SECTION =====
        row += 2
        ws.merge_cells(f'F{row}:I{row}')
        section_pol = ws[f'F{row}']
        section_pol.value = "BIAYA POL"
        section_pol.font = Font(bold=True, size=11)
        section_pol.alignment = Alignment(horizontal='center', vertical='center')
        self._apply_border_to_range(ws, row, 6, row, 9, border_thin)
        
        # POL Headers
        row += 1
        pol_headers = ['KETERANGAN', 'UNIT', 'BIAYA POL', 'TOTAL']
        for col_idx, header in enumerate(pol_headers, start=6):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ambil POL Data dari Treeview
        row += 1
        pol_start_row = row
        
        for item in self.pol_tree.get_children():
            values = self.pol_tree.item(item)['values']
            
            # Keterangan
            cell = ws.cell(row=row, column=6, value=values[0])
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Unit
            cell = ws.cell(row=row, column=7, value=values[1])
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Biaya POL
            biaya_str = str(values[2]).replace('Rp', '').replace(',', '').strip()
            try:
                biaya = float(biaya_str)
            except:
                biaya = 0
            
            cell = ws.cell(row=row, column=8, value=biaya)
            cell.border = border_thin
            cell.number_format = '"Rp "#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Total
            total_str = str(values[3]).replace('Rp', '').replace(',', '').strip()
            try:
                total = float(total_str)
            except:
                total = 0
            
            cell = ws.cell(row=row, column=9, value=total)
            cell.border = border_thin
            cell.number_format = '"Rp "#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        pol_end_row = row
        
        # ===== BIAYA POD SECTION =====
        row += 1
        ws.merge_cells(f'F{row}:I{row}')
        section_pod = ws[f'F{row}']
        section_pod.value = "BIAYA POD"
        section_pod.font = Font(bold=True, size=11)
        section_pod.alignment = Alignment(horizontal='center', vertical='center')
        self._apply_border_to_range(ws, row, 6, row, 9, border_thin)
        
        # POD Headers
        row += 1
        pod_headers = ['KETERANGAN', 'UNIT', 'BIAYA POD', 'TOTAL']
        for col_idx, header in enumerate(pod_headers, start=6):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ambil POD Data dari Treeview
        row += 1
        pod_start_row = row
        
        for item in self.pod_tree.get_children():
            values = self.pod_tree.item(item)['values']
            
            # Keterangan
            cell = ws.cell(row=row, column=6, value=values[0])
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Unit
            cell = ws.cell(row=row, column=7, value=values[1])
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Biaya POD
            biaya_str = str(values[2]).replace('Rp', '').replace(',', '').strip()
            try:
                biaya = float(biaya_str)
            except:
                biaya = 0
            
            cell = ws.cell(row=row, column=8, value=biaya)
            cell.border = border_thin
            cell.number_format = '"Rp "#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Total
            total_str = str(values[3]).replace('Rp', '').replace(',', '').strip()
            try:
                total = float(total_str)
            except:
                total = 0
            
            cell = ws.cell(row=row, column=9, value=total)
            cell.border = border_thin
            cell.number_format = '"Rp "#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        pod_end_row = row
        
        # ===== TOTALS =====
        max_row = max(sales_end_row, pod_end_row)
        
        # TOTAL POL
        row = pol_end_row
        cell = ws.cell(row=row, column=6, value="TOTAL POL:")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        pol_total_str = self.pol_purchase_total_label.cget('text').replace('Rp', '').replace(',', '').strip()
        try:
            pol_total_val = float(pol_total_str)
        except:
            pol_total_val = 0
        
        cell = ws.cell(row=row, column=9, value=pol_total_val)
        cell.font = Font(bold=True, size=11)
        cell.number_format = '"Rp "#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # TOTAL POD
        row = pod_end_row
        cell = ws.cell(row=row, column=6, value="TOTAL POD:")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        pod_total_str = self.pod_purchase_total_label.cget('text').replace('Rp', '').replace(',', '').strip()
        try:
            pod_total_val = float(pod_total_str)
        except:
            pod_total_val = 0
        
        cell = ws.cell(row=row, column=9, value=pod_total_val)
        cell.font = Font(bold=True, size=11)
        cell.number_format = '"Rp "#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # PURCHASE TOTAL
        row = max_row + 2
        cell = ws.cell(row=row, column=6, value="PURCHASE TOTAL:")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        purchase_total_str = self.purchase_total_label.cget('text').replace('Rp', '').replace(',', '').strip()
        try:
            purchase_total_val = float(purchase_total_str)
        except:
            purchase_total_val = 0
        
        cell = ws.cell(row=row, column=9, value=purchase_total_val)
        cell.font = Font(bold=True, size=12)
        cell.number_format = '"Rp "#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # SALES TOTAL
        row = max_row + 2
        cell = ws.cell(row=row, column=1, value="SALES TOTAL:")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        sales_total_str = self.sales_total_label.cget('text').replace('Rp', '').replace(',', '').strip()
        try:
            sales_total_val = float(sales_total_str)
        except:
            sales_total_val = 0
        
        cell = ws.cell(row=row, column=2, value=sales_total_val)
        cell.font = Font(bold=True, size=12)
        cell.number_format = '"Rp "#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # ==================== COLUMN WIDTHS ====================
        ws.column_dimensions['A'].width = 35  # Customer
        ws.column_dimensions['B'].width = 18  # Rp
        ws.column_dimensions['C'].width = 12  # Kubikasi
        ws.column_dimensions['D'].width = 12  # Tonase
        ws.column_dimensions['E'].width = 3   # Spacing
        ws.column_dimensions['F'].width = 25  # Keterangan Purchase
        ws.column_dimensions['G'].width = 10  # Unit
        ws.column_dimensions['H'].width = 15  # Biaya
        ws.column_dimensions['I'].width = 15  # Total
        
        # Set row heights
        for r in range(1, row + 1):
            if ws.row_dimensions[r].height is None:
                ws.row_dimensions[r].height = 18

    def print_preview(self):
        """Show print preview window"""
        if not self.selected_joa:
            messagebox.showwarning("Warning", "Pilih JOA terlebih dahulu!")
            return
        
        preview_window = tk.Toplevel(self.window)
        preview_window.title(f"Print Preview - JOA {self.selected_joa}")
        
        preview_width = int(self.window.winfo_width() * 0.7)
        preview_height = int(self.window.winfo_height() * 0.8)
        preview_window.geometry(f"{preview_width}x{preview_height}")
        preview_window.minsize(600, 400)
        
        text_frame = ttk.Frame(preview_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(text_frame, yscrollcommand=scrollbar.set, 
                             font=('Courier', 10), wrap=tk.WORD)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        joa_info = self.joa_data.get(self.selected_joa, {})
        
        content = f"""
{'='*80}
                    JOB ORDER ACCOUNT
                    PURCHASE INVOICE
{'='*80}

No. JOA        : {self.selected_joa}
Container      : {joa_info.get('container', '-')}
Feeder         : {joa_info.get('feeder', '-')}
Destination    : {joa_info.get('destination', '-')}
ETD            : {joa_info.get('etd_sub', '-')}
Tanggal Print  : {datetime.now().strftime('%d-%m-%Y %H:%M')}

{'-'*80}
{'KETERANGAN':<35} {'UNIT':<12} {'BIAYA PER':>15} {'TOTAL':>15}
{'-'*80}
"""
        
        total = 0
        for data in self.current_data:
            content += f"{data['keterangan']:<35} {data['unit']:<12} "
            content += f"Rp {data['biaya_per']:>12,.0f} Rp {data['total']:>12,.0f}\n"
            total += data['total']
        
        content += f"""{'-'*80}
{'TOTAL':>62} Rp {total:>12,.0f}
{'='*80}
"""
        
        text_widget.insert('1.0', content)
        text_widget.config(state=tk.DISABLED)
        
        btn_frame = ttk.Frame(preview_window)
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        ttk.Button(btn_frame, text="🖨️ Print", 
                  command=lambda: messagebox.showinfo("Info", "Fungsi print akan dihubungkan ke printer")).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Close", 
                  command=preview_window.destroy).pack(side=tk.RIGHT, padx=5)