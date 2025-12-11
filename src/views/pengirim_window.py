import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
import re
from src.models.database import AppDatabase
from PIL import Image, ImageTk
from src.utils.helpers import setup_window_restore_behavior

class SenderWindow:
    def __init__(self, parent, db, refresh_callback=None):
        self.parent = parent
        self.db = db
        self.refresh_callback = refresh_callback
        self.original_sender_data = []  # Store original data for filtering
        self.create_window()
    
    def create_window(self):
        """Create sender management window"""
        self.window = tk.Toplevel(self.parent)
        self.window.title("📋 Data Pengirim")

        # Calculate responsive window size
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()

        # Use 70% of screen width and 75% of screen height (better for 1366x768)
        window_width = min(int(screen_width * 0.70), 1000)
        window_height = min(int(screen_height * 0.75), 700)

        # Minimum size for usability
        window_width = max(window_width, 850)
        window_height = max(window_height, 500)

        self.window.geometry(f"{window_width}x{window_height}")
        self.window.configure(bg='#ecf0f1')
        self.window.transient(self.parent)
        self.window.grab_set()

        # Setup window restore behavior (fix minimize/restore issue)
        setup_window_restore_behavior(self.window)

        try:
            # Load dan resize image
            icon_image = Image.open("assets/logo.jpg")
            icon_image = icon_image.resize((32, 32), Image.Resampling.LANCZOS)
            icon_photo = ImageTk.PhotoImage(icon_image)

            # Set sebagai window icon
            self.window.iconphoto(False, icon_photo)

        except Exception as e:
            print(f"Icon tidak ditemukan: {e}")

        # Center window
        self.center_window()
        
        # Header
        header = tk.Label(
            self.window,
            text="📤 KELOLA DATA PENGIRIM",
            font=('Arial', 18, 'bold'),
            bg='#f7e30c',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Tab 1: Manual Input
        manual_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(manual_frame, text='✍️ Input Manual')
        self.create_manual_tab(manual_frame)
        
        # Tab 2: Excel Upload
        excel_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(excel_frame, text='📊 Upload Excel')
        self.create_excel_tab(excel_frame)
        
        # Tab 3: Sender List
        list_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(list_frame, text='📋 Daftar Pengirim')
        self.create_list_tab(list_frame)
        
        # Close button
        close_btn = tk.Button(
            self.window,
            text="❌ Tutup",
            font=('Arial', 12, 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=30,
            pady=10,
            command=self.window.destroy
        )
        close_btn.pack(pady=10)
    
    def create_manual_tab(self, parent):
        """Create manual input tab"""
        # Form frame
        form_frame = tk.Frame(parent, bg='#ecf0f1')
        form_frame.pack(fill='x', padx=20, pady=20)
        
        # Instructions
        instruction_label = tk.Label(
            form_frame,
            text="📝 Tambah Pengirim Satu per Satu",
            font=('Arial', 14, 'bold'),
            fg='#2c3e50',
            bg='#ecf0f1'
        )
        instruction_label.pack(pady=(0, 20))
        
        # Name field
        tk.Label(form_frame, text="Nama Pengirim:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.name_entry = tk.Entry(form_frame, font=('Arial', 12), width=50)
        self.name_entry.pack(fill='x', pady=(5, 10))
        
        # Buttons
        btn_frame = tk.Frame(form_frame, bg='#ecf0f1')
        btn_frame.pack(fill='x', pady=20)
        
        add_btn = tk.Button(
            btn_frame,
            text="➕ Tambah Pengirim",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.add_sender
        )
        add_btn.pack(side='left', padx=(0, 10))
        
        clear_btn = tk.Button(
            btn_frame,
            text="🗑️ Bersihkan",
            font=('Arial', 12, 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=10,
            pady=5,
            command=self.clear_form
        )
        clear_btn.pack(side='left')
        
        # Focus on name entry
        self.name_entry.focus()
    
    def create_excel_tab(self, parent):
        """Create Excel upload tab with enhanced error handling"""
        # Main container
        main_container = tk.Frame(parent, bg='#ecf0f1')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Instructions
        instruction_frame = tk.Frame(main_container, bg='#ffffff', relief='solid', bd=1)
        instruction_frame.pack(fill='x', pady=(0, 20))
        
        instruction_title = tk.Label(
            instruction_frame,
            text="📊 Upload Data Pengirim dari Excel",
            font=('Arial', 14, 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        instruction_title.pack(pady=(10, 5))
        
        instruction_text = tk.Label(
            instruction_frame,
            text="Format Excel yang dibutuhkan:\n\n" +
                 "• Kolom A: Nama Pengirim (WAJIB)\n" +
                 "• Baris pertama adalah header (Nama)\n" +
                 "• Pastikan tidak ada duplikasi nama pengirim\n\n" +
                 "Tips: Download template untuk format yang benar!",
            font=('Arial', 10),
            fg='#34495e',
            bg='#ffffff',
            justify='left'
        )
        instruction_text.pack(pady=(0, 10), padx=20)
        
        # File selection
        file_frame = tk.Frame(main_container, bg='#ecf0f1')
        file_frame.pack(fill='x', pady=10)
        
        tk.Label(file_frame, text="Pilih File Excel:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        
        file_input_frame = tk.Frame(file_frame, bg='#ecf0f1')
        file_input_frame.pack(fill='x', pady=(5, 0))
        
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(file_input_frame, textvariable=self.file_path_var, font=('Arial', 11), state='readonly')
        self.file_entry.pack(side='left', fill='x', expand=True, ipady=5)
        
        browse_btn = tk.Button(
            file_input_frame,
            text="📁 Browse",
            font=('Arial', 10, 'bold'),
            bg='#3498db',
            fg='white',
            padx=15,
            pady=5,
            command=self.browse_file
        )
        browse_btn.pack(side='right', padx=(5, 0))
        
        # Preview area
        preview_frame = tk.Frame(main_container, bg='#ecf0f1')
        preview_frame.pack(fill='both', expand=True, pady=10)
        
        tk.Label(preview_frame, text="📋 Preview Data:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        
        # Preview treeview with scrollbars
        preview_tree_frame = tk.Frame(preview_frame, bg='#ecf0f1')
        preview_tree_frame.pack(fill='x', pady=5)
        
        preview_container = tk.Frame(preview_tree_frame, bg='#ecf0f1')
        preview_container.pack(fill='both', expand=True)
        
        self.preview_tree = ttk.Treeview(preview_container, 
                                       columns=('Status', 'Nama'), 
                                       show='headings', height=8)
        
        self.preview_tree.heading('Status', text='Status')
        self.preview_tree.heading('Nama', text='Nama Pengirim')
        
        self.preview_tree.column('Status', width=80)
        self.preview_tree.column('Nama', width=400)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(preview_container, orient='vertical', command=self.preview_tree.yview)
        h_scrollbar = ttk.Scrollbar(preview_container, orient='horizontal', command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)
        
        # Upload buttons
        upload_btn_frame = tk.Frame(main_container, bg='#ecf0f1')
        upload_btn_frame.pack(fill='x', pady=15)
        
        self.upload_btn = tk.Button(
            upload_btn_frame,
            text="⬆️ Upload ke Database",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.upload_excel_data,
            state='disabled'
        )
        self.upload_btn.pack(side='left', padx=(0, 15))
        
        download_template_btn = tk.Button(
            upload_btn_frame,
            text="📥 Download Template",
            font=('Arial', 12, 'bold'),
            bg='#f39c12',
            fg='white',
            padx=10,
            pady=5,
            command=self.download_template
        )
        download_template_btn.pack(side='left')
        
        # Status label
        self.status_label = tk.Label(
            main_container,
            text="",
            font=('Arial', 11),
            fg='#e74c3c',
            bg='#ecf0f1',
            wraplength=1000,
            justify='left'
        )
        self.status_label.pack(pady=10, fill='x')
    
    def create_list_tab(self, parent):
        """Create sender list tab with search, update, delete functionality"""
        # Container
        list_container = tk.Frame(parent, bg='#ecf0f1')
        list_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Header
        header_frame = tk.Frame(list_container, bg='#ecf0f1')
        header_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(header_frame, text="📋 DAFTAR PENGIRIM", font=('Arial', 14, 'bold'), bg='#ecf0f1').pack(side='left')
        
        refresh_btn = tk.Button(
            header_frame,
            text="🔄 Refresh",
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=15,
            pady=5,
            command=self.load_senders
        )
        refresh_btn.pack(side='right')
        
        # Search/Filter Frame
        search_frame = tk.Frame(list_container, bg='#ffffff', relief='solid', bd=1)
        search_frame.pack(fill='x', pady=(0, 10))
        
        # Search label
        search_label = tk.Label(
            search_frame,
            text="🔍 Pencarian Pengirim:",
            font=('Arial', 12, 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        search_label.pack(anchor='w', padx=10, pady=(10, 5))
        
        # Search controls frame
        search_controls = tk.Frame(search_frame, bg='#ffffff')
        search_controls.pack(fill='x', padx=10, pady=(0, 10))
        
        # Search by name
        tk.Label(search_controls, text="Nama:", font=('Arial', 10), bg='#ffffff').pack(side='left')
        self.search_name_var = tk.StringVar()
        self.search_name_var.trace('w', self.on_search_change)
        search_name_entry = tk.Entry(search_controls, textvariable=self.search_name_var, font=('Arial', 10), width=40)
        search_name_entry.pack(side='left', padx=(5, 20))
        
        # Clear search button
        clear_search_btn = tk.Button(
            search_controls,
            text="❌ Clear",
            font=('Arial', 9),
            bg='#e67e22',
            fg='white',
            padx=10,
            pady=2,
            command=self.clear_search
        )
        clear_search_btn.pack(side='left', padx=5)
        
        # Action buttons frame
        action_frame = tk.Frame(list_container, bg='#ecf0f1')
        action_frame.pack(fill='x', pady=(0, 10))
        
        # Update button
        update_btn = tk.Button(
            action_frame,
            text="✏️ Edit Pengirim",
            font=('Arial', 11, 'bold'),
            bg='#3498db',
            fg='white',
            padx= 10,
            pady=5,
            command=self.update_sender
        )
        update_btn.pack(side='left', padx=(0, 10))
        
        # Delete button
        delete_btn = tk.Button(
            action_frame,
            text="🗑️ Hapus Pengirim",
            font=('Arial', 11, 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=10,
            pady=5,
            command=self.delete_sender
        )
        delete_btn.pack(side='left', padx=(0, 10))
        
        # Export button
        export_btn = tk.Button(
            action_frame,
            text="📤 Export Excel",
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.export_senders
        )
        export_btn.pack(side='left')
        
        # Info label
        self.info_label = tk.Label(
            action_frame,
            text="💡 Pilih pengirim dari tabel untuk edit/hapus",
            font=('Arial', 10),
            fg='#7f8c8d',
            bg='#ecf0f1'
        )
        self.info_label.pack(side='right')
        
        # Treeview for sender list with scrollbars
        tree_frame = tk.Frame(list_container, bg='#ecf0f1')
        tree_frame.pack(fill='both', expand=True)
        
        tree_container = tk.Frame(tree_frame, bg='#ecf0f1')
        tree_container.pack(fill='both', expand=True)
        
        self.tree = ttk.Treeview(tree_container,
                               columns=('ID', 'Nama', 'Created'),
                               show='headings', height=12)
        
        # Configure columns
        self.tree.heading('ID', text='ID')
        self.tree.heading('Nama', text='Nama Pengirim')
        self.tree.heading('Created', text='Tanggal Dibuat')
        
        self.tree.column('ID', width=80)
        self.tree.column('Nama', width=400)
        self.tree.column('Created', width=150)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient='vertical', command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Bind double-click to edit
        self.tree.bind('<Double-1>', lambda e: self.update_sender())
        
        # Bind selection change to update info
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
        # Load existing senders
        self.load_senders()
        
        # Add tab change event to refresh data when switching to this tab
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
    
    def on_search_change(self, *args):
        """Handle search input changes"""
        self.filter_senders()
    
    def clear_search(self):
        """Clear all search filters"""
        self.search_name_var.set('')
        self.filter_senders()
    
    def filter_senders(self):
        """Filter senders based on search criteria"""
        # Clear current display
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Get search criteria
        search_name = self.search_name_var.get().strip()
        
        # Filter data
        filtered_data = []
        
        for sender in self.original_sender_data:
            # Check name filter
            if search_name:
                nama_pengirim = sender['nama_pengirim'].lower()
                search_terms = search_name.lower().split()
                
                # Check if any search term matches
                match_found = False
                for term in search_terms:
                    # Create regex pattern for flexible matching
                    pattern = '.*'.join(re.escape(char) for char in term)
                    if re.search(pattern, nama_pengirim, re.IGNORECASE):
                        match_found = True
                        break
                    
                    # Also check direct substring match
                    if term in nama_pengirim:
                        match_found = True
                        break
                
                if not match_found:
                    continue
            
            filtered_data.append(sender)
        
        # Display filtered data
        for sender in filtered_data:
            # Format date
            created_date = sender['created_at'][:10] if sender['created_at'] else '-'
            
            self.tree.insert('', tk.END, values=(
                sender['pengirim_id'],
                sender['nama_pengirim'],
                created_date
            ))
        
        # Update info label
        total_count = len(self.original_sender_data)
        filtered_count = len(filtered_data)
        if total_count != filtered_count:
            self.info_label.config(text=f"📊 Menampilkan {filtered_count} dari {total_count} pengirim")
        else:
            self.info_label.config(text="💡 Pilih pengirim dari tabel untuk edit/hapus")
    
    def on_tree_select(self, event):
        """Handle tree selection change"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            sender_id = item['values'][0]
            nama_pengirim = item['values'][1]
            self.info_label.config(text=f"✅ Terpilih: {nama_pengirim} (ID: {sender_id})")
        else:
            self.info_label.config(text="💡 Pilih pengirim dari tabel untuk edit/hapus")
    
    def update_sender(self):
        """Update selected sender"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih pengirim yang akan diedit dari tabel!")
            return
        
        # Get selected item data
        item = self.tree.item(selection[0])
        sender_id = item['values'][0]
        
        # Find full sender data
        selected_sender = None
        for sender in self.original_sender_data:
            if sender['sender_id'] == sender_id:
                selected_sender = sender
                break
        
        if not selected_sender:
            messagebox.showerror("Error", "Data pengirim tidak ditemukan!")
            return
        
        # Open update dialog
        self.open_update_dialog(selected_sender)
    
    def open_update_dialog(self, sender_data):
        """Open dialog to update sender data"""
        # Create update window
        update_window = tk.Toplevel(self.window)
        update_window.title(f"✏️ Edit Pengirim - {sender_data['nama_pengirim']}")
        update_window.geometry("500x300")  
        update_window.configure(bg='#ecf0f1')
        update_window.transient(self.window)
        update_window.grab_set()
        
        # Center window
        update_window.update_idletasks()
        x = (update_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (update_window.winfo_screenheight() // 2) - (300 // 2) 
        update_window.geometry(f"500x300+{x}+{y}")
        
        # Header
        header = tk.Label(
            update_window,
            text="✏️ EDIT DATA PENGIRIM",
            font=('Arial', 16, 'bold'),
            bg='#f7e30c',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Form frame
        form_frame = tk.Frame(update_window, bg='#ecf0f1')
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Sender ID (read-only)
        tk.Label(form_frame, text="ID Pengirim:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        id_label = tk.Label(
            form_frame, 
            text=str(sender_data['sender_id']), 
            font=('Arial', 11),
            bg='#ffffff',
            relief='solid',
            bd=1,
            padx=5,
            pady=5
        )
        id_label.pack(fill='x', pady=(5, 10))
        
        # Sender Name
        tk.Label(form_frame, text="Nama Pengirim:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        nama_var = tk.StringVar(value=sender_data['nama_pengirim'])
        nama_entry = tk.Entry(form_frame, textvariable=nama_var, font=('Arial', 11))
        nama_entry.pack(fill='x', pady=(5, 20))
        
        def on_save():
            """Save updated sender data"""
            try:
                new_nama = nama_var.get().strip()
                
                if not new_nama:
                    messagebox.showerror("Error", "Nama pengirim harus diisi!")
                    nama_entry.focus()
                    return
                
                # Update sender in database
                self.db.update_sender(
                    sender_id=sender_data['sender_id'],
                    nama_pengirim=new_nama
                )
                
                messagebox.showinfo("Sukses", "Data pengirim berhasil diupdate!")
                
                # Refresh data and close dialog
                self.load_senders()
                if self.refresh_callback:
                    self.refresh_callback()
                
                update_window.destroy()
                
            except ValueError as ve:
                messagebox.showerror("Error Validasi", f"Data tidak valid!\nError: {str(ve)}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal mengupdate pengirim!\nError: {str(e)}")
        
        # Buttons
        btn_frame = tk.Frame(update_window, bg='#ecf0f1')  
        btn_frame.pack(fill='x', padx=20, pady=10, side='bottom')  
        
        save_btn = tk.Button(
            btn_frame,
            text="💾 Simpan",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=20,
            pady=8,
            command=on_save
        )
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(
            btn_frame,
            text="❌ Batal",
            font=('Arial', 12, 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=20,
            pady=8,
            command=update_window.destroy
        )
        cancel_btn.pack(side='right')
        
        # Focus on name entry
        nama_entry.focus()
        nama_entry.select_range(0, tk.END)
    
    def delete_sender(self):
        """Delete selected sender"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih pengirim yang akan dihapus dari tabel!")
            return
        
        # Get selected item data
        item = self.tree.item(selection[0])
        sender_id = item['values'][0]
        nama_pengirim = item['values'][1]
        
        # Check if sender has associated barang
        try:
            barang_count = self.db.execute_one(
                "SELECT COUNT(*) as count FROM barang WHERE sender_id = ?",
                (sender_id,)
            )
            
            has_barang = barang_count['count'] > 0 if barang_count else False
            
            # Confirm deletion
            confirm_msg = f"Yakin ingin menghapus pengirim?\n\n" + \
                         f"ID: {sender_id}\n" + \
                         f"Nama: {nama_pengirim}\n\n"
            
            if has_barang:
                confirm_msg += f"⚠️ PERINGATAN: Pengirim ini memiliki {barang_count['count']} barang terkait!\n" + \
                              f"Menghapus pengirim akan mempengaruhi data barang terkait.\n\n"
            
            confirm_msg += f"⚠️ Aksi ini tidak dapat dibatalkan!"
            
            if not messagebox.askyesno("Konfirmasi Hapus", confirm_msg):
                return
            
            # Delete sender
            self.db.execute("DELETE FROM senders WHERE sender_id = ?", (sender_id,))
            
            messagebox.showinfo("Sukses", f"Pengirim '{nama_pengirim}' berhasil dihapus!")
            
            # Refresh data
            self.load_senders()
            if self.refresh_callback:
                self.refresh_callback()
                
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menghapus pengirim:\n{str(e)}")
    
    def export_senders(self):
        """Export sender data to Excel"""
        try:
            if not self.original_sender_data:
                messagebox.showwarning("Peringatan", "Tidak ada data pengirim untuk diekspor!")
                return
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                parent=self.window,
                title="Export Data Pengirim ke Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"data_pengirim_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Prepare data for export
            export_data = []
            for sender in self.original_sender_data:
                export_data.append({
                    'ID': sender['sender_id'],
                    'Nama Pengirim': sender['nama_pengirim'],
                    'Tanggal Dibuat': sender.get('created_at', '')
                })
            
            # Create DataFrame and export
            df = pd.DataFrame(export_data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data Pengirim')
                
                # Format the Excel file
                workbook = writer.book
                worksheet = writer.sheets['Data Pengirim']
                
                # Style headers
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="F7E30C", end_color="F7E30C", fill_type="solid")

                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo(
                "Export Berhasil",
                f"Data pengirim berhasil diekspor ke:\n{filename}\n\n" +
                f"📊 Total: {len(export_data)} pengirim"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export data:\n{str(e)}")
    
    def show_error_details(self, errors, duplicate_list, success_count, total_count):
        """Show detailed error modal with specific error information"""
        error_window = tk.Toplevel(self.window)
        error_window.title("📊 Detail Error Upload")
        error_window.geometry("800x600")
        error_window.configure(bg='#ecf0f1')
        error_window.transient(self.window)
        error_window.grab_set()
        
        # Center the error window
        error_window.update_idletasks()
        x = (error_window.winfo_screenwidth() // 2) - (800 // 2)
        y = (error_window.winfo_screenheight() // 2) - (600 // 2)
        error_window.geometry(f"800x600+{x}+{y}")
        
        # Header
        header = tk.Label(
            error_window,
            text="📊 DETAIL HASIL UPLOAD",
            font=('Arial', 16, 'bold'),
            bg='#e74c3c',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Summary frame
        summary_frame = tk.Frame(error_window, bg='#ffffff', relief='solid', bd=1)
        summary_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        summary_text = f"""📈 RINGKASAN UPLOAD:
        
✅ Berhasil: {success_count} dari {total_count} pengirim
❌ Gagal: {len(errors)} pengirim  
⚠️ Duplikat dilewati: {len(duplicate_list)} pengirim
        
📋 Total diproses: {total_count} baris data"""
        
        summary_label = tk.Label(
            summary_frame,
            text=summary_text,
            font=('Arial', 12),
            fg='#2c3e50',
            bg='#ffffff',
            justify='left',
            padx=20,
            pady=15
        )
        summary_label.pack(fill='x')
        
        # Create notebook for error details
        notebook = ttk.Notebook(error_window)
        notebook.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Tab 1: Processing Errors
        if errors:
            error_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(error_frame, text=f'❌ Error Processing ({len(errors)})')
            
            error_label = tk.Label(
                error_frame,
                text="Pengirim yang gagal diproses karena error teknis:",
                font=('Arial', 12, 'bold'),
                fg='#e74c3c',
                bg='#ecf0f1'
            )
            error_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            # Error treeview
            error_main_frame = tk.Frame(error_frame, bg='#ecf0f1')
            error_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            error_tree = ttk.Treeview(error_main_frame, 
                                    columns=('No', 'Nama Pengirim', 'Error'), 
                                    show='headings', height=12)
            
            error_tree.heading('No', text='No')
            error_tree.heading('Nama Pengirim', text='Nama Pengirim')
            error_tree.heading('Error', text='Detail Error')
            
            error_tree.column('No', width=50)
            error_tree.column('Nama Pengirim', width=300)
            error_tree.column('Error', width=400)
            
            error_v_scroll = ttk.Scrollbar(error_main_frame, orient='vertical', command=error_tree.yview)
            error_tree.configure(yscrollcommand=error_v_scroll.set)
            
            error_tree.pack(side='left', fill='both', expand=True)
            error_v_scroll.pack(side='right', fill='y')
            
            # Add error data
            for i, error_info in enumerate(errors, 1):
                error_tree.insert('', tk.END, values=(
                    i,
                    error_info.get('nama_pengirim', 'N/A'),
                    error_info.get('error', 'Unknown error')
                ))
        
        # Tab 2: Duplicate Records
        if duplicate_list:
            duplicate_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(duplicate_frame, text=f'⚠️ Duplikat Dilewati ({len(duplicate_list)})')
            
            duplicate_label = tk.Label(
                duplicate_frame,
                text="Pengirim yang sudah ada di database (dilewati):",
                font=('Arial', 12, 'bold'),
                fg='#f39c12',
                bg='#ecf0f1'
            )
            duplicate_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            duplicate_main_frame = tk.Frame(duplicate_frame, bg='#ecf0f1')
            duplicate_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            duplicate_tree = ttk.Treeview(duplicate_main_frame, 
                                        columns=('No', 'Nama Pengirim'), 
                                        show='headings', height=12)
            
            duplicate_tree.heading('No', text='No')
            duplicate_tree.heading('Nama Pengirim', text='Nama Pengirim')
            
            duplicate_tree.column('No', width=50)
            duplicate_tree.column('Nama Pengirim', width=400)
            
            duplicate_v_scroll = ttk.Scrollbar(duplicate_main_frame, orient='vertical', command=duplicate_tree.yview)
            duplicate_tree.configure(yscrollcommand=duplicate_v_scroll.set)
            
            duplicate_tree.pack(side='left', fill='both', expand=True)
            duplicate_v_scroll.pack(side='right', fill='y')
            
            # Add duplicate data
            for i, duplicate_info in enumerate(duplicate_list, 1):
                duplicate_tree.insert('', tk.END, values=(
                    i,
                    duplicate_info.get('nama_pengirim', 'N/A')
                ))
        
        # Tab 3: Upload Tips
        tips_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(tips_frame, text='💡 Tips Upload')
        
        tips_canvas = tk.Canvas(tips_frame, bg='#ecf0f1')
        tips_scrollbar = ttk.Scrollbar(tips_frame, orient="vertical", command=tips_canvas.yview)
        tips_scrollable_frame = tk.Frame(tips_canvas, bg='#ecf0f1')
        
        tips_scrollable_frame.bind(
            "<Configure>",
            lambda e: tips_canvas.configure(scrollregion=tips_canvas.bbox("all"))
        )
        
        tips_canvas.create_window((0, 0), window=tips_scrollable_frame, anchor="nw")
        tips_canvas.configure(yscrollcommand=tips_scrollbar.set)
        
        tips_text = """📋 TIPS UNTUK UPLOAD PENGIRIM YANG SUKSES:

1. 🔍 Format Excel yang Benar:
   • Gunakan template yang sudah disediakan
   • Kolom 'Nama' wajib diisi
   • Tidak perlu kolom alamat untuk pengirim

2. 📊 Data Quality:
   • Pastikan nama pengirim tidak kosong
   • Hindari duplikasi nama pengirim
   • Gunakan format teks yang konsisten

3. 🚨 Mengatasi Error:
   • Perbaiki baris yang error sesuai detail di tab Error
   • Pastikan tidak ada karakter khusus yang merusak format
   • Upload ulang file yang sudah diperbaiki

4. 💾 Backup Data:
   • Simpan file Excel asli sebagai backup
   • Export data yang sudah ada sebelum upload besar

5. 🔧 Troubleshooting Common Errors:
   • "Nama pengirim harus diisi" → Pastikan kolom nama tidak kosong
   • "Database error" → Periksa koneksi database
   • "File tidak bisa dibaca" → Pastikan file Excel tidak sedang dibuka

6. 📝 Best Practices:
   • Gunakan nama pengirim yang unik dan deskriptif
   • Periksa data di preview sebelum upload
   • Upload dalam batch kecil untuk file besar"""
        
        tips_label = tk.Label(
            tips_scrollable_frame,
            text=tips_text,
            font=('Arial', 11),
            fg='#2c3e50',
            bg='#ecf0f1',
            justify='left',
            padx=20,
            pady=20,
            wraplength=700
        )
        tips_label.pack(fill='both', expand=True)
        
        tips_canvas.pack(side="left", fill="both", expand=True)
        tips_scrollbar.pack(side="right", fill="y")
        
        # Close button
        close_btn = tk.Button(
            error_window,
            text="✅ Tutup",
            font=('Arial', 14, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=40,
            pady=15,
            command=error_window.destroy
        )
        close_btn.pack(pady=20)
    
    def center_window(self):
        """Center window on parent"""
        self.window.update_idletasks()
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        x = parent_x + (parent_width // 2) - (1000 // 2)
        y = parent_y + (parent_height // 2) - (800 // 2)
        
        self.window.geometry(f"1000x800+{x}+{y}")
    
    def browse_file(self):
        """Browse for Excel file"""
        file_types = [
            ('Excel files', '*.xlsx *.xls'),
            ('All files', '*.*')
        ]
        
        filename = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=file_types,
            parent=self.window
        )
        
        if filename:
            self.file_path_var.set(filename)
            self.preview_excel_file(filename)
    
    def preview_excel_file(self, filename):
        """Preview Excel file content with enhanced validation"""
        try:
            self.status_label.config(text="📄 Membaca file Excel...", fg='#3498db')
            
            # Clear previous preview
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            # Check file exists
            if not os.path.exists(filename):
                self.status_label.config(
                    text="❌ File tidak ditemukan!", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            # Try to read Excel file
            df = None
            try:
                df = pd.read_excel(filename, engine='openpyxl')
            except:
                try:
                    df = pd.read_excel(filename, engine='xlrd')
                except Exception as e:
                    self.status_label.config(
                        text=f"❌ Error membaca Excel: {str(e)}\n\nPastikan file Excel tidak sedang dibuka!", 
                        fg='#e74c3c'
                    )
                    self.upload_btn.config(state='disabled')
                    return
            
            if df is None or df.empty:
                self.status_label.config(
                    text="❌ File Excel kosong atau tidak valid!", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            
            # Find nama column (case insensitive)
            nama_col = None
            
            for col in df.columns:
                col_lower = col.lower().strip()
                if col_lower in ['nama', 'name', 'nama pengirim', 'pengirim', 'sender']:
                    nama_col = col
                    break
            
            if nama_col is None:
                available_cols = ', '.join(df.columns.tolist())
                self.status_label.config(
                    text=f"❌ Kolom 'Nama' tidak ditemukan!\n\nKolom tersedia: {available_cols}", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            # Get existing senders for duplicate check
            existing_senders = {s['nama_pengirim'].upper() for s in self.db.get_all_senders()}
            
            # Filter and validate data
            valid_rows = df[df[nama_col].notna() & (df[nama_col].astype(str).str.strip() != '')]
            
            if len(valid_rows) == 0:
                self.status_label.config(
                    text="❌ Tidak ada data valid! Pastikan kolom 'Nama' tidak kosong.", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            # Preview data with status
            preview_data = valid_rows.head(100)
            row_count = len(valid_rows)
            duplicate_count = 0
            
            for _, row in preview_data.iterrows():
                nama = str(row[nama_col]).strip() if pd.notna(row[nama_col]) else ''
                
                if nama:
                    # Check if sender exists
                    status = "⚠️" if nama.upper() in existing_senders else "✅"
                    if nama.upper() in existing_senders:
                        duplicate_count += 1
                    
                    self.preview_tree.insert('', tk.END, values=(status, nama))
            
            # Store column mapping for upload
            self.nama_column = nama_col
            
            # Create status message
            status_msg = f"✅ File berhasil dibaca: {row_count} baris data valid\n"
            status_msg += f"📋 Kolom: {nama_col}\n"
            if duplicate_count > 0:
                status_msg += f"⚠️ {duplicate_count} pengirim sudah ada (akan dilewati)\n"
            status_msg += f"💡 ✅ = Baru, ⚠️ = Sudah ada"
            
            self.status_label.config(text=status_msg, fg='#27ae60')
            self.upload_btn.config(state='normal')
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Preview error: {error_detail}")
            
            self.status_label.config(
                text=f"❌ Error membaca file: {str(e)}", 
                fg='#e74c3c'
            )
            self.upload_btn.config(state='disabled')
    
    def upload_excel_data(self):
        """Upload Excel data to database with enhanced error handling"""
        filename = self.file_path_var.get()
        if not filename:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu!")
            return
        
        try:
            print(f"🔄 Starting enhanced sender upload from file: {filename}")
            
            # Read Excel file
            df = pd.read_excel(filename)
            df.columns = df.columns.astype(str).str.strip()
            
            # Use stored column mapping
            nama_col = getattr(self, 'nama_column', 'Nama')
            
            # Validate data
            if nama_col not in df.columns:
                messagebox.showerror("Error", f"Kolom '{nama_col}' tidak ditemukan!")
                return
            
            # Filter valid data
            valid_rows = df[df[nama_col].notna() & (df[nama_col].astype(str).str.strip() != '')]
            
            if len(valid_rows) == 0:
                messagebox.showerror("Error", "Tidak ada data valid untuk diupload!")
                return
            
            # Confirm upload
            if not messagebox.askyesno(
                "Konfirmasi Upload", 
                f"Upload {len(valid_rows)} pengirim ke database?\n\n" +
                f"Data yang sudah ada tidak akan terduplikasi."
            ):
                return
            
            # Get existing senders for duplicate check
            existing_senders = {s['nama_pengirim'].upper(): s['pengirim_id'] for s in self.db.get_all_senders()}
            
            # Initialize tracking variables
            success_count = 0
            errors = []
            duplicate_list = []
            
            for idx, (_, row) in enumerate(valid_rows.iterrows()):
                try:
                    nama = str(row[nama_col]).strip()
                    
                    if nama:
                        # Check if sender already exists
                        if nama.upper() in existing_senders:
                            duplicate_list.append({
                                'nama_pengirim': nama
                            })
                            continue
                        
                        # Create new sender
                        sender_id = self.db.create_sender(nama)
                        success_count += 1
                        
                        # Add to existing senders to prevent duplicates in this batch
                        existing_senders[nama.upper()] = sender_id
                
                except Exception as e:
                    error_detail = str(e)
                    errors.append({
                        'nama_pengirim': nama,
                        'error': error_detail,
                        'row_number': idx + 2
                    })
            
            # Show results
            total_processed = len(valid_rows)
            
            if errors or duplicate_list:
                self.show_error_details(errors, duplicate_list, success_count, total_processed)
            else:
                messagebox.showinfo(
                    "Upload Berhasil! 🎉", 
                    f"Semua data berhasil diupload!\n\n" +
                    f"✅ Total berhasil: {success_count} pengirim\n" +
                    f"📊 Total diproses: {total_processed} baris data"
                )
            
            # Refresh display and switch to list tab
            self.load_senders()
            if self.refresh_callback:
                self.refresh_callback()
            
            # Switch to sender list tab
            self.notebook.select(2)
            
            # Clear file selection
            self.file_path_var.set("")
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            self.upload_btn.config(state='disabled')
            self.status_label.config(text="")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Upload error: {error_detail}")
            messagebox.showerror("Error", f"Gagal upload data:\n{str(e)}")
    
    def download_template(self):
        """Download Excel template with sample data"""
        try:
            print("📥 Creating sender template...")
            
            # Create sample data
            template_data = {
                'Nama': [
                    'PT. LOGISTIK JAKARTA',
                    'CV. EXPEDISI MANDIRI',
                    'PT. TRANSPORT NUSANTARA',
                    'UD. KIRIM CEPAT',
                    'PT. CARGO EXPRESS'
                ]
            }
            
            df = pd.DataFrame(template_data)
            print("✅ Template data created")
            
            # Save file
            filename = filedialog.asksaveasfilename(
                parent=self.window,
                title="Simpan Template Excel Pengirim",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="template_pengirim.xlsx"
            )
            
            if filename:
                print(f"💾 Saving template to: {filename}")
                
                # Create Excel with styling and instructions
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Main data sheet
                    df.to_excel(writer, index=False, sheet_name='Data Pengirim')
                    
                    # Create instructions sheet
                    instructions_data = {
                        'Kolom': [
                            'Nama'
                        ],
                        'Keterangan': [
                            'Nama pengirim yang unik dan lengkap (WAJIB)'
                        ],
                        'Contoh': [
                            'PT. LOGISTIK JAKARTA'
                        ]
                    }
                    
                    instructions_df = pd.DataFrame(instructions_data)
                    instructions_df.to_excel(writer, index=False, sheet_name='Petunjuk')
                    
                    # Style the sheets
                    workbook = writer.book
                    data_worksheet = writer.sheets['Data Pengirim']
                    instructions_worksheet = writer.sheets['Petunjuk']
                    
                    # Import styling
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Header styling
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="F7E30C", end_color="F7E30C", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Border
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Style both sheets
                    for worksheet, df_data in [(data_worksheet, df), (instructions_worksheet, instructions_df)]:
                        # Style headers
                        for col_num, column_title in enumerate(df_data.columns, 1):
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            cell.border = thin_border
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 3, 60)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Add borders to data cells
                        for row in worksheet.iter_rows(min_row=2, max_row=len(df_data)+1):
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical="center")
                
                print("✅ Template saved successfully")
                messagebox.showinfo(
                    "Sukses", 
                    f"Template berhasil disimpan:\n{filename}\n\n" +
                    "Template berisi:\n" +
                    "• Sheet 'Data Pengirim' - contoh data pengirim\n" +
                    "• Sheet 'Petunjuk' - penjelasan format\n" +
                    "• Contoh nama pengirim yang lengkap\n\n" +
                    "Pastikan nama pengirim unik untuk menghindari duplikasi!"
                )
            else:
                print("❌ Save cancelled by user")
        
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Template download error: {error_detail}")
            
            messagebox.showerror("Error", f"Gagal membuat template:\n\nError: {str(e)}")
    
    def add_sender(self):
        """Add new sender manually with error handling"""
        name = self.name_entry.get().strip()
        
        if not name:
            messagebox.showerror("Error", "Nama pengirim harus diisi!")
            self.name_entry.focus()
            return
        
        try:
            sender_id = self.db.create_sender(name)
            messagebox.showinfo("Sukses", f"Pengirim berhasil ditambahkan dengan ID: {sender_id}")
            self.clear_form()
            
            # Refresh and switch to list tab
            self.load_senders()
            if self.refresh_callback:
                self.refresh_callback()
            
            self.notebook.select(2)
            
        except ValueError as ve:
            messagebox.showerror("Error Validasi", f"Data tidak valid!\nError: {str(ve)}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menambahkan pengirim:\n{str(e)}")
    
    def clear_form(self):
        """Clear form fields"""
        self.name_entry.delete(0, tk.END)
        self.name_entry.focus()
    
    def load_senders(self):
        """Load senders into treeview with error handling"""
        try:
            print("🔄 Loading senders from database...")
            
            # Clear existing items
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Load senders from database
            senders = self.db.get_all_senders()
            self.original_sender_data = senders  # Store original data for filtering
            print(f"📊 Found {len(senders)} senders in database")
            
            for sender in senders:
                try:
                    # Format date
                    created_date = sender['created_at'][:10] if sender['created_at'] else '-'
                except KeyError:
                    created_date = '-'
                
                self.tree.insert('', tk.END, values=(
                    sender['pengirim_id'],
                    sender['nama_pengirim'],
                    created_date
                ))
            
            print("✅ Sender list loaded successfully")
            
        except Exception as e:
            print(f"💥 Error loading senders: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal memuat daftar pengirim: {str(e)}")
    
    def on_tab_changed(self, event):
        """Handle tab change event"""
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        
        # If switching to sender list tab, refresh the data
        if "Daftar Pengirim" in tab_text:
            self.load_senders()
            print("Tab changed to Sender List - data refreshed")

# Example usage and testing functions
def test_sender_window():
    """Test function for SenderWindow"""
    try:
        # Create test database
        db = AppDatabase("test_sender.db")
        
        # Create root window
        root = tk.Tk()
        root.title("Test Sender Window")
        root.geometry("400x300")
        
        def open_sender_window():
            SenderWindow(root, db)
        
        # Test button
        test_btn = tk.Button(
            root,
            text="Open Sender Window",
            font=('Arial', 14),
            command=open_sender_window,
            padx=20,
            pady=10
        )
        test_btn.pack(expand=True)
        
        root.mainloop()
        
    except Exception as e:
        print(f"Test error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_sender_window()