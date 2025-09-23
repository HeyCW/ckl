import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
from src.models.database import AppDatabase
import re
from PIL import Image, ImageTk

from src.widget.paginated_tree_view import PaginatedTreeView

class BarangWindow:
    def __init__(self, parent, db, refresh_callback=None):
        self.parent = parent
        self.db = db
        self.refresh_callback = refresh_callback
        self.create_window()
    
    def create_window(self):
        """Create barang management window"""
        self.window = tk.Toplevel(self.parent)
        self.window.title("📦 Data Barang")
        self.window.geometry("1200x800")
        self.window.configure(bg='#ecf0f1')
        self.window.transient(self.parent)
        self.window.grab_set()
        
        try:
            # Load dan resize image
            icon_image = Image.open("assets/logo.jpg")
            icon_image = icon_image.resize((32, 32), Image.Resampling.LANCZOS)
            icon_photo = ImageTk.PhotoImage(icon_image)
            
            # Set sebagai window icon
            self.window.iconphoto(False, icon_photo)
            
        except Exception as e:
            print(f"Icon tidak ditemukan: {e}")
        
        # Center window
        self.center_window()
        
        # Header
        header = tk.Label(
            self.window,
            text="📦 KELOLA DATA BARANG",
            font=('Arial', 18, 'bold'),
            bg='#27ae60',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Tab 1: Manual Input
        manual_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(manual_frame, text='✍️ Input Manual')
        self.create_manual_tab(manual_frame)
        
        # Tab 2: Excel Upload
        excel_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(excel_frame, text='📊 Upload Excel')
        self.create_excel_tab(excel_frame)
        
        # Tab 3: Barang List
        list_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(list_frame, text='📋 Daftar Barang')
        self.create_list_tab(list_frame)
        
        # Close button
        close_btn = tk.Button(
            self.window,
            text="❌ Tutup",
            font=('Arial', 12, 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=10,
            pady=5,
            command=self.window.destroy
        )
        close_btn.pack(pady=10)
    
    def create_manual_tab(self, parent):
        """Create manual input tab with scrollable content"""
        # Main container
        main_container = tk.Frame(parent, bg='#ecf0f1')
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(main_container, bg='#ecf0f1', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
        
        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mouse wheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_from_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind('<Enter>', _bind_to_mousewheel)
        canvas.bind('<Leave>', _unbind_from_mousewheel)
        
        # Form frame (now inside scrollable_frame)
        form_frame = tk.Frame(scrollable_frame, bg='#ecf0f1')
        form_frame.pack(fill='x', padx=20, pady=20)
        
        # Instructions
        instruction_label = tk.Label(
            form_frame,
            text="📝 Tambah Barang Satu per Satu",
            font=('Arial', 14, 'bold'),
            fg='#2c3e50',
            bg='#ecf0f1'
        )
        instruction_label.pack(pady=(0, 20))
        
        # Customer selection frame
        customer_frame = tk.Frame(form_frame, bg='#ecf0f1')
        customer_frame.pack(fill='x', pady=10)
        
        # Pengirim selection
        pengirim_frame = tk.Frame(customer_frame, bg='#ecf0f1')
        pengirim_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(pengirim_frame, text="Pilih Pengirim:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.pengirim_var = tk.StringVar()
        self.pengirim_combo = ttk.Combobox(pengirim_frame, textvariable=self.pengirim_var, font=('Arial', 11), width=47, state='normal')
        self.pengirim_combo.pack(fill='x', pady=(5, 0))
        
        self.pengirim_combo.bind('<KeyRelease>', self.filter_pengirim)
        
        # Penerima selection
        penerima_frame = tk.Frame(customer_frame, bg='#ecf0f1')
        penerima_frame.pack(fill='x', pady=(10, 0))
        
        tk.Label(penerima_frame, text="Pilih Penerima:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.penerima_var = tk.StringVar()
        self.penerima_combo = ttk.Combobox(penerima_frame, textvariable=self.penerima_var, font=('Arial', 11), width=47, state='normal')
        self.penerima_combo.pack(fill='x', pady=(5, 0))
        
        self.penerima_combo.bind('<KeyRelease>', self.filter_penerima)
        
        self.load_customer_combo()
        self.load_pengirim_combo()
        
        # Barang name
        tk.Label(form_frame, text="Nama Barang:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.barang_entry = tk.Entry(form_frame, font=('Arial', 12), width=50)
        self.barang_entry.pack(fill='x', pady=(5, 10))
        
        # Dimensions frame
        dim_frame = tk.Frame(form_frame, bg='#ecf0f1')
        dim_frame.pack(fill='x', pady=10)
        
        # Panjang
        tk.Label(dim_frame, text="Panjang (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        self.panjang_entry = tk.Entry(dim_frame, font=('Arial', 10), width=10)
        self.panjang_entry.pack(side='left', padx=(5, 20))
        
        # Lebar
        tk.Label(dim_frame, text="Lebar (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        self.lebar_entry = tk.Entry(dim_frame, font=('Arial', 10), width=10)
        self.lebar_entry.pack(side='left', padx=(5, 20))
        
        # Tinggi
        tk.Label(dim_frame, text="Tinggi (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        self.tinggi_entry = tk.Entry(dim_frame, font=('Arial', 10), width=10)
        self.tinggi_entry.pack(side='left', padx=5)
        
        # Other fields frame
        other_frame = tk.Frame(form_frame, bg='#ecf0f1')
        other_frame.pack(fill='x', pady=10)
        
        # M3
        tk.Label(other_frame, text="Volume (m³):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        self.m3_entry = tk.Entry(other_frame, font=('Arial', 10), width=10)
        self.m3_entry.pack(side='left', padx=(5, 20))
        
        # Ton
        tk.Label(other_frame, text="Berat (ton):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        self.ton_entry = tk.Entry(other_frame, font=('Arial', 10), width=10)
        self.ton_entry.pack(side='left', padx=(5, 20))
        
        # Price frame with multiple pricing options
        price_frame = tk.Frame(form_frame, bg='#ecf0f1')
        price_frame.pack(fill='x', pady=10)
        
        tk.Label(price_frame, text="Harga Satuan:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(0, 10))
        
        # Pricing options sub-frame
        pricing_subframe = tk.Frame(price_frame, bg='#ecf0f1')
        pricing_subframe.pack(fill='x')
        
        # === HARGA PER M³ SECTION ===
        m3_section = tk.LabelFrame(pricing_subframe, text="💰 Harga per m³", font=('Arial', 10, 'bold'), 
                                  bg='#ecf0f1', fg='#2c3e50', padx=10, pady=5)
        m3_section.pack(fill='x', pady=(0, 10))
        
        # Harga m3 - Pelabuhan ke Pelabuhan
        m3_pp_frame = tk.Frame(m3_section, bg='#ecf0f1')
        m3_pp_frame.pack(fill='x', pady=2)
        tk.Label(m3_pp_frame, text="Pelabuhan → Pelabuhan:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_m3_pp_entry = tk.Entry(m3_pp_frame, font=('Arial', 9), width=15)
        self.harga_m3_pp_entry.pack(side='right', padx=(5, 0))
        
        # Harga m3 - Pelabuhan ke Door
        m3_pd_frame = tk.Frame(m3_section, bg='#ecf0f1')
        m3_pd_frame.pack(fill='x', pady=2)
        tk.Label(m3_pd_frame, text="Pelabuhan → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_m3_pd_entry = tk.Entry(m3_pd_frame, font=('Arial', 9), width=15)
        self.harga_m3_pd_entry.pack(side='right', padx=(5, 0))
        
        # Harga m3 - Door ke Door
        m3_dd_frame = tk.Frame(m3_section, bg='#ecf0f1')
        m3_dd_frame.pack(fill='x', pady=2)
        tk.Label(m3_dd_frame, text="Door → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_m3_dd_entry = tk.Entry(m3_dd_frame, font=('Arial', 9), width=15)
        self.harga_m3_dd_entry.pack(side='right', padx=(5, 0))
        
        # === HARGA PER TON SECTION ===
        ton_section = tk.LabelFrame(pricing_subframe, text="⚖️ Harga per Ton", font=('Arial', 10, 'bold'), 
                                   bg='#ecf0f1', fg='#2c3e50', padx=10, pady=5)
        ton_section.pack(fill='x', pady=(0, 10))
        
        # Harga ton - Pelabuhan ke Pelabuhan
        ton_pp_frame = tk.Frame(ton_section, bg='#ecf0f1')
        ton_pp_frame.pack(fill='x', pady=2)
        tk.Label(ton_pp_frame, text="Pelabuhan → Pelabuhan:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_ton_pp_entry = tk.Entry(ton_pp_frame, font=('Arial', 9), width=15)
        self.harga_ton_pp_entry.pack(side='right', padx=(5, 0))
        
        # Harga ton - Pelabuhan ke Door
        ton_pd_frame = tk.Frame(ton_section, bg='#ecf0f1')
        ton_pd_frame.pack(fill='x', pady=2)
        tk.Label(ton_pd_frame, text="Pelabuhan → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_ton_pd_entry = tk.Entry(ton_pd_frame, font=('Arial', 9), width=15)
        self.harga_ton_pd_entry.pack(side='right', padx=(5, 0))
        
        # Harga ton - Door ke Door
        ton_dd_frame = tk.Frame(ton_section, bg='#ecf0f1')
        ton_dd_frame.pack(fill='x', pady=2)
        tk.Label(ton_dd_frame, text="Door → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_ton_dd_entry = tk.Entry(ton_dd_frame, font=('Arial', 9), width=15)
        self.harga_ton_dd_entry.pack(side='right', padx=(5, 0))
        
        # === HARGA PER COLLI SECTION ===
        colli_section = tk.LabelFrame(pricing_subframe, text="📦 Harga per Colli", font=('Arial', 10, 'bold'), 
                                     bg='#ecf0f1', fg='#2c3e50', padx=10, pady=5)
        colli_section.pack(fill='x', pady=(0, 10))
        
        # Harga colli - Pelabuhan ke Pelabuhan
        colli_pp_frame = tk.Frame(colli_section, bg='#ecf0f1')
        colli_pp_frame.pack(fill='x', pady=2)
        tk.Label(colli_pp_frame, text="Pelabuhan → Pelabuhan:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_colli_pp_entry = tk.Entry(colli_pp_frame, font=('Arial', 9), width=15)
        self.harga_colli_pp_entry.pack(side='right', padx=(5, 0))
        
        # Harga colli - Pelabuhan ke Door
        colli_pd_frame = tk.Frame(colli_section, bg='#ecf0f1')
        colli_pd_frame.pack(fill='x', pady=2)
        tk.Label(colli_pd_frame, text="Pelabuhan → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_colli_pd_entry = tk.Entry(colli_pd_frame, font=('Arial', 9), width=15)
        self.harga_colli_pd_entry.pack(side='right', padx=(5, 0))
        
        # Harga colli - Door ke Door
        colli_dd_frame = tk.Frame(colli_section, bg='#ecf0f1')
        colli_dd_frame.pack(fill='x', pady=2)
        tk.Label(colli_dd_frame, text="Door → Door:", font=('Arial', 9), bg='#ecf0f1').pack(side='left')
        self.harga_colli_dd_entry = tk.Entry(colli_dd_frame, font=('Arial', 9), width=15)
        self.harga_colli_dd_entry.pack(side='right', padx=(5, 0))
        
        # Note
        note_label = tk.Label(
            price_frame,
            text="💡 Isi sesuai jenis layanan dan metode pricing yang digunakan",
            font=('Arial', 9),
            fg='#7f8c8d',
            bg='#ecf0f1'
        )
        note_label.pack(anchor='w', pady=(5, 0))
        
        # Buttons
        btn_frame = tk.Frame(form_frame, bg='#ecf0f1')
        btn_frame.pack(fill='x', pady=20)
        
        add_btn = tk.Button(
            btn_frame,
            text="➕ Tambah Barang",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.add_barang
        )
        add_btn.pack(side='left', padx=(0, 10))
        
        clear_btn = tk.Button(
            btn_frame,
            text="🗑️ Bersihkan",
            font=('Arial', 12, 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=10,
            pady=5,
            command=self.clear_form
        )
        clear_btn.pack(side='left')
        
        # Store canvas reference for later use if needed
        self.canvas = canvas
        self.scrollable_frame = scrollable_frame
        
        # Focus on pengirim combo
        self.pengirim_combo.focus()
        
    
    def create_excel_tab(self, parent):
        """Create Excel upload tab with scroll isolation - COMPLETE FIXED VERSION"""
        # Main container with scroll
        main_container = tk.Frame(parent, bg='#ecf0f1')
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        canvas = tk.Canvas(main_container, bg='#ecf0f1', highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(main_container, orient="horizontal", command=canvas.xview)

        scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Pack
        h_scrollbar.pack(side="bottom", fill="x")
        v_scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Content frame (now inside scrollable_frame)
        content_frame = tk.Frame(scrollable_frame, bg='#ecf0f1')
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Instructions
        instruction_frame = tk.Frame(content_frame, bg='#ffffff', relief='solid', bd=1)
        instruction_frame.pack(fill='x', pady=(0, 20))
        
        instruction_title = tk.Label(
            instruction_frame,
            text="📊 Upload Data Barang dari Excel",
            font=('Arial', 14, 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        instruction_title.pack(pady=(10, 5))
        
        instruction_text = tk.Label(
            instruction_frame,
            text="Format Excel yang dibutuhkan:\n\n" +
                "• Pengirim: Nama pengirim yang sudah terdaftar (WAJIB)\n" +
                "• Penerima: Nama penerima yang sudah terdaftar (WAJIB)\n" +
                "• Nama Barang: Nama produk/barang (WAJIB)\n" +
                "• P, L, T: Panjang, Lebar, Tinggi (cm)\n" +
                "• M3: Volume (m³), Ton: Berat (ton), Colli: Jumlah kemasan\n\n" +
                "Harga per Jenis Layanan:\n" +
                "• M3_PP, M3_PD, M3_DD: Harga/m³ (Pelabuhan-Pelabuhan, Pelabuhan-Door, Door-Door)\n" +
                "• TON_PP, TON_PD, TON_DD: Harga/ton (Pelabuhan-Pelabuhan, Pelabuhan-Door, Door-Door)\n" +
                "• COLLI_PP, COLLI_PD, COLLI_DD: Harga/colli (Pelabuhan-Pelabuhan, Pelabuhan-Door, Door-Door)\n\n" +
                "Pastikan pengirim dan penerima sudah terdaftar di sistem!",
            font=('Arial', 10),
            fg='#34495e',
            bg='#ffffff',
            justify='left'
        )
        instruction_text.pack(pady=(0, 10), padx=20)
        
        # File selection
        file_frame = tk.Frame(content_frame, bg='#ecf0f1')
        file_frame.pack(fill='x', pady=10)
        
        tk.Label(file_frame, text="Pilih File Excel:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        
        file_input_frame = tk.Frame(file_frame, bg='#ecf0f1')
        file_input_frame.pack(fill='x', pady=(5, 0))
        
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(file_input_frame, textvariable=self.file_path_var, font=('Arial', 11), state='readonly')
        self.file_entry.pack(side='left', fill='x', expand=True, ipady=5)
        
        browse_btn = tk.Button(
            file_input_frame,
            text="📁 Browse",
            font=('Arial', 10, 'bold'),
            bg='#3498db',
            fg='white',
            padx=15,
            pady=5,
            command=self.browse_file
        )
        browse_btn.pack(side='right', padx=(5, 0))
        
        # Preview area dengan scroll isolation
        preview_frame = tk.Frame(content_frame, bg='#ecf0f1')
        preview_frame.pack(fill='both', expand=True, pady=10)
        
        tk.Label(preview_frame, text="📋 Preview Data:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(0, 5))
        
        # ✅ KUNCI UTAMA: Container khusus untuk TreeView dengan scroll isolation
        tree_container = tk.Frame(preview_frame, bg='#ecf0f1', relief='solid', bd=1)
        tree_container.pack(fill='both', expand=True, pady=5)
        
        # Updated columns
        columns = ('Pengirim', 'Penerima', 'Nama', 'P', 'L', 'T', 'M3', 'Ton', 
                'M3_PP', 'M3_PD', 'M3_DD', 'TON_PP', 'TON_PD', 'TON_DD', 'COLLI_PP', 'COLLI_PD', 'COLLI_DD')
        
        # ✅ GUNAKAN TREEVIEW BIASA (sesuai kode asli Anda)
        self.preview_tree = ttk.Treeview(tree_container, 
                                    columns=columns, 
                                    show='headings',
                                    height=20)
        
        # Configure column headers
        headers = {
            'Pengirim': 'Pengirim',
            'Penerima': 'Penerima',
            'Nama': 'Nama Barang',
            'P': 'P(cm)',
            'L': 'L(cm)', 
            'T': 'T(cm)',
            'M3': 'M³',
            'Ton': 'Ton',
            'M3_PP': 'M³ P→P',
            'M3_PD': 'M³ P→D',
            'M3_DD': 'M³ D→D',
            'TON_PP': 'Ton P→P',
            'TON_PD': 'Ton P→D',
            'TON_DD': 'Ton D→D',
            'COLLI_PP': 'Colli P→P',
            'COLLI_PD': 'Colli P→D',
            'COLLI_DD': 'Colli D→D'
        }
        
        for col_id, header_text in headers.items():
            self.preview_tree.heading(col_id, text=header_text)
        
        # Configure column widths
        column_widths = {
            'Pengirim': 150,
            'Penerima': 150,
            'Nama': 180,
            'P': 70,
            'L': 70,
            'T': 70,
            'M3': 80,
            'Ton': 80,
            'M3_PP': 100,
            'M3_PD': 100,
            'M3_DD': 100,
            'TON_PP': 100,
            'TON_PD': 100,
            'TON_DD': 100,
            'COLLI_PP': 100,
            'COLLI_PD': 100,
            'COLLI_DD': 100
        }
        
        for col_id, width in column_widths.items():
            self.preview_tree.column(col_id, width=width, minwidth=width, stretch=False)
        
        # Create scrollbars untuk TreeView
        tree_v_scrollbar = ttk.Scrollbar(tree_container, orient='vertical', command=self.preview_tree.yview)
        tree_h_scrollbar = ttk.Scrollbar(tree_container, orient='horizontal', command=self.preview_tree.xview)
        
        # Configure treeview scrollbars
        self.preview_tree.configure(yscrollcommand=tree_v_scrollbar.set, xscrollcommand=tree_h_scrollbar.set)
        
        # Pack TreeView dan scrollbars
        tree_h_scrollbar.pack(side='bottom', fill='x')
        tree_v_scrollbar.pack(side='right', fill='y')
        self.preview_tree.pack(fill='both', expand=True)
        
        # ✅ IMPLEMENTASI SCROLL ISOLATION YANG BENAR
        def setup_scroll_isolation():
            """Setup scroll isolation untuk preview tree"""
            
            # Mouse tracking untuk tree area
            self.mouse_in_tree_area = False
            
            def on_tree_area_enter(event):
                """Mouse masuk ke area tree - disable canvas scroll"""
                self.mouse_in_tree_area = True
                # Unbind canvas scroll
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")  # Linux
                canvas.unbind_all("<Button-5>")  # Linux
                print("Mouse entered tree area - canvas scroll disabled")
            
            def on_tree_area_leave(event):
                """Mouse keluar dari area tree - enable canvas scroll"""
                self.mouse_in_tree_area = False
                # Re-enable canvas scroll dengan delay untuk menghindari conflict
                self.window.after(50, restore_canvas_scroll)
                print("Mouse left tree area - scheduling canvas scroll restore")
            
            def restore_canvas_scroll():
                """Restore canvas scrolling functionality"""
                if not getattr(self, 'mouse_in_tree_area', False):
                    def _on_canvas_mousewheel(event):
                        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                    
                    def _on_canvas_button4(event):  # Linux scroll up
                        canvas.yview_scroll(-1, "units")
                        
                    def _on_canvas_button5(event):  # Linux scroll down
                        canvas.yview_scroll(1, "units")
                    
                    canvas.bind_all("<MouseWheel>", _on_canvas_mousewheel)
                    canvas.bind_all("<Button-4>", _on_canvas_button4)
                    canvas.bind_all("<Button-5>", _on_canvas_button5)
                    print("Canvas scroll restored")
            
            def on_tree_scroll(event):
                """Handle scroll events dalam tree area"""
                if self.mouse_in_tree_area:
                    # Allow TreeView to handle its own scrolling
                    try:
                        # Determine scroll direction
                        if event.delta > 0:  # Scroll up
                            self.preview_tree.yview_scroll(-1, "units")
                        else:  # Scroll down
                            self.preview_tree.yview_scroll(1, "units")
                    except:
                        # Fallback for systems without delta
                        if event.num == 4:  # Linux scroll up
                            self.preview_tree.yview_scroll(-1, "units")
                        elif event.num == 5:  # Linux scroll down
                            self.preview_tree.yview_scroll(1, "units")
                    
                    # Stop event propagation
                    return "break"
                
                # If not in tree area, let canvas handle
                return None
            
            # Bind events ke tree container dan semua children
            def bind_tree_isolation_events(widget):
                """Bind isolation events ke widget dan children secara recursive"""
                try:
                    widget.bind("<Enter>", on_tree_area_enter, '+')
                    widget.bind("<Leave>", on_tree_area_leave, '+')
                    widget.bind("<MouseWheel>", on_tree_scroll, '+')
                    widget.bind("<Button-4>", on_tree_scroll, '+')  # Linux
                    widget.bind("<Button-5>", on_tree_scroll, '+')  # Linux
                    
                    # Recursive untuk children
                    for child in widget.winfo_children():
                        bind_tree_isolation_events(child)
                        
                except Exception as e:
                    print(f"Error binding events to {widget}: {e}")
            
            # Apply ke tree container dan semua contents
            bind_tree_isolation_events(tree_container)
            
            # Special handling untuk TreeView scrollbars
            try:
                tree_v_scrollbar.bind("<Enter>", on_tree_area_enter, '+')
                tree_v_scrollbar.bind("<Leave>", on_tree_area_leave, '+')
                tree_h_scrollbar.bind("<Enter>", on_tree_area_enter, '+')
                tree_h_scrollbar.bind("<Leave>", on_tree_area_leave, '+')
            except:
                pass
            
            print("✅ Scroll isolation setup completed")
        
        # ✅ SETUP CANVAS SCROLL (untuk area di luar tree)
        def setup_canvas_scroll():
            """Setup default canvas scrolling"""
            def _on_canvas_mousewheel(event):
                if not getattr(self, 'mouse_in_tree_area', False):
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
            def _on_canvas_button4(event):  # Linux scroll up
                if not getattr(self, 'mouse_in_tree_area', False):
                    canvas.yview_scroll(-1, "units")
                    
            def _on_canvas_button5(event):  # Linux scroll down
                if not getattr(self, 'mouse_in_tree_area', False):
                    canvas.yview_scroll(1, "units")
            
            def _bind_canvas_mousewheel(event):
                if not getattr(self, 'mouse_in_tree_area', False):
                    canvas.bind_all("<MouseWheel>", _on_canvas_mousewheel)
                    canvas.bind_all("<Button-4>", _on_canvas_button4)
                    canvas.bind_all("<Button-5>", _on_canvas_button5)
            
            def _unbind_canvas_mousewheel(event):
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")
            
            # Bind canvas enter/leave events
            canvas.bind('<Enter>', _bind_canvas_mousewheel)
            canvas.bind('<Leave>', _unbind_canvas_mousewheel)
            
            # Initial canvas scroll binding
            canvas.bind_all("<MouseWheel>", _on_canvas_mousewheel)
            canvas.bind_all("<Button-4>", _on_canvas_button4)
            canvas.bind_all("<Button-5>", _on_canvas_button5)
        
        # Bind keyboard events untuk TreeView navigation (dari kode asli)
        def on_key_press(event):
            if event.keysym == 'Right':
                self.preview_tree.xview_scroll(1, "units")
                return "break"
            elif event.keysym == 'Left':
                self.preview_tree.xview_scroll(-1, "units")
                return "break"
            elif event.keysym == 'Down':
                self.preview_tree.yview_scroll(1, "units")
                return "break"
            elif event.keysym == 'Up':
                self.preview_tree.yview_scroll(-1, "units")
                return "break"
        
        # Bind keyboard events
        self.preview_tree.bind('<Key>', on_key_press)
        self.preview_tree.bind('<Button-1>', lambda e: self.preview_tree.focus_set())
        
        # Apply scroll isolation setelah semua widget dibuat
        self.window.after(100, setup_scroll_isolation)
        self.window.after(200, setup_canvas_scroll)
        
        # Upload buttons
        upload_btn_frame = tk.Frame(content_frame, bg='#ecf0f1')
        upload_btn_frame.pack(fill='x', pady=15)
        
        self.upload_btn = tk.Button(
            upload_btn_frame,
            text="⬆️ Upload ke Database",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.upload_excel_data,
            state='disabled'
        )
        self.upload_btn.pack(side='left', padx=(0, 15))
        
        download_template_btn = tk.Button(
            upload_btn_frame,
            text="📥 Download Template",
            font=('Arial', 12, 'bold'),
            bg='#f39c12',
            fg='white',
            padx=10,
            pady=5,
            command=self.download_template
        )
        download_template_btn.pack(side='left')
        
        # Status label
        self.status_label = tk.Label(
            content_frame,
            text="",
            font=('Arial', 11),
            fg='#e74c3c',
            bg='#ecf0f1',
            wraplength=1000,
            justify='left'
        )
        self.status_label.pack(pady=10, fill='x')
        
        # Store references
        self.excel_canvas = canvas
        self.excel_scrollable_frame = scrollable_frame
        self.tree_container = tree_container
    
    def create_list_tab(self, parent):
        """Create barang list tab with search, update, delete functionality"""
        # Container
        list_container = tk.Frame(parent, bg='#ecf0f1')
        list_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Header
        header_frame = tk.Frame(list_container, bg='#ecf0f1')
        header_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(header_frame, text="📋 DAFTAR BARANG", font=('Arial', 14, 'bold'), bg='#ecf0f1').pack(side='left')
        
        # Search/Filter Frame
        search_frame = tk.Frame(list_container, bg='#ffffff', relief='solid', bd=1)
        search_frame.pack(fill='x', pady=(0, 10))
        
        # Search label
        search_label = tk.Label(
            search_frame,
            text="🔍 Filter & Pencarian:",
            font=('Arial', 12, 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        search_label.pack(anchor='w', padx=10, pady=(10, 5))
        
        # Search controls frame - First row
        search_controls_row1 = tk.Frame(search_frame, bg='#ffffff')
        search_controls_row1.pack(fill='x', padx=10, pady=(0, 5))
        
        # Search by name
        tk.Label(search_controls_row1, text="Nama Barang:", font=('Arial', 10), bg='#ffffff').pack(side='left')
        self.search_name_var = tk.StringVar()
        self.search_name_var.trace('w', self.on_search_change)
        search_name_entry = tk.Entry(search_controls_row1, textvariable=self.search_name_var, font=('Arial', 10), width=20)
        search_name_entry.pack(side='left', padx=(5, 15))
        
        # Clear search button
        clear_search_btn = tk.Button(
            search_controls_row1,
            text="❌ Clear",
            font=('Arial', 9),
            bg='#e67e22',
            fg='white',
            padx=10,
            pady=2,
            command=self.clear_barang_filter
        )
        clear_search_btn.pack(side='right', padx=5)
        
        # Search controls frame - Second row
        search_controls_row2 = tk.Frame(search_frame, bg='#ffffff')
        search_controls_row2.pack(fill='x', padx=10, pady=(0, 10))
        
        # Filter by Pengirim
        tk.Label(search_controls_row2, text="Pengirim:", font=('Arial', 10), bg='#ffffff').pack(side='left')
        self.filter_pengirim_var = tk.StringVar()
        self.filter_pengirim_var.trace('w', self.on_search_change)
        self.filter_pengirim_combo = ttk.Combobox(
            search_controls_row2, 
            textvariable=self.filter_pengirim_var, 
            font=('Arial', 10), 
            width=18,
            state='normal'
        )
        self.filter_pengirim_combo.pack(side='left', padx=(5, 15))
        self.filter_pengirim_combo.bind('<KeyRelease>', self.filter_pengirim_list)
        
        # Filter by Penerima
        tk.Label(search_controls_row2, text="Penerima:", font=('Arial', 10), bg='#ffffff').pack(side='left')
        self.filter_penerima_var = tk.StringVar()
        self.filter_penerima_var.trace('w', self.on_search_change)
        self.filter_penerima_combo = ttk.Combobox(
            search_controls_row2, 
            textvariable=self.filter_penerima_var, 
            font=('Arial', 10), 
            width=18,
            state='normal'
        )
        self.filter_penerima_combo.pack(side='left', padx=(5, 15))
        self.filter_penerima_combo.bind('<KeyRelease>', self.filter_penerima_list)

        # Action buttons frame
        action_frame = tk.Frame(list_container, bg='#ecf0f1')
        action_frame.pack(fill='x', pady=(0, 10))
        
        # Update button
        update_btn = tk.Button(
            action_frame,
            text="✏️ Edit Barang",
            font=('Arial', 11, 'bold'),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=5,
            command=self.update_barang
        )
        update_btn.pack(side='left', padx=(0, 10))
        
        # Delete button
        delete_btn = tk.Button(
            action_frame,
            text="🗑️ Hapus Barang",
            font=('Arial', 11, 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=10,
            pady=5,
            command=self.delete_barang
        )
        delete_btn.pack(side='left', padx=(0, 10))
        
        # Export button
        export_btn = tk.Button(
            action_frame,
            text="📤 Export Excel",
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.export_barang
        )
        export_btn.pack(side='left')
        
        # Info label
        self.info_label = tk.Label(
            action_frame,
            text="💡 Pilih barang dari tabel untuk edit/hapus",
            font=('Arial', 10),
            fg='#7f8c8d',
            bg='#ecf0f1'
        )
        self.info_label.pack(side='right')
        
        # Treeview for barang list with scrollbars
        tree_frame = tk.Frame(list_container, bg='#ecf0f1')
        tree_frame.pack(fill='both', expand=True)
        
        tree_container = tk.Frame(tree_frame, bg='#ecf0f1')
        tree_container.pack(fill='both', expand=True)
        
        columns = ('ID', 'Pengirim', 'Penerima', 'Nama', 'Dimensi', 'Volume', 'Berat', 
               'Harga/M3_PP', 'Harga/M3_PD', 'Harga/M3_DD', 'Harga/Ton_PP', 
               'Harga/Ton_PD', 'Harga/Ton_DD', 'Harga/Col_PP', 'Harga/Col_PD', 
               'Harga/Col_DD', 'Created')
        
        self.tree = PaginatedTreeView(
            parent=tree_container,
            columns=columns,
            show='headings',
            height=12,
            items_per_page=100  
        )
        
        # Configure columns
        self.tree.heading('ID', text='ID')
        self.tree.heading('Pengirim', text='Pengirim')
        self.tree.heading('Penerima', text='Penerima')
        self.tree.heading('Nama', text='Nama Barang')
        self.tree.heading('Dimensi', text='P×L×T (cm)')
        self.tree.heading('Volume', text='Volume (m³)')
        self.tree.heading('Berat', text='Berat (ton)')
        self.tree.heading('Harga/M3_PP', text='Harga/M3_PP (Rp)')
        self.tree.heading('Harga/M3_PD', text='Harga/M3_PD (Rp)')
        self.tree.heading('Harga/M3_DD', text='Harga/M3_DD (Rp)')
        self.tree.heading('Harga/Ton_PP', text='Harga/Ton_PP (Rp)')
        self.tree.heading('Harga/Ton_PD', text='Harga/Ton_PD (Rp)')
        self.tree.heading('Harga/Ton_DD', text='Harga/Ton_DD (Rp)')
        self.tree.heading('Harga/Col_PP', text='Harga/Col_PP (Rp)')
        self.tree.heading('Harga/Col_PD', text='Harga/Col_PD (Rp)')
        self.tree.heading('Harga/Col_DD', text='Harga/Col_DD (Rp)')
        self.tree.heading('Created', text='Tanggal Dibuat')
        
        self.tree.column('ID', width=40)
        self.tree.column('Pengirim', width=150)
        self.tree.column('Penerima', width=150)
        self.tree.column('Nama', width=200)
        self.tree.column('Dimensi', width=100)
        self.tree.column('Volume', width=80)
        self.tree.column('Berat', width=80)
        self.tree.column('Harga/M3_PP', width=100)
        self.tree.column('Harga/M3_PD', width=100)
        self.tree.column('Harga/M3_DD', width=100)
        self.tree.column('Harga/Ton_PP', width=100)
        self.tree.column('Harga/Ton_PD', width=100)
        self.tree.column('Harga/Ton_DD', width=100)
        self.tree.column('Harga/Col_PP', width=100)
        self.tree.column('Harga/Col_PD', width=100)
        self.tree.column('Harga/Col_DD', width=100)
        self.tree.column('Created', width=120)
        
        # Bind double-click to edit
        self.tree.bind('<Double-1>', lambda e: self.update_barang())
        
        # Bind selection change to update info
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
        self.tree.pack(fill='both', expand=True)
        
        # Store original data for filtering
        self.original_barang_data = []
        self.original_pengirim_data = []
        self.original_penerima_data = []
        
        # Load existing barang
        self.load_pengirim_penerima_filter()
        
        # Add tab change event to refresh data when switching to this tab
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
    def load_pengirim_penerima_filter(self):
        """Load unique pengirim and penerima for filter dropdowns"""
        try:
            customers = self.db.get_all_customers()
            customers_list = [c['nama_customer'] for c in customers if 'nama_customer' in c]

            # Update combobox values
            customers_list = sorted(list(customers_list))

            self.filter_pengirim_combo['values'] = customers_list
            self.filter_penerima_combo['values'] = customers_list
            
            
        except Exception as e:
            print(f"Error loading pengirim/penerima filter: {e}")
            
    def on_search_change(self, *args):
        """Handle search input changes"""
        self.filter_barang()
    
    def clear_barang_filter(self):
        """Clear semua filter dan reload semua data barang"""
        try:
            # Clear filter inputs (sesuaikan dengan nama field Anda)
            if hasattr(self, 'search_nama'):
                self.search_nama.delete(0, tk.END)
            if hasattr(self, 'search_pengirim'):
                if hasattr(self.search_pengirim, 'set'):  # Combobox
                    self.search_pengirim.set('')
                else:  # Entry
                    self.search_pengirim.delete(0, tk.END)
            if hasattr(self, 'search_penerima'):
                if hasattr(self.search_penerima, 'set'):  # Combobox
                    self.search_penerima.set('')
                else:  # Entry
                    self.search_penerima.delete(0, tk.END)
            
            # Reload semua data
            self.load_barang()
            
        except Exception as e:
            print(f"Error clearing filter: {str(e)}")
    
    def filter_barang(self):
        """Filter barang based on search criteria for PaginatedTreeView"""
        try:
            if not hasattr(self, 'original_barang_data') or not self.original_barang_data:
                self.load_barang()  # Load data if not available
                return
            
            # Get search criteria
            search_name = self.search_name_var.get().strip()
            filter_pengirim = self.filter_pengirim_var.get()
            filter_penerima = self.filter_penerima_var.get()
            
            print(f"Filtering with: name='{search_name}', pengirim='{filter_pengirim}', penerima='{filter_penerima}'")
            
            # Filter data
            filtered_data = []
            
            for barang in self.original_barang_data:
                show_item = True
                
                print(f"Processing item: {barang.get('nama_barang', 'Unknown')}")

                # Extract values for filtering (assuming dict structure based on your code)
                if isinstance(barang, dict):
                    # Dictionary structure
                    nama_barang = str(barang.get('nama_barang', '')).lower()
                    
                    # Get pengirim name
                    pengirim_id = barang.get('pengirim', '')
                    try:
                        pengirim_data = self.db.get_customer_by_id(pengirim_id) if pengirim_id else {}
                        pengirim = pengirim_data.get('nama_customer', '') if pengirim_data else ''
                    except Exception as e:
                        print(f"Error getting pengirim data: {e}")
                        pengirim = ''
                    
                    # Get penerima name
                    penerima_id = barang.get('penerima', '')
                    try:
                        penerima_data = self.db.get_customer_by_id(penerima_id) if penerima_id else {}
                        penerima = penerima_data.get('nama_customer', '') if penerima_data else ''
                    except Exception as e:
                        print(f"Error getting penerima data: {e}")
                        penerima = ''
                    
                    print(f"Pengirim ID: {pengirim_id}, Name: {pengirim}")
                    print(f"Penerima ID: {penerima_id}, Name: {penerima}")
                    
                else:
                    # List/tuple structure - adjust indices based on your column order
                    nama_barang = str(barang[3]).lower() if len(barang) > 3 else ''
                    pengirim = str(barang[1]) if len(barang) > 1 else ''
                    penerima = str(barang[2]) if len(barang) > 2 else ''
                
                # Check name filter with flexible matching
                if search_name:
                    search_terms = search_name.lower().split()
                    match_found = False
                    
                    for term in search_terms:
                        if term in nama_barang:
                            match_found = True
                            break
                    
                    if not match_found:
                        show_item = False
                        print(f"Name filter failed for: {nama_barang}")
                
                # Check pengirim filter
                if filter_pengirim and filter_pengirim.strip() != '':
                    if filter_pengirim.lower() not in str(pengirim).lower():
                        show_item = False
                        print(f"Pengirim filter failed: {pengirim} vs {filter_pengirim}")
                
                # Check penerima filter
                if filter_penerima and filter_penerima.strip() != '':
                    if filter_penerima.lower() not in str(penerima).lower():
                        show_item = False
                        print(f"Penerima filter failed: {penerima} vs {filter_penerima}")
                
                if show_item:
                    filtered_data.append(barang)
            
            print(f"Filtered {len(filtered_data)} items from {len(self.original_barang_data)} total")
            
            # Format filtered data for PaginatedTreeView
            formatted_data = []
            
            for barang in filtered_data:
                if isinstance(barang, dict):
                    # Dictionary structure - format for display
                    dimensi = f"{barang.get('panjang_barang', '-')}×{barang.get('lebar_barang', '-')}×{barang.get('tinggi_barang', '-')}"
                    
                    # Format currency
                    harga_m3_pp = f"Rp {barang.get('m3_pp', 0):,.0f}" if barang.get('m3_pp') else '-'
                    harga_m3_pd = f"Rp {barang.get('m3_pd', 0):,.0f}" if barang.get('m3_pd') else '-'
                    harga_m3_dd = f"Rp {barang.get('m3_dd', 0):,.0f}" if barang.get('m3_dd') else '-'
                    harga_ton_pp = f"Rp {barang.get('ton_pp', 0):,.0f}" if barang.get('ton_pp') else '-'
                    harga_ton_pd = f"Rp {barang.get('ton_pd', 0):,.0f}" if barang.get('ton_pd') else '-'
                    harga_ton_dd = f"Rp {barang.get('ton_dd', 0):,.0f}" if barang.get('ton_dd') else '-'
                    harga_col_pp = f"Rp {barang.get('col_pp', 0):,.0f}" if barang.get('col_pp') else '-'
                    harga_col_pd = f"Rp {barang.get('col_pd', 0):,.0f}" if barang.get('col_pd') else '-'
                    harga_col_dd = f"Rp {barang.get('col_dd', 0):,.0f}" if barang.get('col_dd') else '-'
                    
                    # Format date
                    created_date = barang.get('created_at', '')[:10] if barang.get('created_at') else '-'
                    
                    # Create row tuple
                    row_data = (
                        barang.get('barang_id', ''),
                        barang.get('sender_name', ''),
                        barang.get('receiver_name', ''),
                        barang.get('nama_barang', ''),
                        dimensi,
                        barang.get('m3_barang', '-'),
                        barang.get('ton_barang', '-'),
                        harga_m3_pp,
                        harga_m3_pd,
                        harga_m3_dd,
                        harga_ton_pp,
                        harga_ton_pd,
                        harga_ton_dd,
                        harga_col_pp,
                        harga_col_pd,
                        harga_col_dd,
                        created_date
                    )
                    
                    formatted_data.append(row_data)
                else:
                    # List/tuple structure - add as is
                    formatted_data.append(barang)
            
            # Set filtered data to PaginatedTreeView
            self.tree.set_data(formatted_data)
            
            # Update info label
            total_count = len(self.original_barang_data)
            filtered_count = len(filtered_data)
            
            if hasattr(self, 'info_label'):
                if total_count != filtered_count:
                    self.info_label.config(text=f"Menampilkan {filtered_count} dari {total_count} barang")
                else:
                    self.info_label.config(text="Pilih barang dari tabel untuk edit/hapus")
            
            print(f"Filter completed successfully: {filtered_count} items displayed")
            
        except Exception as e:
            print(f"Error in filter_barang: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal menerapkan filter: {str(e)}")
                
    def update_barang(self):
        """Update selected barang"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih barang yang akan diedit dari tabel!")
            return
        
        # Get selected item data
        item = self.tree.item(selection[0])
        barang_id = item['values'][0]
        
        # Find full barang data
        selected_barang = None
        for barang in self.original_barang_data:
            # Handle both dictionary and list/tuple structures
            if isinstance(barang, dict):
                if barang['barang_id'] == barang_id:
                    selected_barang = barang
                    break
            else:
                # For list/tuple structure, ID is at index 0
                if str(barang[0]) == str(barang_id):
                    # Convert to dictionary structure for easier handling
                    selected_barang = {
                        'barang_id': barang[0] if len(barang) > 0 else '',
                        'pengirim': barang[1] if len(barang) > 1 else '',
                        'penerima': barang[2] if len(barang) > 2 else '',
                        'nama_barang': barang[3] if len(barang) > 3 else '',
                        'dimensi': barang[4] if len(barang) > 4 else '',
                        'volume_barang': barang[5] if len(barang) > 5 else '',
                        'berat_barang': barang[6] if len(barang) > 6 else '',
                        'colli_barang': barang[7] if len(barang) > 7 else '',
                        'harga_m3_pp': barang[8] if len(barang) > 8 else '',
                        'harga_m3_pd': barang[9] if len(barang) > 9 else '',
                        'harga_m3_dd': barang[10] if len(barang) > 10 else '',
                        'harga_ton_pp': barang[11] if len(barang) > 11 else '',
                        'harga_ton_pd': barang[12] if len(barang) > 12 else '',
                        'harga_ton_dd': barang[13] if len(barang) > 13 else '',
                        'harga_col_pp': barang[14] if len(barang) > 14 else '',
                        'harga_col_pd': barang[15] if len(barang) > 15 else '',
                        'harga_col_dd': barang[16] if len(barang) > 16 else '',
                        'created_at': barang[17] if len(barang) > 17 else ''
                    }
                    # Parse dimensions if it's in "PxLxT" format
                    if selected_barang['dimensi'] and '×' in selected_barang['dimensi']:
                        dims = selected_barang['dimensi'].split('×')
                        selected_barang['panjang_barang'] = dims[0] if len(dims) > 0 and dims[0] != '-' else ''
                        selected_barang['lebar_barang'] = dims[1] if len(dims) > 1 and dims[1] != '-' else ''
                        selected_barang['tinggi_barang'] = dims[2] if len(dims) > 2 and dims[2] != '-' else ''
                    break
        
        print(f"Selected barang: {selected_barang}")

        if not selected_barang:
            messagebox.showerror("Error", "Data barang tidak ditemukan!")
            return
        
        # Open update dialog
        self.open_update_dialog(selected_barang)

    def save_changes(self, updated_barang):
        try:
            print(f"Updated data: {updated_barang}")
            self.db.update_barang(updated_barang)
            messagebox.showinfo("Sukses", "Data barang berhasil disimpan!")
            self.load_barang()
            self.load_pengirim_penerima_filter()  # Refresh filter options
        except Exception as e:
            print(f"Error saat menyimpan data: {e}")
            messagebox.showerror("Error", f"Gagal menyimpan data barang!\nError: {str(e)}")

    def open_update_dialog(self, barang_data):
        """Open dialog to update barang data"""
        # Create update window
        update_window = tk.Toplevel(self.window)
        update_window.title(f"✏️ Edit Barang - {barang_data.get('nama_barang', 'Unknown')}")
        update_window.geometry("700x800")
        update_window.configure(bg='#ecf0f1')
        update_window.transient(self.window)
        update_window.grab_set()
        
        # Center window
        update_window.update_idletasks()
        x = (update_window.winfo_screenwidth() // 2) - (700 // 2)
        y = (update_window.winfo_screenheight() // 2) - (800 // 2)
        update_window.geometry(f"700x800+{x}+{y}")
        
        # Header
        header = tk.Label(
            update_window,
            text="✏️ EDIT DATA BARANG",
            font=('Arial', 16, 'bold'),
            bg='#3498db',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Form frame with scrollbar
        canvas = tk.Canvas(update_window, bg='#ecf0f1')
        scrollbar = ttk.Scrollbar(update_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        scrollbar.pack(side="right", fill="y")
        
        form_frame = scrollable_frame
        
        customers = self.db.get_all_customers()
        customer_names = [c['nama_customer'] for c in customers]

        pengirim_name = self.db.get_customer_by_id(barang_data.get('pengirim', '')).get('nama_customer', '') if barang_data.get('pengirim') else ''
        penerima_name = self.db.get_customer_by_id(barang_data.get('penerima', '')).get('nama_customer', '') if barang_data.get('penerima') else ''
        # Pengirim
        tk.Label(form_frame, text="Pengirim:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(10, 0))
        pengirim_var = tk.StringVar(value=pengirim_name)
        pengirim_combo = ttk.Combobox(
            form_frame,
            textvariable=pengirim_var,
            font=('Arial', 11),
            values=customer_names,
            state='normal'  # Allow typing new values
        )
        pengirim_combo.pack(fill='x', pady=(5, 10))
        
        # Penerima
        tk.Label(form_frame, text="Penerima:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        penerima_var = tk.StringVar(value=penerima_name)
        penerima_combo = ttk.Combobox(
            form_frame,
            textvariable=penerima_var,
            font=('Arial', 11),
            values=customer_names,
            state='normal'  # Allow typing new values
        )
        penerima_combo.pack(fill='x', pady=(5, 10))
        
        # Nama Barang
        tk.Label(form_frame, text="Nama Barang:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        nama_barang_var = tk.StringVar(value=barang_data.get('nama_barang', ''))
        nama_barang_entry = tk.Entry(form_frame, textvariable=nama_barang_var, font=('Arial', 11))
        nama_barang_entry.pack(fill='x', pady=(5, 10))
        

        # Dimensions
        tk.Label(form_frame, text="Dimensi Barang:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w')
        dim_frame = tk.Frame(form_frame, bg='#ecf0f1')
        dim_frame.pack(fill='x', pady=(5, 10))

        tk.Label(dim_frame, text="Panjang (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        panjang_var = tk.StringVar(value=str(barang_data.get('panjang_barang', '') or '') or '-')
        panjang_entry = tk.Entry(dim_frame, textvariable=panjang_var, font=('Arial', 10), width=10)
        panjang_entry.pack(side='left', padx=(5, 20))

        tk.Label(dim_frame, text="Lebar (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        lebar_var = tk.StringVar(value=str(barang_data.get('lebar_barang', '') or '') or '-')
        lebar_entry = tk.Entry(dim_frame, textvariable=lebar_var, font=('Arial', 10), width=10)
        lebar_entry.pack(side='left', padx=(5, 20))

        tk.Label(dim_frame, text="Tinggi (cm):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        tinggi_var = tk.StringVar(value=str(barang_data.get('tinggi_barang', '') or '') or '-')
        tinggi_entry = tk.Entry(dim_frame, textvariable=tinggi_var, font=('Arial', 10), width=10)
        tinggi_entry.pack(side='left', padx=5)

        # Other fields
        tk.Label(form_frame, text="Spesifikasi Barang:", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(10, 0))
        other_frame = tk.Frame(form_frame, bg='#ecf0f1')
        other_frame.pack(fill='x', pady=(5, 10))

        tk.Label(other_frame, text="Volume (m³):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        volume_var = tk.StringVar(value=str(barang_data.get('m3_barang', '') or '') or '-')
        volume_entry = tk.Entry(other_frame, textvariable=volume_var, font=('Arial', 10), width=10)
        volume_entry.pack(side='left', padx=(5, 20))

        tk.Label(other_frame, text="Berat (ton):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        berat_var = tk.StringVar(value=str(barang_data.get('ton_barang', '') or '') or '-')
        berat_entry = tk.Entry(other_frame, textvariable=berat_var, font=('Arial', 10), width=10)
        berat_entry.pack(side='left', padx=(5, 20))

        tk.Label(other_frame, text="Colli:", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side='left')
        colli_var = tk.StringVar(value=str(barang_data.get('col_barang', '') or '') or '-')
        colli_entry = tk.Entry(other_frame, textvariable=colli_var, font=('Arial', 10), width=10)
        colli_entry.pack(side='left', padx=5)
        
        # Pricing section
        tk.Label(form_frame, text="Harga Pickup to Pickup (PP):", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(20, 5))
        
        # PP Prices
        pp_frame = tk.Frame(form_frame, bg='#ecf0f1')
        pp_frame.pack(fill='x', pady=(0, 10))
        
        # Harga M3 PP
        harga_m3_pp_frame = tk.Frame(pp_frame, bg='#ecf0f1')
        harga_m3_pp_frame.pack(fill='x', pady=2)
        tk.Label(harga_m3_pp_frame, text="Harga/m³ (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_m3_pp_var = tk.StringVar(value=str(barang_data.get('m3_pp', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_m3_pp_entry = tk.Entry(harga_m3_pp_frame, textvariable=harga_m3_pp_var, font=('Arial', 10), width=20)
        harga_m3_pp_entry.pack(side='left', padx=(5, 0))

        # Harga Ton PP
        harga_ton_pp_frame = tk.Frame(pp_frame, bg='#ecf0f1')
        harga_ton_pp_frame.pack(fill='x', pady=2)
        tk.Label(harga_ton_pp_frame, text="Harga/ton (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_ton_pp_var = tk.StringVar(value=str(barang_data.get('ton_pp', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_ton_pp_entry = tk.Entry(harga_ton_pp_frame, textvariable=harga_ton_pp_var, font=('Arial', 10), width=20)
        harga_ton_pp_entry.pack(side='left', padx=(5, 0))

        # Harga Col PP
        harga_col_pp_frame = tk.Frame(pp_frame, bg='#ecf0f1')
        harga_col_pp_frame.pack(fill='x', pady=2)
        tk.Label(harga_col_pp_frame, text="Harga/colli (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_col_pp_var = tk.StringVar(value=str(barang_data.get('col_pp', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_col_pp_entry = tk.Entry(harga_col_pp_frame, textvariable=harga_col_pp_var, font=('Arial', 10), width=20)
        harga_col_pp_entry.pack(side='left', padx=(5, 0))
        
        # PD Prices
        tk.Label(form_frame, text="Harga Pickup to Door (PD):", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(10, 5))

        pd_frame = tk.Frame(form_frame, bg='#ecf0f1')
        pd_frame.pack(fill='x', pady=(0, 10))

        # Similar structure for PD prices
        harga_m3_pd_frame = tk.Frame(pd_frame, bg='#ecf0f1')
        harga_m3_pd_frame.pack(fill='x', pady=2)
        tk.Label(harga_m3_pd_frame, text="Harga/m³ (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_m3_pd_var = tk.StringVar(value=str(barang_data.get('m3_pd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_m3_pd_entry = tk.Entry(harga_m3_pd_frame, textvariable=harga_m3_pd_var, font=('Arial', 10), width=20)
        harga_m3_pd_entry.pack(side='left', padx=(5, 0))

        harga_ton_pd_frame = tk.Frame(pd_frame, bg='#ecf0f1')
        harga_ton_pd_frame.pack(fill='x', pady=2)
        tk.Label(harga_ton_pd_frame, text="Harga/ton (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_ton_pd_var = tk.StringVar(value=str(barang_data.get('ton_pd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_ton_pd_entry = tk.Entry(harga_ton_pd_frame, textvariable=harga_ton_pd_var, font=('Arial', 10), width=20)
        harga_ton_pd_entry.pack(side='left', padx=(5, 0))

        harga_col_pd_frame = tk.Frame(pd_frame, bg='#ecf0f1')
        harga_col_pd_frame.pack(fill='x', pady=2)
        tk.Label(harga_col_pd_frame, text="Harga/colli (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_col_pd_var = tk.StringVar(value=str(barang_data.get('col_pd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_col_pd_entry = tk.Entry(harga_col_pd_frame, textvariable=harga_col_pd_var, font=('Arial', 10), width=20)
        harga_col_pd_entry.pack(side='left', padx=(5, 0))

        # DD Prices
        tk.Label(form_frame, text="Harga Door to Door (DD):", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(anchor='w', pady=(10, 5))

        dd_frame = tk.Frame(form_frame, bg='#ecf0f1')
        dd_frame.pack(fill='x', pady=(0, 10))

        # Similar structure for DD prices
        harga_m3_dd_frame = tk.Frame(dd_frame, bg='#ecf0f1')
        harga_m3_dd_frame.pack(fill='x', pady=2)
        tk.Label(harga_m3_dd_frame, text="Harga/m³ (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_m3_dd_var = tk.StringVar(value=str(barang_data.get('m3_dd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_m3_dd_entry = tk.Entry(harga_m3_dd_frame, textvariable=harga_m3_dd_var, font=('Arial', 10), width=20)
        harga_m3_dd_entry.pack(side='left', padx=(5, 0))

        harga_ton_dd_frame = tk.Frame(dd_frame, bg='#ecf0f1')
        harga_ton_dd_frame.pack(fill='x', pady=2)
        tk.Label(harga_ton_dd_frame, text="Harga/ton (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_ton_dd_var = tk.StringVar(value=str(barang_data.get('ton_dd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_ton_dd_entry = tk.Entry(harga_ton_dd_frame, textvariable=harga_ton_dd_var, font=('Arial', 10), width=20)
        harga_ton_dd_entry.pack(side='left', padx=(5, 0))

        harga_col_dd_frame = tk.Frame(dd_frame, bg='#ecf0f1')
        harga_col_dd_frame.pack(fill='x', pady=2)
        tk.Label(harga_col_dd_frame, text="Harga/colli (Rp):", font=('Arial', 10), bg='#ecf0f1', width=15, anchor='w').pack(side='left')
        harga_col_dd_var = tk.StringVar(value=str(barang_data.get('col_dd', '') or '').replace('Rp ', '').replace(',', '') or '-')
        harga_col_dd_entry = tk.Entry(harga_col_dd_frame, textvariable=harga_col_dd_var, font=('Arial', 10), width=20)
        harga_col_dd_entry.pack(side='left', padx=(5, 0))
        
        def validate_update_form():
            """Validate update form data"""
            
            # 1. Pengirim dan Penerima wajib
            if not pengirim_var.get().strip():
                messagebox.showwarning("Peringatan", "Pengirim tidak boleh kosong.")
                pengirim_var.focus()
                return False
                
            if not penerima_var.get().strip():
                messagebox.showwarning("Peringatan", "Penerima tidak boleh kosong.")
                penerima_var.focus()
                return False
            
            # 2. Nama Barang wajib
            if not nama_barang_var.get().strip():
                messagebox.showwarning("Peringatan", "Nama Barang tidak boleh kosong.")
                nama_barang_entry.focus()
                return False
            
            # 3. Validasi format angka untuk dimensi (jika diisi)
            try:
                if panjang_var.get().strip():
                    value = panjang_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = float(value)
                        if val <= 0:
                            raise ValueError("Panjang harus lebih besar dari 0")
            
                if lebar_var.get().strip():
                    value = lebar_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = float(value)
                        if val <= 0:
                            raise ValueError("Lebar harus lebih besar dari 0")
            
                if tinggi_var.get().strip():
                    value = tinggi_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = float(value)
                        if val <= 0:
                            raise ValueError("Tinggi harus lebih besar dari 0")
                        
            except ValueError as e:
                messagebox.showwarning("Format Tidak Valid", f"Dimensi tidak valid: {str(e)}")
                return False

            # 4. Validasi volume, berat, colli (jika diisi)
            try:
                if volume_var.get().strip():
                    value = volume_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = float(value)
                        if val <= 0:
                            raise ValueError("Volume harus lebih besar dari 0")
            
                if berat_var.get().strip():
                    value = berat_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = float(value)
                        if val <= 0:
                            raise ValueError("Berat harus lebih besar dari 0")
            
                if colli_var.get().strip():
                    value = colli_var.get().strip()
                    if value == '-':
                        pass  # Skip validasi untuk nilai '-'
                    else:
                        val = int(float(value))
                        if val <= 0:
                            raise ValueError("Colli harus lebih besar dari 0")
                        
            except ValueError as e:
                messagebox.showwarning("Format Tidak Valid", f"Volume/Berat/Colli tidak valid: {str(e)}")
                return False
            
            # 5. Validasi harga - minimal salah satu kategori harus diisi
            all_prices = [
                harga_m3_pp_var.get().strip(), harga_ton_pp_var.get().strip(), harga_col_pp_var.get().strip(),
                harga_m3_pd_var.get().strip(), harga_ton_pd_var.get().strip(), harga_col_pd_var.get().strip(),
                harga_m3_dd_var.get().strip(), harga_ton_dd_var.get().strip(), harga_col_dd_var.get().strip()
            ]
            
            if not any(all_prices):
                messagebox.showwarning(
                    "Peringatan", 
                    "Minimal salah satu harga harus diisi!\n\n" +
                    "💰 Pilihan pricing:\n" +
                    "• Harga per m³ (untuk volume)\n" +
                    "• Harga per ton (untuk berat)\n" +
                    "• Harga per colli (untuk jumlah kemasan)\n\n" +
                    "Dan minimal salah satu kategori (PP/PD/DD)"
                )
                return False
            
            # 6. Validasi format harga yang diisi
            try:
                for price_var, name in [
                    (harga_m3_pp_var, "Harga M3 PP"), (harga_ton_pp_var, "Harga Ton PP"), (harga_col_pp_var, "Harga Col PP"),
                    (harga_m3_pd_var, "Harga M3 PD"), (harga_ton_pd_var, "Harga Ton PD"), (harga_col_pd_var, "Harga Col PD"),
                    (harga_m3_dd_var, "Harga M3 DD"), (harga_ton_dd_var, "Harga Ton DD"), (harga_col_dd_var, "Harga Col DD")
                ]:
                    value = price_var.get().strip()
                    if value:
                        if value == '-':
                            continue  # Skip validasi untuk nilai '-'
                    
                        try:
                            val = float(value)
                            if val <= 0:
                                raise ValueError(f"{name} harus lebih besar dari 0")
                        except ValueError:
                            raise ValueError(f"Format {name} tidak valid: '{value}' (gunakan angka atau '-')")
                                    
            except ValueError as e:
                messagebox.showwarning("Format Error", str(e))
                return False
            
            return True
        
        def on_save():
            """Save updated barang data with validation"""
            
            if not validate_update_form():
                return
            
            pengirim_id = self.db.get_customer_id_by_name(pengirim_var.get().strip())
            penerima_id = self.db.get_customer_id_by_name(penerima_var.get().strip())
            
            def process_value(value):
                """Convert '-' to None, otherwise return stripped value or None if empty"""
                stripped = value.strip() if value else ''
                if stripped == '-' or stripped == '':
                    return None
                return stripped
            
            updated_barang = {
                'barang_id': barang_data['barang_id'],
                'pengirim': pengirim_id,
                'penerima': penerima_id,
                'nama_barang': nama_barang_var.get().strip(),
                'panjang_barang': process_value(panjang_var.get()),
                'lebar_barang': process_value(lebar_var.get()),
                'tinggi_barang': process_value(tinggi_var.get()),
                'm3_barang': process_value(volume_var.get()),
                'ton_barang': process_value(berat_var.get()),
                'm3_pp': process_value(harga_m3_pp_var.get()),
                'ton_pp': process_value(harga_ton_pp_var.get()),
                'col_pp': process_value(harga_col_pp_var.get()),
                'm3_pd': process_value(harga_m3_pd_var.get()),
                'ton_pd': process_value(harga_ton_pd_var.get()),
                'col_pd': process_value(harga_col_pd_var.get()),
                'm3_dd': process_value(harga_m3_dd_var.get()),
                'ton_dd': process_value(harga_ton_dd_var.get()),
                'col_dd': process_value(harga_col_dd_var.get()),
            }
            self.save_changes(updated_barang)
            update_window.destroy()

        # Buttons
        btn_frame = tk.Frame(form_frame, bg='#ecf0f1')
        btn_frame.pack(fill='x', pady=30)
        
        # Save button
        btn_save = tk.Button(
            btn_frame, 
            text="💾 Simpan Perubahan", 
            bg="#2ecc71", 
            fg="white", 
            font=('Arial', 11, 'bold'),
            padx=20, 
            pady=8, 
            command=on_save
        )
        btn_save.pack(side="right", padx=10)

        # Cancel button
        btn_cancel = tk.Button(
            btn_frame, 
            text="❌ Batal", 
            bg="#e74c3c", 
            fg="white", 
            font=('Arial', 11, 'bold'),
            padx=20, 
            pady=8, 
            command=lambda: update_window.destroy()
        )
        btn_cancel.pack(side="right")
        
        # Bind mouse wheel to canvas for scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Clean up when window is closed
        def on_closing():
            canvas.unbind_all("<MouseWheel>")
            update_window.destroy()
        
        update_window.protocol("WM_DELETE_WINDOW", on_closing)
        
      
    def on_tree_select(self, event):
        """Handle tree selection change"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            barang_id = item['values'][0]
            nama_barang = item['values'][2]
            self.info_label.config(text=f"✅ Terpilih: {nama_barang} (ID: {barang_id})")
        else:
            self.info_label.config(text="💡 Pilih barang dari tabel untuk edit/hapus")
    
    def delete_barang(self):
        """Delete selected barang"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih barang yang akan dihapus dari tabel!")
            return
        
        # Get selected item data
        item = self.tree.item(selection[0])
        barang_id = item['values'][0]
        nama_barang = item['values'][2]
        
        # Confirm deletion
        if not messagebox.askyesno(
            "Konfirmasi Hapus", 
            f"Yakin ingin menghapus barang?\n\n" +
            f"ID: {barang_id}\n" +
            f"Nama: {nama_barang}\n\n" +
            f"⚠️ Aksi ini tidak dapat dibatalkan!"
        ):
            return
        
        try:
            # Delete from database
            self.db.delete_barang(barang_id)
            
            messagebox.showinfo("Sukses", f"Barang '{nama_barang}' berhasil dihapus!")
            
            # Refresh data
            self.load_barang()
            if self.refresh_callback:
                self.refresh_callback()
                
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menghapus barang:\n{str(e)}")
    
    def export_barang(self):
        """Export barang data to Excel - FIXED for pengirim-penerima system"""
        try:
            if not self.original_barang_data:
                messagebox.showwarning("Peringatan", "Tidak ada data barang untuk diekspor!")
                return
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                parent=self.window,
                title="Export Data Barang ke Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"data_barang_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Prepare data for export - UPDATED structure
            export_data = []
            for barang in self.original_barang_data:
                # Format dimensions
                dimensi = f"{barang.get('panjang_barang', '-')}×{barang.get('lebar_barang', '-')}×{barang.get('tinggi_barang', '-')}"
                
                export_data.append({
                    'ID': barang.get('barang_id', ''),
                    'Pengirim': barang.get('sender_name', ''),  # Updated field name
                    'Penerima': barang.get('receiver_name', ''), # Updated field name
                    'Nama Barang': barang.get('nama_barang', ''),
                    'Dimensi (P×L×T cm)': dimensi,
                    'Panjang (cm)': barang.get('panjang_barang', ''),
                    'Lebar (cm)': barang.get('lebar_barang', ''),
                    'Tinggi (cm)': barang.get('tinggi_barang', ''),
                    'Volume (m³)': barang.get('m3_barang', ''),
                    'Berat (ton)': barang.get('ton_barang', ''),
                    
                    # All pricing fields
                    'Harga M³ PP (Rp)': barang.get('m3_pp', ''),
                    'Harga M³ PD (Rp)': barang.get('m3_pd', ''),
                    'Harga M³ DD (Rp)': barang.get('m3_dd', ''),
                    'Harga Ton PP (Rp)': barang.get('ton_pp', ''),
                    'Harga Ton PD (Rp)': barang.get('ton_pd', ''),
                    'Harga Ton DD (Rp)': barang.get('ton_dd', ''),
                    'Harga Colli PP (Rp)': barang.get('col_pp', ''),
                    'Harga Colli PD (Rp)': barang.get('col_pd', ''),
                    'Harga Colli DD (Rp)': barang.get('col_dd', ''),
                    
                    'Tanggal Dibuat': barang.get('created_at', '')[:19] if barang.get('created_at') else ''
                })
            
            # Create DataFrame and export
            df = pd.DataFrame(export_data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data Barang')
                
                # Format the Excel file
                try:
                    workbook = writer.book
                    worksheet = writer.sheets['Data Barang']
                    
                    # Style headers
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    
                    for col_num, column_title in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                except Exception as styling_error:
                    print(f"Warning: Could not apply Excel styling: {styling_error}")
                    # File will still be created without styling
            
            messagebox.showinfo(
                "Export Berhasil",
                f"Data barang berhasil diekspor ke:\n{filename}\n\n" +
                f"📊 Total: {len(export_data)} barang\n" +
                f"📋 Kolom: Pengirim, Penerima, Nama Barang, Dimensi, Harga lengkap"
            )
            
        except Exception as e:
            error_msg = f"Gagal export data: {str(e)}"
            print(f"Export error: {error_msg}")
            messagebox.showerror("Error", error_msg)
        
        
    def show_error_details(self, errors, customer_not_found_list, success_count, total_count):
        """Show detailed error modal with proper horizontal scrolling"""
        error_window = tk.Toplevel(self.window)
        error_window.title("📊 Detail Error Upload")
        error_window.geometry("1200x700")  # ✅ Lebih lebar lagi
        error_window.configure(bg='#ecf0f1')
        error_window.transient(self.window)
        error_window.grab_set()
        
        # Center the error window
        error_window.update_idletasks()
        x = (error_window.winfo_screenwidth() // 2) - (1200 // 2)
        y = (error_window.winfo_screenheight() // 2) - (700 // 2)
        error_window.geometry(f"1200x700+{x}+{y}")
        
        # Header
        header = tk.Label(
            error_window,
            text="📊 DETAIL HASIL UPLOAD",
            font=('Arial', 16, 'bold'),
            bg='#e74c3c',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Summary frame
        summary_frame = tk.Frame(error_window, bg='#ffffff', relief='solid', bd=1)
        summary_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        summary_text = f"""📈 RINGKASAN UPLOAD:
        
    ✅ Berhasil: {success_count} dari {total_count} barang
    ❌ Gagal: {len(errors)} barang  
    ⚠️ Customer tidak ditemukan: {len(customer_not_found_list)} barang
        
    📋 Total diproses: {total_count} baris data"""
        
        summary_label = tk.Label(
            summary_frame,
            text=summary_text,
            font=('Arial', 12),
            fg='#2c3e50',
            bg='#ffffff',
            justify='left',
            padx=20,
            pady=15
        )
        summary_label.pack(fill='x')
        
        # Create notebook for error details
        notebook = ttk.Notebook(error_window)
        notebook.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Tab 1: Processing Errors with PROPER scrolling
        if errors:
            error_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(error_frame, text=f'❌ Error Processing ({len(errors)})')
            
            error_label = tk.Label(
                error_frame,
                text="Barang yang gagal diproses karena error teknis:",
                font=('Arial', 12, 'bold'),
                fg='#e74c3c',
                bg='#ecf0f1'
            )
            error_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            # ✅ Main frame untuk treeview
            error_main_frame = tk.Frame(error_frame, bg='#ecf0f1')
            error_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            # ✅ Create treeview dengan scrollbar yang benar
            error_tree = ttk.Treeview(error_main_frame, 
                                    columns=('No', 'Nama Barang', 'Customer', 'Error'), 
                                    show='headings', height=12)
            
            # Configure headers
            error_tree.heading('No', text='No')
            error_tree.heading('Nama Barang', text='Nama Barang')
            error_tree.heading('Customer', text='Customer')  
            error_tree.heading('Error', text='Detail Error')
            
            # ✅ Column widths yang lebih reasonable
            error_tree.column('No', width=50, minwidth=40, stretch=False)
            error_tree.column('Nama Barang', width=200, minwidth=150, stretch=False)
            error_tree.column('Customer', width=150, minwidth=120, stretch=False)
            error_tree.column('Error', width=600, minwidth=300, stretch=True)
            
            # Create scrollbars
            error_v_scroll = ttk.Scrollbar(error_main_frame, orient='vertical', command=error_tree.yview)
            error_h_scroll = ttk.Scrollbar(error_main_frame, orient='horizontal', command=error_tree.xview)
            
            # Configure treeview scrollbars
            error_tree.configure(yscrollcommand=error_v_scroll.set, xscrollcommand=error_h_scroll.set)
            
            # ✅ GRID LAYOUT untuk scrollbar yang proper
            error_tree.grid(row=0, column=0, sticky='nsew')
            error_v_scroll.grid(row=0, column=1, sticky='ns')
            error_h_scroll.grid(row=1, column=0, sticky='ew')
            
            # ✅ Configure grid weights
            error_main_frame.grid_rowconfigure(0, weight=1)
            error_main_frame.grid_columnconfigure(0, weight=1)
            
            # Add error data
            for i, error_info in enumerate(errors, 1):
                error_tree.insert('', tk.END, values=(
                    i,
                    error_info.get('nama_barang', 'N/A'),
                    error_info.get('customer', 'N/A'),
                    error_info.get('error', 'Unknown error')  # ✅ Error message bisa panjang
                ))
        
        # Tab 2: Customer Not Found dengan scrolling yang sama
        if customer_not_found_list:
            customer_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(customer_frame, text=f'⚠️ Customer Tidak Ditemukan ({len(customer_not_found_list)})')
            
            customer_label = tk.Label(
                customer_frame,
                text="Barang yang gagal karena customer belum terdaftar di sistem:",
                font=('Arial', 12, 'bold'),
                fg='#f39c12',
                bg='#ecf0f1'
            )
            customer_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            # ✅ Customer treeview dengan grid layout
            customer_main_frame = tk.Frame(customer_frame, bg='#ecf0f1')
            customer_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            customer_tree = ttk.Treeview(customer_main_frame, 
                                    columns=('No', 'Nama Barang', 'Customer yang Dicari'), 
                                    show='headings', height=12)
            
            customer_tree.heading('No', text='No')
            customer_tree.heading('Nama Barang', text='Nama Barang')
            customer_tree.heading('Customer yang Dicari', text='Customer yang Dicari')
            
            customer_tree.column('No', width=50, minwidth=40, stretch=False)
            customer_tree.column('Nama Barang', width=300, minwidth=200, stretch=False)
            customer_tree.column('Customer yang Dicari', width=300, minwidth=200, stretch=True)
            
            # Scrollbars
            customer_v_scroll = ttk.Scrollbar(customer_main_frame, orient='vertical', command=customer_tree.yview)
            customer_h_scroll = ttk.Scrollbar(customer_main_frame, orient='horizontal', command=customer_tree.xview)
            customer_tree.configure(yscrollcommand=customer_v_scroll.set, xscrollcommand=customer_h_scroll.set)
            
            # ✅ Grid layout
            customer_tree.grid(row=0, column=0, sticky='nsew')
            customer_v_scroll.grid(row=0, column=1, sticky='ns')
            customer_h_scroll.grid(row=1, column=0, sticky='ew')
            
            customer_main_frame.grid_rowconfigure(0, weight=1)
            customer_main_frame.grid_columnconfigure(0, weight=1)
            
            # Add customer not found data
            for i, customer_info in enumerate(customer_not_found_list, 1):
                customer_tree.insert('', tk.END, values=(
                    i,
                    customer_info.get('nama_barang', 'N/A'),
                    customer_info.get('customer', 'N/A')
                ))
            
            # Instruction
            instruction_frame = tk.Frame(customer_frame, bg='#fff3cd', relief='solid', bd=1)
            instruction_frame.pack(fill='x', padx=10, pady=10)
            
            instruction_text = tk.Label(
                instruction_frame,
                text="💡 Solusi: Tambahkan customer yang belum terdaftar melalui menu Customer, " +
                    "kemudian upload ulang file Excel ini.",
                font=('Arial', 10),
                fg='#856404',
                bg='#fff3cd',
                wraplength=1000,
                justify='left',
                padx=15,
                pady=10
            )
            instruction_text.pack()
        
        # Tab 3: Tips (sama seperti sebelumnya)
        tips_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(tips_frame, text='💡 Tips Upload')
        
        # ... (kode tips sama seperti sebelumnya)
        
        # Close button
        close_btn = tk.Button(
            error_window,
            text="✅ Tutup",
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=error_window.destroy
        )
        close_btn.pack(pady=20)
    
    def center_window(self):
        """Center window on parent"""
        self.window.update_idletasks()
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        # Ubah ukuran di sini
        window_width = 1400  # dari 1400
        window_height = 850  # dari 800
        
        x = parent_x + (parent_width // 2) - (window_width // 2)
        y = parent_y + (parent_height // 2) - (window_height // 2) - 50
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    def load_customer_combo(self):
        """Load customers into combobox"""
        customers = self.db.get_all_customers()
        customer_list = [f"{c['nama_customer']}" for c in customers]
        self.penerima_combo['values'] = customer_list
        self.original_penerima_data = customer_list

    def load_pengirim_combo(self):
        """Load senders into combobox"""
        senders = self.db.get_all_customers()
        sender_list = [f"{s['nama_customer']}" for s in senders]
        self.pengirim_combo['values'] = sender_list
        self.original_pengirim_data = sender_list

    def filter_pengirim(self, event):
        """Filter pengirim combobox saat user mengetik"""
        typed = self.pengirim_var.get().lower()
        
        if not hasattr(self, 'original_pengirim_values'):
            # Simpan nilai asli jika belum ada
            self.original_pengirim_values = list(self.pengirim_combo['values'])
        
        if typed == '':
            # Jika kosong, tampilkan semua
            self.pengirim_combo['values'] = self.original_pengirim_values
        else:
            # Filter berdasarkan yang diketik
            filtered = [item for item in self.original_pengirim_values 
                    if typed in item.lower()]
            self.pengirim_combo['values'] = filtered
        
        # Buka dropdown untuk menampilkan hasil filter
        # self.pengirim_combo.event_generate('<Button-1>')

    def filter_pengirim_list(self, event):
        typed = self.filter_pengirim_var.get().lower()
        
        if not hasattr(self, 'original_pengirim_values'):
            # Simpan nilai asli jika belum ada
            self.original_pengirim_values = list(self.filter_pengirim_combo['values'])
        
        if typed == '':
            # Jika kosong, tampilkan semua
            self.filter_pengirim_combo['values'] = self.original_pengirim_values
        else:
            # Filter berdasarkan yang diketik
            filtered = [item for item in self.original_pengirim_values 
                    if typed in item.lower()]
            self.filter_pengirim_combo['values'] = filtered
        
        # Buka dropdown untuk menampilkan hasil filter
        # self.pengirim_combo.event_generate('<Button-1>')

    def filter_penerima(self, event):
        """Filter penerima combobox saat user mengetik"""
        typed = self.penerima_var.get().lower()
        
        if not hasattr(self, 'original_penerima_values'):
            # Simpan nilai asli jika belum ada
            self.original_penerima_values = list(self.penerima_combo['values'])
        
        if typed == '':
            # Jika kosong, tampilkan semua
            self.penerima_combo['values'] = self.original_penerima_values
        else:
            # Filter berdasarkan yang diketik
            filtered = [item for item in self.original_penerima_values 
                    if typed in item.lower()]
            self.penerima_combo['values'] = filtered
        
        # Buka dropdown untuk menampilkan hasil filter
        # self.penerima_combo.event_generate('<Button-1>')

    def filter_penerima_list(self, event):
        """Filter penerima combobox saat user mengetik"""
        typed = self.filter_penerima_var.get().lower()
        
        if not hasattr(self, 'original_penerima_values'):
            # Simpan nilai asli jika belum ada
            self.original_penerima_values = list(self.filter_penerima_combo['values'])
        
        if typed == '':
            # Jika kosong, tampilkan semua
            self.filter_penerima_combo['values'] = self.original_penerima_values
        else:
            # Filter berdasarkan yang diketik
            filtered = [item for item in self.original_penerima_values 
                    if typed in item.lower()]
            self.filter_penerima_combo['values'] = filtered
        
        # Buka dropdown untuk menampilkan hasil filter
        # self.penerima_combo.event_generate('<Button-1>')

    def browse_file(self):
        """Browse for Excel file"""
        file_types = [
            ('Excel files', '*.xlsx *.xls'),
            ('All files', '*.*')
        ]
        
        filename = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=file_types,
            parent=self.window
        )
        
        if filename:
            self.file_path_var.set(filename)
            self.preview_excel_file(filename)
    
    def preview_excel_file(self, filename):
        """Preview Excel file content with enhanced field mapping for Pengirim-Penerima system"""
        try:
            self.status_label.config(text="📄 Membaca file Excel...", fg='#3498db')
            
            # Clear previous preview
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            # Read Excel file
            df = pd.read_excel(filename, engine='openpyxl')
            print(f"📋 Columns found: {list(df.columns)}")
            
            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            
            # ✅ UPDATED: Enhanced column mapping for Pengirim-Penerima system
            column_mapping = {
                # Required fields
                'pengirim': ['pengirim', 'sender', 'from', 'asal', 'customer_asal'],
                'penerima': ['penerima', 'receiver', 'to', 'tujuan', 'customer_tujuan'],
                'nama_barang': ['nama barang', 'nama_barang', 'barang', 'product', 'nama', 'item'],
                
                # Optional fields  
                'panjang': ['p', 'panjang', 'length'],
                'lebar': ['l', 'lebar', 'width'], 
                'tinggi': ['t', 'tinggi', 'height'],
                'm3': ['m3', 'volume', 'vol'],
                'ton': ['ton', 'berat', 'weight'],
                
                # ✅ NEW: All pricing fields (PP, PD, DD)
                'harga_m3_pp': ['m3_pp', 'harga_m3_pp', 'harga m3 pp', 'm³ p→p', 'm3 pelabuhan-pelabuhan'],
                'harga_m3_pd': ['m3_pd', 'harga_m3_pd', 'harga m3 pd', 'm³ p→d', 'm3 pelabuhan-door'],
                'harga_m3_dd': ['m3_dd', 'harga_m3_dd', 'harga m3 dd', 'm³ d→d', 'm3 door-door'],
                
                'harga_ton_pp': ['ton_pp', 'harga_ton_pp', 'harga ton pp', 'ton p→p', 'ton pelabuhan-pelabuhan'],
                'harga_ton_pd': ['ton_pd', 'harga_ton_pd', 'harga ton pd', 'ton p→d', 'ton pelabuhan-door'],
                'harga_ton_dd': ['ton_dd', 'harga_ton_dd', 'harga ton dd', 'ton d→d', 'ton door-door'],
                
                'harga_col_pp': ['colli_pp', 'harga_col_pp', 'harga colli pp', 'colli p→p', 'colli pelabuhan-pelabuhan'],
                'harga_col_pd': ['colli_pd', 'harga_col_pd', 'harga colli pd', 'colli p→d', 'colli pelabuhan-door'],
                'harga_col_dd': ['colli_dd', 'harga_col_dd', 'harga colli dd', 'colli d→d', 'colli door-door'],
                
                # Legacy fallback fields
                'harga_m3': ['harga/m3', 'harga per m3', 'harga_m3', 'price_m3'],
                'harga_ton': ['harga/ton', 'harga per ton', 'harga_ton', 'price_ton'],
                'harga_coll': ['harga/col', 'harga per coll', 'harga_coll', 'price_coll', 'harga/colli'],
                'harga': ['harga', 'price']  # Generic price field as fallback
            }
            
            # ✅ IMPROVED: Enhanced column finding algorithm
            found_columns = {}
            used_columns = set()
            
            # First pass: Look for exact matches (highest priority)
            for field, possible_names in column_mapping.items():
                best_match = None
                best_score = 0
                
                for col in df.columns:
                    print(f"Matching for field '{field}': checking column '{col}'")
                    
                    if col in used_columns:
                        continue
                        
                    col_lower = col.lower().strip()
                    
                    # Calculate match score
                    for possible_name in possible_names:
                        possible_lower = possible_name.lower()
                        
                        if col_lower == possible_lower:  # Exact match
                            best_match = col
                            best_score = 100
                            break
                        elif possible_lower in col_lower:  # Partial match
                            # Prefer shorter matches and those that start with the pattern
                            score = len(possible_lower) / len(col_lower) * 50
                            if col_lower.startswith(possible_lower):
                                score += 25
                            # Bonus for service type indicators (PP, PD, DD)
                            if any(indicator in col_lower for indicator in ['pp', 'pd', 'dd', 'p→p', 'p→d', 'd→d']):
                                score += 15
                            if score > best_score:
                                best_match = col
                                best_score = score
                    
                    if best_score >= 100:  # Perfect match found
                        break
                
                if best_match and best_score >= 25:  # Minimum threshold
                    found_columns[field] = best_match
                    used_columns.add(best_match)
                    print(f"🎯 Mapped '{field}' to '{best_match}' (score: {best_score:.1f})")
            
            print(f"🎯 Found columns mapping: {found_columns}")
            
            # ✅ UPDATED: Check required columns for new system
            required_fields = ['pengirim', 'penerima', 'nama_barang']
            missing_fields = [field for field in required_fields if field not in found_columns]
            
            if missing_fields:
                available_cols = ', '.join(df.columns.tolist())
                self.status_label.config(
                    text=f"❌ Kolom wajib tidak ditemukan: {missing_fields}\n\nKolom tersedia: {available_cols}", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            # ✅ UPDATED: Preview data with pengirim-penerima validation
            valid_rows = df.dropna(subset=[found_columns['pengirim'], found_columns['penerima'], found_columns['nama_barang']])
            preview_data = valid_rows.head(50)
            
            # Get existing customers for validation
            existing_customers = {c['nama_customer'].upper(): c['customer_id'] for c in self.db.get_all_customers()}

            preview_count = 0
            pengirim_errors = set()
            penerima_errors = set()
            
            for _, row in preview_data.iterrows():
                pengirim = str(row[found_columns['pengirim']]).strip()
                penerima = str(row[found_columns['penerima']]).strip()
                nama_barang = str(row[found_columns['nama_barang']]).strip()
                
                # ✅ ENHANCED: Better field value extraction
                def get_field_value(field_name, default=''):
                    print(f"Extracting field '{field_name}'")
                    if field_name in found_columns:
                        value = row.get(found_columns[field_name], default)
                        print(f"  Raw value: {value}")
                        if pd.isna(value) or str(value).strip().lower() in ['nan', 'none', '']:
                            return ''
                        return str(value).strip()
                    return default
                
                # Get basic fields
                panjang = get_field_value('panjang')
                lebar = get_field_value('lebar') 
                tinggi = get_field_value('tinggi')
                m3 = get_field_value('m3')
                ton = get_field_value('ton')
                
                # ✅ NEW: Get all pricing fields
                harga_m3_pp = get_field_value('harga_m3_pp')
                harga_m3_pd = get_field_value('harga_m3_pd')
                harga_m3_dd = get_field_value('harga_m3_dd')
                harga_ton_pp = get_field_value('harga_ton_pp')
                harga_ton_pd = get_field_value('harga_ton_pd')
                harga_ton_dd = get_field_value('harga_ton_dd')
                harga_col_pp = get_field_value('harga_col_pp')
                harga_col_pd = get_field_value('harga_col_pd')
                harga_col_dd = get_field_value('harga_col_dd')
                
                # Fallback to legacy pricing if new fields not found
                if not any([harga_m3_pp, harga_m3_pd, harga_m3_dd]):
                    legacy_harga_m3 = get_field_value('harga_m3')
                    if legacy_harga_m3:
                        harga_m3_pp = legacy_harga_m3  # Default to PP
                
                if not any([harga_ton_pp, harga_ton_pd, harga_ton_dd]):
                    legacy_harga_ton = get_field_value('harga_ton')
                    if legacy_harga_ton:
                        harga_ton_pp = legacy_harga_ton
                
                if not any([harga_col_pp, harga_col_pd, harga_col_dd]):
                    legacy_harga_coll = get_field_value('harga_coll')
                    if legacy_harga_coll:
                        harga_col_pp = legacy_harga_coll
                
                print(f"Previewing row: {row.name}")
                print(f"Pengirim: {pengirim}, Penerima: {penerima}, Nama Barang: {nama_barang}")
                print(f"Pricing - M3(PP/PD/DD): {harga_m3_pp}/{harga_m3_pd}/{harga_m3_dd}")
                print(f"Pricing - TON(PP/PD/DD): {harga_ton_pp}/{harga_ton_pd}/{harga_ton_dd}")
                print(f"Pricing - COL(PP/PD/DD): {harga_col_pp}/{harga_col_pd}/{harga_col_dd}")
                
                if pengirim and penerima and nama_barang:
                    # ✅ UPDATED: Check if pengirim and penerima exist
                    pengirim_exists = pengirim.upper() in existing_customers
                    penerima_exists = penerima.upper() in existing_customers
                    
                    if not pengirim_exists:
                        pengirim_errors.add(pengirim)
                    if not penerima_exists:
                        penerima_errors.add(penerima)
                    
                    pengirim_status = "✅" if pengirim_exists else "❌"
                    penerima_status = "✅" if penerima_exists else "❌"
                    
                    display_pengirim = f"{pengirim_status} {pengirim}"
                    display_penerima = f"{penerima_status} {penerima}"
                    
                    # ✅ ENHANCED: Format currency values
                    def format_currency(value):
                        if value and str(value).strip():
                            try:
                                # Handle Indonesian number format
                                clean_value = str(value).replace(',', '').replace(' ', '')
                                return f"{float(clean_value):,.0f}"
                            except:
                                return value
                        return ''
                    
                    # ✅ UPDATED: Insert with all new columns (matching treeview structure)
                    self.preview_tree.insert('', tk.END, values=(
                        display_pengirim,           # Pengirim
                        display_penerima,           # Penerima  
                        nama_barang,                # Nama
                        panjang,                    # P
                        lebar,                      # L
                        tinggi,                     # T
                        m3,                         # M3
                        ton,                        # Ton
                        format_currency(harga_m3_pp),   # M3_PP
                        format_currency(harga_m3_pd),   # M3_PD
                        format_currency(harga_m3_dd),   # M3_DD
                        format_currency(harga_ton_pp),  # TON_PP
                        format_currency(harga_ton_pd),  # TON_PD
                        format_currency(harga_ton_dd),  # TON_DD
                        format_currency(harga_col_pp),  # COLLI_PP
                        format_currency(harga_col_pd),  # COLLI_PD
                        format_currency(harga_col_dd)   # COLLI_DD
                    ))
                    preview_count += 1
            
            # ✅ UPDATED: Store column mapping for upload
            self.column_mapping = found_columns
            
            # ✅ ENHANCED: Create comprehensive status message
            found_fields = list(found_columns.keys())
            required_fields_found = [f for f in required_fields if f in found_columns]
            optional_fields = [f for f in found_fields if f not in required_fields]
            
            # Categorize optional fields
            basic_fields = [f for f in optional_fields if f in ['panjang', 'lebar', 'tinggi', 'm3', 'ton', 'colli']]
            pricing_fields = [f for f in optional_fields if 'harga' in f]
            
            status_msg = f"✅ File berhasil dibaca: {preview_count} baris data\n\n"
            status_msg += f"📋 Kolom wajib: {', '.join([found_columns[f] for f in required_fields_found])}\n"
            
            if basic_fields:
                status_msg += f"📊 Data barang: {', '.join([found_columns[f] for f in basic_fields])}\n"
            
            if pricing_fields:
                status_msg += f"💰 Harga: {', '.join([found_columns[f] for f in pricing_fields])}\n"
            
            # ✅ UPDATED: Warning for missing customers
            warning_msg = ""
            if pengirim_errors:
                warning_msg += f"⚠️ Pengirim tidak terdaftar: {', '.join(list(pengirim_errors)[:3])}{'...' if len(pengirim_errors) > 3 else ''}\n"
            if penerima_errors:
                warning_msg += f"⚠️ Penerima tidak terdaftar: {', '.join(list(penerima_errors)[:3])}{'...' if len(penerima_errors) > 3 else ''}\n"
            
            if warning_msg:
                status_msg += f"\n{warning_msg}"
                status_msg += "❗ Pastikan customer dengan tanda ❌ sudah terdaftar sebelum upload!"
            
            self.status_label.config(text=status_msg, fg='#27ae60' if not warning_msg else '#f39c12')
            self.upload_btn.config(state='normal')
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"💥 Preview error: {error_detail}")
            
            self.status_label.config(
                text=f"❌ Error membaca file: {str(e)}", 
                fg='#e74c3c'
            )
            self.upload_btn.config(state='disabled')

    def download_template(self):
        """Download Excel template with updated column structure"""
        try:
            # ✅ UPDATED: Create template with new column structure
            template_data = {
                'Pengirim': [
                    'PT. ANEKA PLASTIK INDONESIA',
                    'CV. MAJU BERSAMA',
                    'PT. TEKNOLOGI MODERN'
                ],
                'Penerima': [
                    'PT. TEKNOLOGI MODERN',
                    'UD. SUMBER REZEKI',
                    'PT. GLOBAL TRADING'
                ],
                'Nama Barang': ['TV LED 32 inch', 'Mie Instan 1 Karton', 'Kain Katun 50 Yard'],
                'P': [50, 40, 30],
                'L': [30, 30, 25],
                'T': [10, 25, 20],
                'M3': [0.015, 0.030, 0.015],
                'Ton': [0.012, 0.008, 0.005],
                'M3_PP': [150000, 175000, 200000],
                'M3_PD': [180000, 200000, 230000],
                'M3_DD': [220000, 250000, 280000],
                'TON_PP': [500000, 450000, 600000],
                'TON_PD': [600000, 550000, 700000],
                'TON_DD': [750000, 700000, 850000],
                'COLLI_PP': [25000, 30000, 35000],
                'COLLI_PD': [35000, 40000, 45000],
                'COLLI_DD': [50000, 55000, 60000]
            }
            
            # Create DataFrame
            template_df = pd.DataFrame(template_data)
            
            # Save template
            filename = filedialog.asksaveasfilename(
                parent=self.window,  # ✅ FIXED: Added parent window
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Simpan Template Excel",
                initialfile="template_barang_lengkap.xlsx"
            )
            
            if filename:
                try:
                    # ✅ FIXED: Try different methods to create Excel file
                    # Method 1: Try with openpyxl engine
                    try:
                        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                            template_df.to_excel(writer, sheet_name='Data Barang', index=False)
                            
                            # Add instruction sheet
                            instructions = pd.DataFrame({
                                'PANDUAN PENGGUNAAN TEMPLATE': [
                                    '1. Isi kolom Pengirim dengan nama customer yang sudah terdaftar',
                                    '2. Isi kolom Penerima dengan nama customer yang sudah terdaftar',
                                    '3. Nama Barang wajib diisi',
                                    '4. P, L, T dalam satuan cm',
                                    '5. M3 dalam satuan meter kubik',
                                    '6. Ton dalam satuan ton',
                                    '7. Colli adalah jumlah kemasan',
                                    '8. Harga menggunakan kode:',
                                    '   - PP = Pelabuhan ke Pelabuhan',
                                    '   - PD = Pelabuhan ke Door',
                                    '   - DD = Door ke Door',
                                    '9. Hapus baris contoh ini sebelum upload',
                                    '10. Pastikan tidak ada baris kosong di tengah data'
                                ]
                            })
                            instructions.to_excel(writer, sheet_name='Panduan', index=False)
                            
                    except Exception as openpyxl_error:
                        print(f"⚠️ Openpyxl method failed: {openpyxl_error}")
                        
                        # Method 2: Try with xlsxwriter engine
                        try:
                            with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                                template_df.to_excel(writer, sheet_name='Data Barang', index=False)
                                
                                # Add instruction sheet
                                instructions = pd.DataFrame({
                                    'PANDUAN PENGGUNAAN TEMPLATE': [
                                        '1. Isi kolom Pengirim dengan nama customer yang sudah terdaftar',
                                        '2. Isi kolom Penerima dengan nama customer yang sudah terdaftar',
                                        '3. Nama Barang wajib diisi',
                                        '4. P, L, T dalam satuan cm',
                                        '5. M3 dalam satuan meter kubik',
                                        '6. Ton dalam satuan ton',
                                        '7. Colli adalah jumlah kemasan',
                                        '8. Harga menggunakan kode:',
                                        '   - PP = Pelabuhan ke Pelabuhan',
                                        '   - PD = Pelabuhan ke Door',
                                        '   - DD = Door ke Door',
                                        '9. Hapus baris contoh ini sebelum upload',
                                        '10. Pastikan tidak ada baris kosong di tengah data'
                                    ]
                                })
                                instructions.to_excel(writer, sheet_name='Panduan', index=False)
                                
                        except Exception as xlsxwriter_error:
                            print(f"⚠️ Xlsxwriter method failed: {xlsxwriter_error}")
                            
                            # Method 3: Fallback - save as simple Excel without multiple sheets
                            template_df.to_excel(filename, index=False, sheet_name='Data Barang')
                            print("✅ Template saved using fallback method (single sheet)")
                    
                    messagebox.showinfo(
                        "Template Berhasil Disimpan",
                        f"Template Excel berhasil disimpan:\n{filename}\n\n" +
                        "📋 Template berisi:\n" +
                        "• Sheet 'Data Barang' dengan contoh data\n" +
                        "• Sheet 'Panduan' dengan instruksi penggunaan\n\n" +
                        "Silakan isi data sesuai panduan yang tersedia."
                    )
                    
                except Exception as save_error:
                    raise Exception(f"Gagal menyimpan file Excel: {str(save_error)}")
                    
            else:
                print("👤 User cancelled template download")
                
        except Exception as e:
            error_msg = str(e)
            print(f"💥 Error in download_template: {error_msg}")
            
            # ✅ ENHANCED: Better error handling with suggestions
            if "initialname" in error_msg.lower():
                error_dialog = (
                    "❌ GAGAL MENYIMPAN TEMPLATE\n\n"
                    "Error terkait pengaturan nama file.\n\n"
                    "💡 SOLUSI:\n"
                    "• Pastikan lokasi penyimpanan dapat diakses\n"
                    "• Coba simpan dengan nama file yang berbeda\n"
                    "• Periksa izin folder tujuan"
                )
            elif "engine" in error_msg.lower() or "openpyxl" in error_msg.lower():
                error_dialog = (
                    "❌ GAGAL MENYIMPAN TEMPLATE\n\n"
                    "Error terkait engine Excel.\n\n"
                    "💡 SOLUSI:\n"
                    "• Install: pip install openpyxl\n"
                    "• Install: pip install xlsxwriter\n"
                    "• Restart aplikasi setelah install"
                )
            else:
                error_dialog = (
                    f"❌ GAGAL MENYIMPAN TEMPLATE\n\n"
                    f"Error: {error_msg}\n\n"
                    f"💡 SOLUSI:\n"
                    f"• Pastikan folder tujuan dapat diakses\n"
                    f"• Tutup Excel jika sedang membuka file yang sama\n"
                    f"• Coba lokasi penyimpanan yang berbeda"
                )
            
            messagebox.showerror("Error Download Template", error_dialog)     
           
    def validate_excel_row(self, row_data, column_mapping, existing_customers, row_index):
        """Validate single row from Excel data - UPDATED for Pengirim-Penerima system"""
        errors = []
        
        try:
            # ✅ UPDATED: 1. Validasi Pengirim
            pengirim_name = str(row_data.get(column_mapping.get('pengirim', ''), '')).strip()
            pengirim_id = None
            
            if not pengirim_name or pengirim_name.upper() == 'NAN':
                errors.append("Pengirim tidak boleh kosong")
            else:
                pengirim_id = existing_customers.get(pengirim_name.upper())
                if not pengirim_id:
                    errors.append(f"Pengirim '{pengirim_name}' tidak ditemukan di database")
            
            # ✅ UPDATED: 2. Validasi Penerima
            penerima_name = str(row_data.get(column_mapping.get('penerima', ''), '')).strip()
            penerima_id = None
            
            if not penerima_name or penerima_name.upper() == 'NAN':
                errors.append("Penerima tidak boleh kosong")
            else:
                penerima_id = existing_customers.get(penerima_name.upper())
                if not penerima_id:
                    errors.append(f"Penerima '{penerima_name}' tidak ditemukan di database")
            
            # 3. Validasi Nama Barang (sama seperti sebelumnya)
            nama_barang = str(row_data.get(column_mapping.get('nama_barang', ''), '')).strip()
            if not nama_barang or nama_barang.upper() == 'NAN':
                errors.append("Nama Barang tidak boleh kosong")
            
            # 4. Validasi Dimensi (sama seperti sebelumnya, tapi lebih robust)
            dimension_fields = {
                'panjang': 'Panjang',
                'lebar': 'Lebar', 
                'tinggi': 'Tinggi'
            }
            
            for field, field_name in dimension_fields.items():
                if field in column_mapping:
                    value = row_data.get(column_mapping[field])
                    if pd.notna(value) and str(value).strip() != '' and str(value).upper() != 'NAN':
                        try:
                            # Handle Indonesian number format (comma as thousand separator)
                            clean_value = str(value).replace(',', '').replace(' ', '')
                            float_val = float(clean_value)
                            if float_val <= 0:
                                errors.append(f"{field_name} harus lebih besar dari 0")
                        except (ValueError, TypeError):
                            errors.append(f"Format {field_name} tidak valid: '{value}' (gunakan angka)")
            
            # 5. Validasi Volume, Berat, Colli (enhanced)
            measurement_fields = {
                'm3': ('Volume (M³)', 'float'),
                'ton': ('Berat (Ton)', 'float'),
                'colli': ('Colli', 'int')
            }
            
            for field, (field_name, data_type) in measurement_fields.items():
                if field in column_mapping:
                    value = row_data.get(column_mapping[field])
                    if pd.notna(value) and str(value).strip() != '' and str(value).upper() != 'NAN':
                        try:
                            clean_value = str(value).replace(',', '').replace(' ', '')
                            if data_type == 'int':
                                int_val = int(float(clean_value))
                                if int_val <= 0:
                                    errors.append(f"{field_name} harus lebih besar dari 0")
                            else:
                                float_val = float(clean_value)
                                if float_val <= 0:
                                    errors.append(f"{field_name} harus lebih besar dari 0")
                        except (ValueError, TypeError):
                            errors.append(f"Format {field_name} tidak valid: '{value}' (gunakan angka)")
            
            # ✅ UPDATED: 6. Validasi Harga - Support all pricing fields (PP/PD/DD)
            pricing_fields = {
                # M3 pricing
                'harga_m3_pp': 'Harga M³ (Pelabuhan-Pelabuhan)',
                'harga_m3_pd': 'Harga M³ (Pelabuhan-Door)',
                'harga_m3_dd': 'Harga M³ (Door-Door)',
                
                # TON pricing  
                'harga_ton_pp': 'Harga Ton (Pelabuhan-Pelabuhan)',
                'harga_ton_pd': 'Harga Ton (Pelabuhan-Door)',
                'harga_ton_dd': 'Harga Ton (Door-Door)',
                
                # COLLI pricing
                'harga_col_pp': 'Harga Colli (Pelabuhan-Pelabuhan)',
                'harga_col_pd': 'Harga Colli (Pelabuhan-Door)',
                'harga_col_dd': 'Harga Colli (Door-Door)',
                
                # Legacy pricing (backward compatibility)
                'harga_m3': 'Harga M³',
                'harga_ton': 'Harga Ton',
                'harga_coll': 'Harga Colli'
            }
            
            has_valid_price = False
            price_errors = []
            
            for field, field_name in pricing_fields.items():
                if field in column_mapping:
                    value = row_data.get(column_mapping[field])
                    if pd.notna(value) and str(value).strip() != '' and str(value).upper() != 'NAN':
                        try:
                            # Handle Indonesian currency format
                            clean_value = str(value).replace(',', '').replace(' ', '').replace('Rp', '').replace('.', '')
                            price = float(clean_value)
                            if price > 0:
                                has_valid_price = True
                            elif price == 0:
                                price_errors.append(f"{field_name} tidak boleh 0")
                            else:
                                price_errors.append(f"{field_name} harus lebih besar dari 0")
                        except (ValueError, TypeError):
                            price_errors.append(f"Format {field_name} tidak valid: '{value}' (gunakan angka)")
            
            # Add price-related errors
            errors.extend(price_errors)
            
            # Check if at least one valid price exists
            if not has_valid_price and any(field in column_mapping for field in pricing_fields.keys()):
                # Only require price if price columns exist in the mapping
                price_columns_found = [field for field in pricing_fields.keys() if field in column_mapping]
                if price_columns_found:
                    errors.append(f"Minimal salah satu harga harus diisi dengan nilai > 0 dari kolom: {', '.join(price_columns_found)}")
            
            # ✅ UPDATED: 7. Business Logic Validation
            # Check if pengirim and penerima are different
            if pengirim_name and penerima_name and pengirim_name.upper() == penerima_name.upper():
                errors.append("Pengirim dan Penerima tidak boleh sama")
            
            # ✅ UPDATED: Return comprehensive validation result
            return {
                'valid': len(errors) == 0,
                'errors': errors,
                'pengirim_name': pengirim_name,
                'penerima_name': penerima_name,
                'pengirim_id': pengirim_id,
                'penerima_id': penerima_id,
                'nama_barang': nama_barang,
                'row_index': row_index
            }
            
        except Exception as e:
            return {
                'valid': False,
                'errors': [f"Error validasi baris {row_index}: {str(e)}"],
                'pengirim_name': '',
                'penerima_name': '',
                'pengirim_id': None,
                'penerima_id': None,
                'nama_barang': '',
                'row_index': row_index
            }

    def validate_excel_row_enhanced(self, row, column_mapping, existing_customers, row_number):
        """Enhanced validation for Excel row with better error handling - UPDATED"""
        errors = []
        result = {
            'valid': False, 
            'errors': [], 
            'row_number': row_number,
            'pengirim_name': '',
            'penerima_name': '',
            'pengirim_id': None,
            'penerima_id': None,
            'nama_barang': ''
        }
        
        try:
            # ✅ UPDATED: Get pengirim
            pengirim_name = str(row.get(column_mapping.get('pengirim', ''), '')).strip()
            print(f"Validating row {row_number}: Pengirim='{pengirim_name}'")
            if not pengirim_name or pengirim_name.upper() == 'NAN':
                errors.append("Pengirim tidak boleh kosong")
            else:
                pengirim_id = existing_customers.get(pengirim_name.upper())
                if not pengirim_id:
                    errors.append(f"Pengirim '{pengirim_name}' tidak ditemukan di database")
                else:
                    result['pengirim_id'] = pengirim_id
                    result['pengirim_name'] = pengirim_name
            
            # ✅ UPDATED: Get penerima
            penerima_name = str(row.get(column_mapping.get('penerima', ''), '')).strip()
            if not penerima_name or penerima_name.upper() == 'NAN':
                errors.append("Penerima tidak boleh kosong")
            else:
                penerima_id = existing_customers.get(penerima_name.upper())
                if not penerima_id:
                    errors.append(f"Penerima '{penerima_name}' tidak ditemukan di database")
                else:
                    result['penerima_id'] = penerima_id
                    result['penerima_name'] = penerima_name
            
            # Get nama barang
            nama_barang = str(row.get(column_mapping.get('nama_barang', ''), '')).strip()
            if not nama_barang or nama_barang.upper() == 'NAN':
                errors.append("Nama barang tidak boleh kosong")
            else:
                result['nama_barang'] = nama_barang
            
            # ✅ UPDATED: Validate numeric fields (optional) - Enhanced validation
            numeric_fields = {
                'panjang': ('Panjang', 'float'),
                'lebar': ('Lebar', 'float'),
                'tinggi': ('Tinggi', 'float'),
                'm3': ('Volume M³', 'float'),
                'ton': ('Berat Ton', 'float'),
                'colli': ('Colli', 'int')
            }
            
            for field, (field_name, data_type) in numeric_fields.items():
                if field in column_mapping:
                    value = row.get(column_mapping[field])
                    if pd.notna(value) and str(value).strip() != '' and str(value).upper() != 'NAN':
                        try:
                            clean_value = str(value).replace(',', '').replace(' ', '')
                            if data_type == 'int':
                                test_val = int(float(clean_value))
                                if test_val <= 0:
                                    errors.append(f"{field_name} harus lebih besar dari 0")
                            else:
                                test_val = float(clean_value)
                                if test_val <= 0:
                                    errors.append(f"{field_name} harus lebih besar dari 0")
                        except (ValueError, TypeError):
                            errors.append(f"{field_name} harus berupa angka valid (ditemukan: '{value}')")
            
            # ✅ UPDATED: Validate pricing fields
            pricing_fields = [
                'harga_m3_pp', 'harga_m3_pd', 'harga_m3_dd',
                'harga_ton_pp', 'harga_ton_pd', 'harga_ton_dd', 
                'harga_col_pp', 'harga_col_pd', 'harga_col_dd',
                'harga_m3', 'harga_ton', 'harga_coll'  # Legacy support
            ]
            
            has_valid_price = False
            for field in pricing_fields:
                if field in column_mapping:
                    value = row.get(column_mapping[field])
                    if pd.notna(value) and str(value).strip() != '' and str(value).upper() != '-':
                        try:
                            clean_value = str(value).replace(',', '').replace(' ', '').replace('Rp', '')
                            test_price = float(clean_value)
                            if test_price > 0:
                                has_valid_price = True
                            elif test_price <= 0:
                                errors.append(f"Harga {field} harus lebih besar dari 0")
                        except (ValueError, TypeError):
                            errors.append(f"Format harga {field} tidak valid (ditemukan: '{value}')")
            
            # Business rules
            if pengirim_name and penerima_name and pengirim_name.upper() == penerima_name.upper():
                errors.append("Pengirim dan Penerima tidak boleh sama")
            
            result['errors'] = errors
            result['valid'] = len(errors) == 0
            
            return result
            
        except Exception as e:
            result['errors'] = [f"Error validasi: {str(e)}"]
            return result
        
    def upload_excel_data(self):
        """Upload Excel data with enhanced validation and error handling - CLEANED VERSION"""
        filename = self.file_path_var.get()
        if not filename:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu!")
            return
        
        # ✅ ENHANCED: Check file exists and accessible
        if not os.path.exists(filename):
            messagebox.showerror("Error", f"File tidak ditemukan: {filename}")
            return
        
        # Disable upload button during process
        original_btn_text = self.upload_btn.cget('text')
        self.upload_btn.config(state='disabled', text="Processing...")
        self.status_label.config(text="Memproses file Excel...", fg='#3498db')
        self.window.update()
        
        try:
            print(f"Starting barang upload from file: {filename}")

            # Step 1: Read and validate Excel file
            try:
                self.status_label.config(text="Membaca file Excel...", fg='#3498db')
                self.window.update()
                
                df = pd.read_excel(filename, engine='openpyxl')
                if df.empty:
                    raise ValueError("File Excel kosong atau tidak memiliki data")
                
                # ✅ ENHANCED: Clean and validate data
                df.columns = df.columns.astype(str).str.strip()
                
                # Remove completely empty rows
                df = df.dropna(how='all')
                
                if df.empty:
                    raise ValueError("File Excel tidak memiliki data yang valid (semua baris kosong)")
                
                print(f"Found {len(df)} rows in Excel file (after removing empty rows)")
                
            except Exception as e:
                raise ValueError(f"Gagal membaca file Excel: {str(e)}")
            
            # Step 2: Check column mapping
            column_mapping = getattr(self, 'column_mapping', {})
            if not column_mapping:
                raise ValueError("Column mapping tidak tersedia. Silakan preview file terlebih dahulu dengan klik 'Browse' dan pilih file.")
            
            # ✅ ENHANCED: Validate required columns exist and accessible
            required_fields = ['pengirim', 'penerima', 'nama_barang']
            missing_fields = []
            invalid_columns = []
            
            for field in required_fields:
                if field not in column_mapping:
                    missing_fields.append(field)
                elif column_mapping[field] not in df.columns:
                    invalid_columns.append(f"{field} -> {column_mapping[field]}")
            
            if missing_fields:
                raise ValueError(f"Kolom wajib tidak ditemukan dalam mapping: {', '.join(missing_fields)}")
            
            if invalid_columns:
                raise ValueError(f"Kolom mapped tidak ada dalam file Excel: {', '.join(invalid_columns)}")
            
            # Step 3: Get existing customers (pengirim & penerima)
            self.status_label.config(text="Memvalidasi data customer...", fg='#3498db')
            self.window.update()

            try:
                existing_customers = {c['nama_customer'].upper(): c['customer_id'] for c in self.db.get_all_customers()}
                print(f"Found {len(existing_customers)} existing customers in database")


                if not existing_customers:
                    raise ValueError("Tidak ada customer yang terdaftar dalam database. Silakan tambahkan customer terlebih dahulu.")
                
            except Exception as e:
                raise ValueError(f"Gagal mengambil data customer: {str(e)}")
            
            # Step 4: Filter and validate data
            self.status_label.config(text="Memfilter data valid...", fg='#3498db')
            self.window.update()
            
            # Filter rows with required data - more thorough checking
            required_cols = [column_mapping['pengirim'], column_mapping['penerima'], column_mapping['nama_barang']]
            
            # Check each required column individually for better error reporting
            for col in required_cols:
                if col not in df.columns:
                    raise ValueError(f"Kolom '{col}' tidak ditemukan dalam file Excel")
            
            # Filter valid rows
            valid_rows = df.dropna(subset=required_cols)
            
            # ✅ ENHANCED: Additional filtering for meaningful data
            def is_meaningful_value(value):
                if pd.isna(value):
                    return False
                str_val = str(value).strip().upper()
                return str_val not in ['', 'NAN', 'NONE', 'NULL', '#N/A', '#NULL!']
            
            # Further filter rows with meaningful data
            meaningful_rows = []
            for _, row in valid_rows.iterrows():
                if (is_meaningful_value(row[column_mapping['pengirim']]) and 
                    is_meaningful_value(row[column_mapping['penerima']]) and 
                    is_meaningful_value(row[column_mapping['nama_barang']])):
                    meaningful_rows.append(row)
            
            if not meaningful_rows:
                raise ValueError(
                    "Tidak ada data valid untuk diupload!\n\n" +
                    "Pastikan kolom Pengirim, Penerima, dan Nama Barang terisi dengan data yang valid.\n" +
                    f"Total baris dalam file: {len(df)}\n" +
                    f"Baris dengan data tidak kosong: {len(valid_rows)}\n" +
                    f"Baris dengan data bermakna: {len(meaningful_rows)}"
                )
            
            valid_rows = pd.DataFrame(meaningful_rows)
            print(f"Processing {len(valid_rows)} valid rows (filtered from {len(df)} total rows)")
            
            # Step 5: Detailed validation with progress
            self.status_label.config(text="Memvalidasi data barang...", fg='#3498db')
            self.window.update()
            
            validation_errors = []
            customer_not_found_list = []
            valid_data_for_upload = []
            
            for idx, (_, row) in enumerate(valid_rows.iterrows()):
                try:
                    validation_result = self.validate_excel_row_enhanced(
                        row, column_mapping, existing_customers,idx + 2
                    )
                    
                    print(f"Validation result: {validation_result}")
                    
                    if validation_result['valid']:
                        valid_data_for_upload.append((idx, row, validation_result))
                    else:
                        # ✅ ENHANCED: Better error categorization
                        customer_errors = [e for e in validation_result['errors'] 
                                        if any(keyword in e.lower() for keyword in ['tidak ditemukan', 'tidak terdaftar', 'not found'])]
                        other_errors = [e for e in validation_result['errors'] 
                                    if not any(keyword in e.lower() for keyword in ['tidak ditemukan', 'tidak terdaftar', 'not found'])]
                        
                        if customer_errors:
                            customer_not_found_list.append({
                                'nama_barang': validation_result.get('nama_barang', 'N/A'),
                                'pengirim': validation_result.get('pengirim_name', 'N/A'),
                                'penerima': validation_result.get('penerima_name', 'N/A'),
                                'row_number': idx + 2,
                                'errors': customer_errors
                            })
                        
                        if other_errors:
                            validation_errors.append({
                                'nama_barang': validation_result.get('nama_barang', 'N/A'),
                                'pengirim': validation_result.get('pengirim_name', 'N/A'),
                                'penerima': validation_result.get('penerima_name', 'N/A'),
                                'error': '; '.join(other_errors),
                                'row_number': idx + 2
                            })
                            
                except Exception as e:
                    print(f"Validation exception at row {idx + 2}: {str(e)}")
                    validation_errors.append({
                        'nama_barang': str(row.get(column_mapping.get('nama_barang', ''), 'N/A'))[:50],
                        'pengirim': str(row.get(column_mapping.get('pengirim', ''), 'N/A'))[:30],
                        'penerima': str(row.get(column_mapping.get('penerima', ''), 'N/A'))[:30],
                        'error': f"Validation exception: {str(e)}",
                        'row_number': idx + 2
                    })
            
            # Step 6: Handle validation errors with detailed reporting
            total_errors = len(validation_errors) + len(customer_not_found_list)
            if total_errors > 0:
                self.status_label.config(
                    text=f"Ditemukan {total_errors} error validasi!", 
                    fg='#e74c3c'
                )
                
                # ✅ ENHANCED: More informative error dialog
                error_summary = (
                    f"HASIL VALIDASI:\n\n"
                    f"Data valid siap upload: {len(valid_data_for_upload)}\n"
                    f"Customer tidak ditemukan: {len(customer_not_found_list)}\n"
                    f"Error validasi lainnya: {len(validation_errors)}\n"
                    f"Total baris diproses: {len(valid_rows)}\n\n"
                )
                
                if customer_not_found_list:
                    # Show sample of missing customers
                    sample_missing = list(set([item['pengirim'] for item in customer_not_found_list[:3]] + 
                                            [item['penerima'] for item in customer_not_found_list[:3]]))
                    error_summary += f"Contoh customer tidak ditemukan: {', '.join(sample_missing[:5])}\n\n"
                
                error_summary += "Lihat detail lengkap error?"
                
                if messagebox.askyesno("Hasil Validasi", error_summary):
                    self.show_enhanced_error_details(
                        validation_errors, 
                        customer_not_found_list, 
                        0,  # success_count = 0 karena belum upload
                        len(valid_rows)
                    )
                return
            
            # Step 7: Confirm upload with detailed summary
            self.status_label.config(text="Semua data valid! Siap upload...", fg='#27ae60')
            
            # ✅ ENHANCED: More informative confirmation dialog
            confirmation_msg = (
                f"VALIDASI BERHASIL!\n\n"
                f"Total barang siap upload: {len(valid_data_for_upload)}\n"
                f"File: {os.path.basename(filename)}\n\n"
                f"Lanjutkan upload ke database?"
            )
            
            if not messagebox.askyesno("Konfirmasi Upload", confirmation_msg):
                self.status_label.config(text="Upload dibatalkan oleh user", fg='#95a5a6')
                return
            
            # Step 8: Upload to database with simple progress tracking
            self.status_label.config(text="Mengupload data ke database...", fg='#3498db')
            self.window.update()
            
            success_count = 0
            upload_errors = []
            
            total_upload_items = len(valid_data_for_upload)
            
            for idx, (original_idx, row, validation_data) in enumerate(valid_data_for_upload):
                
                try:
                    # Simple progress update
                    progress = int((idx + 1) / total_upload_items * 100)
                    self.status_label.config(
                        text=f"Uploading... {progress}% ({idx + 1}/{total_upload_items})",
                        fg='#3498db'
                    )
                    self.window.update()
                    
                    # Get validated data
                    pengirim_id = validation_data['pengirim_id']
                    print("Pengirim ID:", pengirim_id)
                    penerima_id = validation_data['penerima_id']
                    nama_barang = validation_data['nama_barang']
                    
                    # Extract all fields with enhanced error handling
                    extracted_data = self.extract_row_data(row, column_mapping)
                    
                    print(f"Processing row {original_idx + 1}: {nama_barang}")
                    print(f"  Pengirim ID: {pengirim_id}, Penerima ID: {penerima_id}")
                    print(f"  Extracted data keys: {list(extracted_data.keys())}")
                    
                    # Create barang in database
                    barang_id = self.db.create_barang(
                        pengirim=pengirim_id,
                        penerima=penerima_id,
                        nama_barang=nama_barang,
                        **extracted_data  # Spread all extracted data
                    )
                    
                    success_count += 1
                    print(f"Barang created successfully with ID: {barang_id}")
                    
                except Exception as e:
                    error_detail = str(e)
                    print(f"Error creating barang '{nama_barang}': {error_detail}")
                    
                    upload_errors.append({
                        'nama_barang': nama_barang[:50],  # Truncate long names
                        'pengirim': validation_data.get('pengirim_name', 'N/A')[:30],
                        'penerima': validation_data.get('penerima_name', 'N/A')[:30],
                        'error': f"Database error: {error_detail}"[:200],  # Truncate long errors
                        'row_number': original_idx + 2
                    })
            
            # Step 9: Show comprehensive results
            if upload_errors:
                self.status_label.config(
                    text=f"Upload selesai: {success_count} berhasil, {len(upload_errors)} error",
                    fg='#f39c12'
                )
                self.show_enhanced_error_details(
                    upload_errors, [], success_count, len(valid_data_for_upload)
                )
            else:
                self.status_label.config(
                    text=f"Upload berhasil! {success_count} barang ditambahkan",
                    fg='#27ae60'
                )
                
                # ✅ ENHANCED: More detailed success message
                success_msg = (
                    f"UPLOAD BERHASIL!\n\n"
                    f"Total berhasil: {success_count} barang\n"
                    f"File: {os.path.basename(filename)}\n\n"
                    f"Data telah tersimpan dalam database."
                )
                
                messagebox.showinfo("Upload Berhasil!", success_msg)
            
            # Step 10: Refresh and cleanup
            try:
                self.load_barang()
                if hasattr(self, 'refresh_callback') and self.refresh_callback:
                    self.refresh_callback()
            except Exception as e:
                print(f"Warning: Failed to refresh data: {str(e)}")
                    
        except Exception as e:
            error_msg = str(e)
            print(f"Fatal error during upload: {error_msg}")
            
            self.status_label.config(
                text=f"Error: {error_msg[:80]}...",
                fg='#e74c3c'
            )
            
            # ✅ ENHANCED: Better error dialog with troubleshooting tips
            error_dialog = (
                f"GAGAL UPLOAD DATA\n\n"
                f"Error: {error_msg}\n\n"
                f"TIPS MENGATASI:\n"
                f"• Pastikan file Excel tidak sedang dibuka\n"
                f"• Periksa format data dalam file Excel\n"
                f"• Pastikan customer Pengirim & Penerima sudah terdaftar\n"
                f"• Coba preview file terlebih dahulu\n"
                f"• Gunakan template Excel yang disediakan"
            )
            
            messagebox.showerror("Error Upload", error_dialog)
            
        finally:
            # ✅ ENHANCED: Always restore button state
            self.upload_btn.config(state='normal', text=original_btn_text)

    def extract_row_data(self, row, column_mapping):
        """Extract and convert row data with proper type handling"""
        def get_safe_value(field_name, value_type='str', default_value=None):
            if field_name not in column_mapping:
                return default_value
            
            value = row.get(column_mapping[field_name])
            if pd.isna(value) or str(value).strip() == '' or str(value).upper() == 'NAN':
                return default_value
            
            try:
                if value_type == 'float':
                    # Handle comma separated numbers (Indonesian format)
                    if str(value).strip() == "-":
                        return default_value
                    
                    if isinstance(value, str):
                        value = value.replace(',', '').replace(' ', '')
                    return float(str(value).strip()) if str(value).strip() else default_value
                elif value_type == 'int':
                    if isinstance(value, str):
                        value = value.replace(',', '').replace(' ', '')
                    return int(float(str(value).strip())) if str(value).strip() else default_value
                else:
                    return str(value).strip() if str(value).strip() else default_value
            except Exception as e:
                raise ValueError(f"Format {field_name} tidak valid: '{value}' - {str(e)}")
        
        return {
            'panjang_barang': get_safe_value('panjang', 'float'),
            'lebar_barang': get_safe_value('lebar', 'float'),
            'tinggi_barang': get_safe_value('tinggi', 'float'),
            'm3_barang': get_safe_value('m3', 'float'),
            'ton_barang': get_safe_value('ton', 'float'),
            # Pricing fields
            'm3_pp': get_safe_value('harga_m3_pp', 'float'),
            'm3_pd': get_safe_value('harga_m3_pd', 'float'),
            'm3_dd': get_safe_value('harga_m3_dd', 'float'),
            'ton_pp': get_safe_value('harga_ton_pp', 'float'),
            'ton_pd': get_safe_value('harga_ton_pd', 'float'),
            'ton_dd': get_safe_value('harga_ton_dd', 'float'),
            'col_pp': get_safe_value('harga_col_pp', 'float'),
            'col_pd': get_safe_value('harga_col_pd', 'float'),
            'col_dd': get_safe_value('harga_col_dd', 'float'),
        }

    def show_enhanced_error_details(self, validation_errors, customer_not_found, success_count, total_count):
        """Show enhanced error details in a popup window"""
        error_window = tk.Toplevel(self.window)
        error_window.title("📋 Detail Error Upload")
        error_window.geometry("1000x600")
        error_window.configure(bg='#ecf0f1')
        
        # Header
        header_frame = tk.Frame(error_window, bg='#e74c3c', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        tk.Label(
            header_frame,
            text=f"📋 LAPORAN UPLOAD EXCEL",
            font=('Arial', 16, 'bold'),
            fg='white',
            bg='#e74c3c'
        ).pack(expand=True)
        
        # Summary
        summary_frame = tk.Frame(error_window, bg='#ffffff', relief='solid', bd=1)
        summary_frame.pack(fill='x', padx=10, pady=10)
        
        summary_text = f"📊 RINGKASAN:\n"
        summary_text += f"✅ Berhasil: {success_count}/{total_count}\n"
        summary_text += f"❌ Customer tidak ditemukan: {len(customer_not_found)}\n"
        summary_text += f"⚠️ Error validasi: {len(validation_errors)}"
        
        tk.Label(
            summary_frame,
            text=summary_text,
            font=('Arial', 12, 'bold'),
            bg='#ffffff',
            fg='#2c3e50',
            justify='left'
        ).pack(pady=10, padx=20)
        
        # Error details in notebook
        notebook = ttk.Notebook(error_window)
        notebook.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Customer not found tab
        if customer_not_found:
            customer_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(customer_frame, text=f"👥 Customer Tidak Ditemukan ({len(customer_not_found)})")
            
            customer_text = tk.Text(customer_frame, wrap='word', font=('Consolas', 10))
            customer_scroll = ttk.Scrollbar(customer_frame, orient='vertical', command=customer_text.yview)
            customer_text.configure(yscrollcommand=customer_scroll.set)
            
            customer_text.pack(side='left', fill='both', expand=True)
            customer_scroll.pack(side='right', fill='y')
            
            for error in customer_not_found:
                customer_text.insert('end', 
                    f"Baris {error['row_number']}: {error['nama_barang']}\n" +
                    f"  Pengirim: {error['pengirim']}\n" +
                    f"  Penerima: {error['penerima']}\n" +
                    f"  Error: {'; '.join(error['errors'])}\n\n"
                )
        
        # Validation errors tab
        if validation_errors:
            validation_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(validation_frame, text=f"⚠️ Error Validasi ({len(validation_errors)})")
            
            validation_text = tk.Text(validation_frame, wrap='word', font=('Consolas', 10))
            validation_scroll = ttk.Scrollbar(validation_frame, orient='vertical', command=validation_text.yview)
            validation_text.configure(yscrollcommand=validation_scroll.set)
            
            validation_text.pack(side='left', fill='both', expand=True)
            validation_scroll.pack(side='right', fill='y')
            
            for error in validation_errors:
                validation_text.insert('end',
                    f"Baris {error['row_number']}: {error['nama_barang']}\n" +
                    f"  Pengirim: {error['pengirim']}\n" +
                    f"  Penerima: {error['penerima']}\n" +
                    f"  Error: {error['error']}\n\n"
                )
        
        # Close button
        tk.Button(
            error_window,
            text="🚪 Tutup",
            font=('Arial', 12, 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=10,
            pady=5,
            command=error_window.destroy
        ).pack(pady=10)
    
    def add_barang(self):
        """Add new barang manually with enhanced pricing for pengirim-penerima system"""
        
        # Validate form data
        if not self.validated_barang():
            return
        
        try:
            # Get pengirim and penerima names
            pengirim_name = self.pengirim_var.get().strip()
            penerima_name = self.penerima_var.get().strip()
            
            if not pengirim_name:
                messagebox.showerror("Error", "Pilih pengirim terlebih dahulu!")
                self.pengirim_combo.focus()
                return
                
            if not penerima_name:
                messagebox.showerror("Error", "Pilih penerima terlebih dahulu!")
                self.penerima_combo.focus()
                return
            
            # Get customer IDs from names
            pengirim_id = self.db.get_customer_id_by_name(pengirim_name)
            penerima_id = self.db.get_customer_id_by_name(penerima_name)
            
            if not pengirim_id:
                messagebox.showerror("Error", f"Pengirim '{pengirim_name}' tidak ditemukan dalam database!")
                return
                
            if not penerima_id:
                messagebox.showerror("Error", f"Penerima '{penerima_name}' tidak ditemukan dalam database!")
                return
            
            # Get basic barang data
            nama_barang = self.barang_entry.get().strip()
            
            if not nama_barang:
                messagebox.showerror("Error", "Nama barang harus diisi!")
                self.barang_entry.focus()
                return
            
            # Get numeric values with error handling
            def get_numeric_value(entry_widget, field_name):
                try:
                    value = entry_widget.get().strip()
                    
                    if value == "-":
                        return None
                    
                    if not value:
                        return None
                    return float(value)
                except ValueError:
                    raise ValueError(f"Format {field_name} tidak valid: '{value}' (gunakan angka)")
            
            panjang = get_numeric_value(self.panjang_entry, "panjang")
            lebar = get_numeric_value(self.lebar_entry, "lebar")
            tinggi = get_numeric_value(self.tinggi_entry, "tinggi")
            m3 = get_numeric_value(self.m3_entry, "volume")
            ton = get_numeric_value(self.ton_entry, "berat")
            
            # Get pricing information for all service types
            # M3 pricing
            harga_m3_pp = get_numeric_value(self.harga_m3_pp_entry, "Harga M3 PP") if hasattr(self, 'harga_m3_pp_entry') else None
            harga_m3_pd = get_numeric_value(self.harga_m3_pd_entry, "Harga M3 PD") if hasattr(self, 'harga_m3_pd_entry') else None
            harga_m3_dd = get_numeric_value(self.harga_m3_dd_entry, "Harga M3 DD") if hasattr(self, 'harga_m3_dd_entry') else None
            
            # Ton pricing
            harga_ton_pp = get_numeric_value(self.harga_ton_pp_entry, "Harga Ton PP") if hasattr(self, 'harga_ton_pp_entry') else None
            harga_ton_pd = get_numeric_value(self.harga_ton_pd_entry, "Harga Ton PD") if hasattr(self, 'harga_ton_pd_entry') else None
            harga_ton_dd = get_numeric_value(self.harga_ton_dd_entry, "Harga Ton DD") if hasattr(self, 'harga_ton_dd_entry') else None
            
            # Colli pricing (using correct attribute names)
            harga_col_pp = get_numeric_value(self.harga_colli_pp_entry, "Harga Colli PP") if hasattr(self, 'harga_colli_pp_entry') else None
            harga_col_pd = get_numeric_value(self.harga_colli_pd_entry, "Harga Colli PD") if hasattr(self, 'harga_colli_pd_entry') else None
            harga_col_dd = get_numeric_value(self.harga_colli_dd_entry, "Harga Colli DD") if hasattr(self, 'harga_colli_dd_entry') else None
            
            # Create barang in database using the new pengirim-penerima system
            barang_id = self.db.create_barang(
                pengirim=pengirim_id,
                penerima=penerima_id,
                nama_barang=nama_barang,
                panjang_barang=panjang,
                lebar_barang=lebar,
                tinggi_barang=tinggi,
                m3_barang=m3,
                ton_barang=ton,
                # All pricing options
                m3_pp=harga_m3_pp,
                m3_pd=harga_m3_pd,
                m3_dd=harga_m3_dd,
                ton_pp=harga_ton_pp,
                ton_pd=harga_ton_pd,
                ton_dd=harga_ton_dd,
                col_pp=harga_col_pp,
                col_pd=harga_col_pd,
                col_dd=harga_col_dd
            )
            
            # Generate pricing summary for success message
            pricing_methods = []
            if any([harga_m3_pp, harga_m3_pd, harga_m3_dd]):
                pricing_methods.append("M³")
            if any([harga_ton_pp, harga_ton_pd, harga_ton_dd]):
                pricing_methods.append("Ton")
            if any([harga_col_pp, harga_col_pd, harga_col_dd]):
                pricing_methods.append("Colli")
            
            pricing_info = f" dengan metode pricing: {', '.join(pricing_methods)}" if pricing_methods else ""
            
            # Success message
            success_msg = (
                f"Barang berhasil ditambahkan!\n\n"
                f"ID Barang: {barang_id}\n"
                f"Nama: {nama_barang}\n"
                f"Pengirim: {pengirim_name}\n"
                f"Penerima: {penerima_name}{pricing_info}"
            )
            
            messagebox.showinfo("Sukses", success_msg)
            
            # Clear form and refresh
            self.clear_form()
            self.load_barang()
            
            if self.refresh_callback:
                self.refresh_callback()
            
            # Switch to list tab to see the added barang
            self.notebook.select(2)  # Switch to list tab
            
        except ValueError as ve:
            messagebox.showerror("Format Error", str(ve))
        except Exception as e:
            error_msg = f"Gagal menambahkan barang: {str(e)}"
            print(f"Add barang error: {error_msg}")
            messagebox.showerror("Error", error_msg)
            
    def validated_barang(self):
            # Validate input fields for barang
            if not self.pengirim_combo.get():
                messagebox.showwarning("Peringatan", "Pilih Pengirim terlebih dahulu.")
                return False
            if not self.penerima_var.get():
                messagebox.showwarning("Peringatan", "Pilih Penerima terlebih dahulu.")
                return False
            if not self.barang_entry.get():
                messagebox.showwarning("Peringatan", "Nama Barang tidak boleh kosong.")
                return False
            if not self.panjang_entry.get() or not self.lebar_entry.get() or not self.tinggi_entry.get():
                messagebox.showwarning("Peringatan", "Dimensi Barang tidak boleh kosong.")
                return False
            if not self.m3_entry.get() or not self.ton_entry.get():
                messagebox.showwarning("Peringatan", "Volume dan Berat Barang tidak boleh kosong.")
                return False
            
            harga_m3_pp = self.harga_m3_pp_entry.get().strip()
            harga_m3_pd = self.harga_m3_pd_entry.get().strip()
            harga_m3_dd = self.harga_m3_dd_entry.get().strip()
            
            harga_ton_pp = self.harga_ton_pp_entry.get().strip()
            harga_ton_pd = self.harga_ton_pd_entry.get().strip()
            harga_ton_dd = self.harga_ton_dd_entry.get().strip()
            
            harga_col_pp = self.harga_colli_pp_entry.get().strip()
            harga_col_pd = self.harga_colli_pd_entry.get().strip()
            harga_col_dd = self.harga_colli_dd_entry.get().strip()

            if not (harga_m3_pp or harga_m3_pd or harga_m3_dd or 
                       harga_ton_pp or harga_ton_pd or harga_ton_dd or 
                       harga_col_pp or harga_col_pd or harga_col_dd):
                messagebox.showwarning(
                    "Peringatan", 
                    "Minimal salah satu metode pricing harus diisi!\n\n" +
                    "💰 Pilihan pricing:\n" +
                    "• Harga per m³ (untuk volume)\n" +
                    "• Harga per ton (untuk berat)\n" +
                    "• Harga per colli (untuk jumlah kemasan)"
                )
                self.harga_m3_entry.focus()
                return False
            return True
    
    def clear_form(self):
        """Clear form fields - updated for pengirim-penerima system"""
        try:
            # Clear customer selections
            self.pengirim_var.set('')
            self.penerima_var.set('')
            
            # Clear basic fields
            self.barang_entry.delete(0, tk.END)
            self.panjang_entry.delete(0, tk.END)
            self.lebar_entry.delete(0, tk.END)
            self.tinggi_entry.delete(0, tk.END)
            self.m3_entry.delete(0, tk.END)
            self.ton_entry.delete(0, tk.END)
            
            # Clear all pricing fields if they exist
            pricing_entries = [
                'harga_m3_pp_entry', 'harga_m3_pd_entry', 'harga_m3_dd_entry',
                'harga_ton_pp_entry', 'harga_ton_pd_entry', 'harga_ton_dd_entry',
                'harga_colli_pp_entry', 'harga_colli_pd_entry', 'harga_colli_dd_entry'
            ]
            
            for entry_name in pricing_entries:
                if hasattr(self, entry_name):
                    entry = getattr(self, entry_name)
                    entry.delete(0, tk.END)
            
            # Focus on pengirim combo for next input
            self.pengirim_combo.focus()
            
        except Exception as e:
            print(f"Error clearing form: {str(e)}")
            # Try basic clearing even if some fields fail
            try:
                self.barang_entry.delete(0, tk.END)
                self.pengirim_combo.focus()
            except:
                pass
        
    def load_barang(self):
        """Load barang into PaginatedTreeView"""
        try:
            print("Loading barang from database...")
            
            # Load barang from database
            barang_list = self.db.get_all_barang()
            self.original_barang_data = barang_list  # Store original data for filtering
            print(f"Found {len(barang_list)} barang in database")
            
            # Format data untuk PaginatedTreeView
            formatted_data = []
            
            for barang in barang_list:
                # Format dimensions
                dimensi = f"{barang.get('panjang_barang', '-')}×{barang.get('lebar_barang', '-')}×{barang.get('tinggi_barang', '-')}"
                
                # Format currency
                harga_m3_pp = f"Rp {barang.get('m3_pp', 0):,.0f}" if barang.get('m3_pp') and barang.get('m3_pp') != '-' else '-'
                harga_m3_pd = f"Rp {barang.get('m3_pd', 0):,.0f}" if barang.get('m3_pd') and barang.get('m3_pd') != '-' else '-'
                harga_m3_dd = f"Rp {barang.get('m3_dd', 0):,.0f}" if barang.get('m3_dd') and barang.get('m3_dd') != '-' else '-'

                harga_ton_pp = f"Rp {barang.get('ton_pp', 0):,.0f}" if barang.get('ton_pp') and barang.get('ton_pp') != '-' else '-'
                harga_ton_pd = f"Rp {barang.get('ton_pd', 0):,.0f}" if barang.get('ton_pd') and barang.get('ton_pd') != '-' else '-'
                harga_ton_dd = f"Rp {barang.get('ton_dd', 0):,.0f}" if barang.get('ton_dd') and barang.get('ton_dd') != '-' else '-'

                harga_col_pp = f"Rp {barang.get('col_pp', 0):,.0f}" if barang.get('col_pp') and barang.get('col_pp') != '-' else '-'
                harga_col_pd = f"Rp {barang.get('col_pd', 0):,.0f}" if barang.get('col_pd') and barang.get('col_pd') != '-' else '-'
                harga_col_dd = f"Rp {barang.get('col_dd', 0):,.0f}" if barang.get('col_dd') and barang.get('col_dd') != '-' else '-'

                # Format date
                created_date = barang.get('created_at', '')[:10] if barang.get('created_at') else '-'
                
                # Buat tuple data untuk row ini
                row_data = (
                    barang['barang_id'],
                    barang['sender_name'],
                    barang['receiver_name'],
                    barang['nama_barang'],
                    dimensi,
                    barang.get('m3_barang', '-'),
                    barang.get('ton_barang', '-'),
                    harga_m3_pp,
                    harga_m3_pd,
                    harga_m3_dd,
                    harga_ton_pp,
                    harga_ton_pd,
                    harga_ton_dd,
                    harga_col_pp,
                    harga_col_pd,
                    harga_col_dd,
                    created_date
                )
                
                formatted_data.append(row_data)
            
            # Set data ke PaginatedTreeView
            self.tree.set_data(formatted_data)
            
            print("Barang list loaded successfully")
            
        except Exception as e:
            print(f"Error loading barang: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal memuat daftar barang: {str(e)}")
        
    def on_tab_changed(self, event):
        """Handle tab change event"""
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        
        # If switching to barang list tab, refresh the data
        if "Daftar Barang" in tab_text:
            self.load_barang()
            print("Tab changed to Barang List - data refreshed")