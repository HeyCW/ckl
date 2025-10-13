import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
import re
from src.models.database import AppDatabase
from PIL import Image, ImageTk
from src.widget.paginated_tree_view import PaginatedTreeView

class CustomerWindow:
    def __init__(self, parent, db, refresh_callback=None):
        self.parent = parent
        self.db = db
        self.refresh_callback = refresh_callback
        self.original_customer_data = []
        self.create_window()
    
    def get_scale_factor(self):
        """Calculate scale factor based on screen size"""
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Base scale on 1920x1080 as reference
        width_scale = screen_width / 1920
        height_scale = screen_height / 1080
        
        # Use average and clamp between 0.7 and 1.2
        scale = (width_scale + height_scale) / 2
        return max(0.7, min(1.2, scale))
    
    def scaled_font(self, base_size):
        """Return scaled font size"""
        scale = self.get_scale_factor()
        return max(8, int(base_size * scale))
    
    def on_window_resize(self, event):
        """Handle window resize to adjust column widths"""
        if hasattr(self, 'tree') and event.widget == self.window:
            try:
                window_width = self.window.winfo_width()
                available_width = window_width - 100
                
                # Proportional widths: ID(8%), Nama(25%), Alamat(50%), Created(17%)
                self.tree.column('ID', width=int(available_width * 0.08))
                self.tree.column('Nama', width=int(available_width * 0.25))
                self.tree.column('Alamat', width=int(available_width * 0.50))
                self.tree.column('Created', width=int(available_width * 0.17))
            except:
                pass
    
    def create_window(self):
        """Create customer management window"""
        self.window = tk.Toplevel(self.parent)
        self.window.title("📋 Data Customer")
        
        # Get screen dimensions
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Adaptive window size
        window_width = min(int(screen_width * 0.8), 1400)
        window_height = min(int(screen_height * 0.85), 850)
        
        self.window.geometry(f"{window_width}x{window_height}")
        self.window.configure(bg='#ecf0f1')
        self.window.transient(self.parent)
        self.window.grab_set()
        
        # Resizable
        self.window.minsize(1000, 600)
        self.window.resizable(True, True)
        
        try:
            icon_image = Image.open("assets/logo.jpg")
            icon_image = icon_image.resize((32, 32), Image.Resampling.LANCZOS)
            icon_photo = ImageTk.PhotoImage(icon_image)
            self.window.iconphoto(False, icon_photo)
        except Exception as e:
            print(f"Icon tidak ditemukan: {e}")
        
        self.center_window()
        
        # Header with responsive font
        header = tk.Label(
            self.window,
            text="👥 KELOLA DATA CUSTOMER",
            font=('Arial', self.scaled_font(18), 'bold'),
            bg='#3498db',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        # Notebook
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Tabs
        manual_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(manual_frame, text='✍️ Input Manual')
        self.create_manual_tab(manual_frame)
        
        excel_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(excel_frame, text='📊 Upload Excel')
        self.create_excel_tab(excel_frame)
        
        list_frame = tk.Frame(self.notebook, bg='#ecf0f1')
        self.notebook.add(list_frame, text='📋 Daftar Customer')
        self.create_list_tab(list_frame)
        
        # Close button
        close_btn = tk.Button(
            self.window,
            text="❌ Tutup",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=30,
            pady=10,
            command=self.window.destroy
        )
        close_btn.pack(pady=10)
        
        # Bind resize event
        self.window.bind('<Configure>', self.on_window_resize)
    
    def create_manual_tab(self, parent):
        """Create manual input tab"""
        form_frame = tk.Frame(parent, bg='#ecf0f1')
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        instruction_label = tk.Label(
            form_frame,
            text="📝 Tambah Customer Satu per Satu",
            font=('Arial', self.scaled_font(14), 'bold'),
            fg='#2c3e50',
            bg='#ecf0f1'
        )
        instruction_label.pack(pady=(0, 20))
        
        tk.Label(form_frame, text="Nama Customer:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.name_entry = tk.Entry(form_frame, font=('Arial', self.scaled_font(12)))
        self.name_entry.pack(fill='x', pady=(5, 10))
        
        tk.Label(form_frame, text="Alamat Customer:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        self.address_entry = tk.Text(form_frame, font=('Arial', self.scaled_font(12)), height=4)
        self.address_entry.pack(fill='x', pady=(5, 10))
        
        btn_frame = tk.Frame(form_frame, bg='#ecf0f1')
        btn_frame.pack(fill='x', pady=20)
        
        add_btn = tk.Button(
            btn_frame,
            text="➕ Tambah Customer",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.add_customer
        )
        add_btn.pack(side='left', padx=(0, 10))
        
        clear_btn = tk.Button(
            btn_frame,
            text="🗑️ Bersihkan",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=10,
            pady=5,
            command=self.clear_form
        )
        clear_btn.pack(side='left')
        
        self.name_entry.focus()
    
    def create_excel_tab(self, parent):
        """Create Excel upload tab"""
        main_container = tk.Frame(parent, bg='#ecf0f1')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        instruction_frame = tk.Frame(main_container, bg='#ffffff', relief='solid', bd=1)
        instruction_frame.pack(fill='x', pady=(0, 20))
        
        instruction_title = tk.Label(
            instruction_frame,
            text="📊 Upload Data Customer dari Excel",
            font=('Arial', self.scaled_font(14), 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        instruction_title.pack(pady=(10, 5))
        
        instruction_text = tk.Label(
            instruction_frame,
            text="Format Excel yang dibutuhkan:\n\n" +
                 "• Kolom A: Nama Customer (WAJIB)\n" +
                 "• Kolom B: Alamat Customer (WAJIB)\n" +
                 "• Baris pertama adalah header (Nama, Alamat)\n" +
                 "• Pastikan tidak ada duplikasi nama customer\n\n" +
                 "Tips: Download template untuk format yang benar!",
            font=('Arial', self.scaled_font(10)),
            fg='#34495e',
            bg='#ffffff',
            justify='left'
        )
        instruction_text.pack(pady=(0, 10), padx=20)
        
        file_frame = tk.Frame(main_container, bg='#ecf0f1')
        file_frame.pack(fill='x', pady=10)
        
        tk.Label(file_frame, text="Pilih File Excel:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        
        file_input_frame = tk.Frame(file_frame, bg='#ecf0f1')
        file_input_frame.pack(fill='x', pady=(5, 0))
        
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(file_input_frame, textvariable=self.file_path_var, font=('Arial', self.scaled_font(11)), state='readonly')
        self.file_entry.pack(side='left', fill='x', expand=True, ipady=5)
        
        browse_btn = tk.Button(
            file_input_frame,
            text="📁 Browse",
            font=('Arial', self.scaled_font(10), 'bold'),
            bg='#3498db',
            fg='white',
            padx=15,
            pady=5,
            command=self.browse_file
        )
        browse_btn.pack(side='right', padx=(5, 0))
        
        preview_frame = tk.Frame(main_container, bg='#ecf0f1')
        preview_frame.pack(fill='both', expand=True, pady=10)
        
        tk.Label(preview_frame, text="📋 Preview Data:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        
        preview_tree_frame = tk.Frame(preview_frame, bg='#ecf0f1')
        preview_tree_frame.pack(fill='both', expand=True, pady=5)
        
        preview_container = tk.Frame(preview_tree_frame, bg='#ecf0f1')
        preview_container.pack(fill='both', expand=True)
        
        self.preview_tree = ttk.Treeview(preview_container, 
                                       columns=('Status', 'Nama', 'Alamat'), 
                                       show='headings', height=8)
        
        self.preview_tree.heading('Status', text='Status')
        self.preview_tree.heading('Nama', text='Nama Customer')
        self.preview_tree.heading('Alamat', text='Alamat')
        
        # Responsive column widths
        window_width = self.window.winfo_width()
        self.preview_tree.column('Status', width=max(60, int(window_width * 0.05)))
        self.preview_tree.column('Nama', width=max(200, int(window_width * 0.25)))
        self.preview_tree.column('Alamat', width=max(300, int(window_width * 0.40)))
        
        v_scrollbar = ttk.Scrollbar(preview_container, orient='vertical', command=self.preview_tree.yview)
        h_scrollbar = ttk.Scrollbar(preview_container, orient='horizontal', command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.preview_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)
        
        upload_btn_frame = tk.Frame(main_container, bg='#ecf0f1')
        upload_btn_frame.pack(fill='x', pady=15)
        
        self.upload_btn = tk.Button(
            upload_btn_frame,
            text="⬆️ Upload ke Database",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.upload_excel_data,
            state='disabled'
        )
        self.upload_btn.pack(side='left', padx=(0, 15))
        
        download_template_btn = tk.Button(
            upload_btn_frame,
            text="📥 Download Template",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#f39c12',
            fg='white',
            padx=10,
            pady=5,
            command=self.download_template
        )
        download_template_btn.pack(side='left')
        
        self.status_label = tk.Label(
            main_container,
            text="",
            font=('Arial', self.scaled_font(11)),
            fg='#e74c3c',
            bg='#ecf0f1',
            wraplength=1000,
            justify='left'
        )
        self.status_label.pack(pady=10, fill='x')
    
    def create_list_tab(self, parent):
        """Create customer list tab"""
        list_container = tk.Frame(parent, bg='#ecf0f1')
        list_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        header_frame = tk.Frame(list_container, bg='#ecf0f1')
        header_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(header_frame, text="📋 DAFTAR CUSTOMER", font=('Arial', self.scaled_font(14), 'bold'), bg='#ecf0f1').pack(side='left')
        
        refresh_btn = tk.Button(
            header_frame,
            text="🔄 Refresh",
            font=('Arial', self.scaled_font(10)),
            bg='#95a5a6',
            fg='white',
            padx=15,
            pady=5,
            command=self.load_customers
        )
        refresh_btn.pack(side='right')
        
        search_frame = tk.Frame(list_container, bg='#ffffff', relief='solid', bd=1)
        search_frame.pack(fill='x', pady=(0, 10))
        
        search_label = tk.Label(
            search_frame,
            text="🔍 Pencarian Customer:",
            font=('Arial', self.scaled_font(12), 'bold'),
            fg='#2c3e50',
            bg='#ffffff'
        )
        search_label.pack(anchor='w', padx=10, pady=(10, 5))
        
        search_controls = tk.Frame(search_frame, bg='#ffffff')
        search_controls.pack(fill='x', padx=10, pady=(0, 10))
        
        tk.Label(search_controls, text="Nama:", font=('Arial', self.scaled_font(10)), bg='#ffffff').pack(side='left')
        self.search_name_var = tk.StringVar()
        self.search_name_var.trace('w', self.on_search_change)
        search_name_entry = tk.Entry(search_controls, textvariable=self.search_name_var, font=('Arial', self.scaled_font(10)), width=25)
        search_name_entry.pack(side='left', padx=(5, 20))
        
        tk.Label(search_controls, text="Alamat:", font=('Arial', self.scaled_font(10)), bg='#ffffff').pack(side='left')
        self.search_address_var = tk.StringVar()
        self.search_address_var.trace('w', self.on_search_change)
        search_address_entry = tk.Entry(search_controls, textvariable=self.search_address_var, font=('Arial', self.scaled_font(10)), width=25)
        search_address_entry.pack(side='left', padx=(5, 20))
        
        clear_search_btn = tk.Button(
            search_controls,
            text="❌ Clear",
            font=('Arial', self.scaled_font(9)),
            bg='#e67e22',
            fg='white',
            padx=10,
            pady=2,
            command=self.clear_search
        )
        clear_search_btn.pack(side='left', padx=5)
        
        action_frame = tk.Frame(list_container, bg='#ecf0f1')
        action_frame.pack(fill='x', pady=(0, 10))
        
        update_btn = tk.Button(
            action_frame,
            text="✏️ Edit Customer",
            font=('Arial', self.scaled_font(11), 'bold'),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=5,
            command=self.update_customer
        )
        update_btn.pack(side='left', padx=(0, 10))
        
        delete_btn = tk.Button(
            action_frame,
            text="🗑️ Hapus Customer",
            font=('Arial', self.scaled_font(11), 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=10,
            pady=5,
            command=self.delete_customer
        )
        delete_btn.pack(side='left', padx=(0, 10))
        
        export_btn = tk.Button(
            action_frame,
            text="📤 Export Excel",
            font=('Arial', self.scaled_font(11), 'bold'),
            bg='#27ae60',
            fg='white',
            padx=10,
            pady=5,
            command=self.export_customers
        )
        export_btn.pack(side='left')
        
        self.info_label = tk.Label(
            action_frame,
            text="💡 Pilih customer dari tabel untuk edit/hapus",
            font=('Arial', self.scaled_font(10)),
            fg='#7f8c8d',
            bg='#ecf0f1'
        )
        self.info_label.pack(side='right')
        
        tree_frame = tk.Frame(list_container, bg='#ecf0f1')
        tree_frame.pack(fill='both', expand=True)
        
        tree_container = tk.Frame(tree_frame, bg='#ecf0f1')
        tree_container.pack(fill='both', expand=True)
        
        columns = ('ID', 'Nama', 'Alamat', 'Created')
        
        self.tree = PaginatedTreeView(
            parent=tree_container,
            columns=columns,
            show='headings',
            height=12,
            items_per_page=15
        )
        
        self.tree.heading('ID', text='ID')
        self.tree.heading('Nama', text='Nama Customer')
        self.tree.heading('Alamat', text='Alamat')
        self.tree.heading('Created', text='Tanggal Dibuat')
        
        # Responsive initial column widths
        window_width = self.window.winfo_width()
        self.tree.column('ID', width=max(60, int(window_width * 0.06)))
        self.tree.column('Nama', width=max(200, int(window_width * 0.25)))
        self.tree.column('Alamat', width=max(350, int(window_width * 0.48)))
        self.tree.column('Created', width=max(120, int(window_width * 0.13)))
        
        self.tree.pack(fill='both', expand=True)
        
        self.tree.bind('<Double-1>', lambda e: self.update_customer())
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        
        self.load_customers()
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
    
    def on_search_change(self, *args):
        """Handle search input changes"""
        self.filter_customers()
    
    def clear_search(self):
        """Clear all search filters"""
        self.search_name_var.set('')
        self.search_address_var.set('')
        self.filter_customers()
    
    def filter_customers(self):
        """Filter customers based on search criteria"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        search_name = self.search_name_var.get().strip()
        search_address = self.search_address_var.get().strip()
        
        filtered_data = []
        
        for customer in self.original_customer_data:
            if search_name:
                nama_customer = customer['nama_customer'].lower()
                search_terms = search_name.lower().split()
                
                match_found = False
                for term in search_terms:
                    pattern = '.*'.join(re.escape(char) for char in term)
                    if re.search(pattern, nama_customer, re.IGNORECASE):
                        match_found = True
                        break
                    
                    if term in nama_customer:
                        match_found = True
                        break
                
                if not match_found:
                    continue
            
            if search_address:
                alamat_customer = (customer.get('alamat_customer') or '').lower()
                search_address_lower = search_address.lower()
                
                if search_address_lower not in alamat_customer:
                    continue
            
            filtered_data.append(customer)
        
        for customer in filtered_data:
            created_date = customer.get('created_at', '')[:10] if customer.get('created_at') else '-'
            
            self.tree.insert('', tk.END, values=(
                customer['customer_id'],
                customer['nama_customer'],
                customer['alamat_customer'] or '-',
                created_date
            ))
        
        total_count = len(self.original_customer_data)
        filtered_count = len(filtered_data)
        if total_count != filtered_count:
            self.info_label.config(text=f"📊 Menampilkan {filtered_count} dari {total_count} customer")
        else:
            self.info_label.config(text="💡 Pilih customer dari tabel untuk edit/hapus")
    
    def on_tree_select(self, event):
        """Handle tree selection change"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            customer_id = item['values'][0]
            nama_customer = item['values'][1]
            self.info_label.config(text=f"✅ Terpilih: {nama_customer} (ID: {customer_id})")
        else:
            self.info_label.config(text="💡 Pilih customer dari tabel untuk edit/hapus")
    
    def update_customer(self):
        """Update selected customer"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih customer yang akan diedit dari tabel!")
            return
        
        item = self.tree.item(selection[0])
        customer_id = item['values'][0]
        
        selected_customer = None
        for customer in self.original_customer_data:
            if customer['customer_id'] == customer_id:
                selected_customer = customer
                break
        
        if not selected_customer:
            messagebox.showerror("Error", "Data customer tidak ditemukan!")
            return
        
        self.open_update_dialog(selected_customer)
    
    def open_update_dialog(self, customer_data):
        """Open dialog to update customer data"""
        update_window = tk.Toplevel(self.window)
        update_window.title(f"✏️ Edit Customer - {customer_data['nama_customer']}")
        update_window.configure(bg='#ecf0f1')
        update_window.transient(self.window)
        update_window.grab_set()
        
        dialog_width = 500
        dialog_height = 500
        
        parent_x = self.window.winfo_x()
        parent_y = self.window.winfo_y()
        parent_width = self.window.winfo_width()
        parent_height = self.window.winfo_height()
        
        screen_width = update_window.winfo_screenwidth()
        screen_height = update_window.winfo_screenheight()
        
        x = parent_x + (parent_width // 2) - (dialog_width // 2)
        y = parent_y + (parent_height // 2) - (dialog_height // 2)
        
        if x + dialog_width > screen_width:
            x = screen_width - dialog_width - 20
        if x < 0:
            x = 20
        if y + dialog_height > screen_height:
            y = screen_height - dialog_height - 50
        if y < 0:
            y = 20
        
        update_window.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        update_window.lift()
        update_window.focus_force()
        
        header = tk.Label(
            update_window,
            text="✏️ EDIT DATA CUSTOMER",
            font=('Arial', self.scaled_font(16), 'bold'),
            bg='#3498db',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        form_frame = tk.Frame(update_window, bg='#ecf0f1')
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(form_frame, text="ID Customer:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        id_label = tk.Label(
            form_frame, 
            text=str(customer_data['customer_id']), 
            font=('Arial', self.scaled_font(11)),
            bg='#ffffff',
            relief='solid',
            bd=1,
            padx=5,
            pady=5
        )
        id_label.pack(fill='x', pady=(5, 10))
        
        tk.Label(form_frame, text="Nama Customer:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        nama_var = tk.StringVar(value=customer_data['nama_customer'])
        nama_entry = tk.Entry(form_frame, textvariable=nama_var, font=('Arial', self.scaled_font(11)))
        nama_entry.pack(fill='x', pady=(5, 10))
        
        tk.Label(form_frame, text="Alamat Customer:", font=('Arial', self.scaled_font(12), 'bold'), bg='#ecf0f1').pack(anchor='w')
        alamat_text = tk.Text(form_frame, font=('Arial', self.scaled_font(11)), height=6)
        alamat_text.insert('1.0', customer_data['alamat_customer'] or '')
        alamat_text.pack(fill='x', pady=(5, 20))
        
        def on_save():
            try:
                new_nama = nama_var.get().strip()
                new_alamat = alamat_text.get('1.0', tk.END).strip()
                
                if not new_nama:
                    messagebox.showerror("Error", "Nama customer harus diisi!")
                    nama_entry.focus()
                    return
                
                self.db.update_customer(
                    customer_id=customer_data['customer_id'],
                    nama_customer=new_nama,
                    alamat_customer=new_alamat
                )
                
                messagebox.showinfo("Sukses", "Data customer berhasil diupdate!")
                
                self.load_customers()
                if self.refresh_callback:
                    self.refresh_callback()
                
                update_window.destroy()
                
            except ValueError as ve:
                messagebox.showerror("Error Validasi", f"Data tidak valid!\nError: {str(ve)}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal mengupdate customer!\nError: {str(e)}")
        
        btn_frame = tk.Frame(update_window, bg='#ecf0f1')
        btn_frame.pack(fill='x', padx=20, pady=10, side='bottom')
        
        save_btn = tk.Button(
            btn_frame,
            text="💾 Simpan",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#27ae60',
            fg='white',
            padx=20,
            pady=8,
            command=on_save
        )
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(
            btn_frame,
            text="❌ Batal",
            font=('Arial', self.scaled_font(12), 'bold'),
            bg='#e74c3c',
            fg='white',
            padx=20,
            pady=8,
            command=update_window.destroy
        )
        cancel_btn.pack(side='right')
        
        nama_entry.focus()
        nama_entry.select_range(0, tk.END)
    
    def delete_customer(self):
        """Delete selected customer"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Peringatan", "Pilih customer yang akan dihapus dari tabel!")
            return
        
        item = self.tree.item(selection[0])
        customer_id = item['values'][0]
        nama_customer = item['values'][1]
        
        try:
            barang_count = self.db.execute_one(
                "SELECT COUNT(*) as count FROM barang WHERE customer_id = ?",
                (customer_id,)
            )
            
            has_barang = barang_count['count'] > 0 if barang_count else False
            
            confirm_msg = f"Yakin ingin menghapus customer?\n\n" + \
                         f"ID: {customer_id}\n" + \
                         f"Nama: {nama_customer}\n\n"
            
            if has_barang:
                confirm_msg += f"⚠️ PERINGATAN: Customer ini memiliki {barang_count['count']} barang terkait!\n" + \
                              f"Menghapus customer akan menghapus semua barang terkait.\n\n"
            
            confirm_msg += f"⚠️ Aksi ini tidak dapat dibatalkan!"
            
            if not messagebox.askyesno("Konfirmasi Hapus", confirm_msg):
                return
            
            self.db.execute("DELETE FROM customers WHERE customer_id = ?", (customer_id,))
            
            messagebox.showinfo("Sukses", f"Customer '{nama_customer}' berhasil dihapus!")
            
            self.load_customers()
            if self.refresh_callback:
                self.refresh_callback()
                
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menghapus customer:\n{str(e)}")
    
    def export_customers(self):
        """Export customer data to Excel"""
        try:
            if not self.original_customer_data:
                messagebox.showwarning("Peringatan", "Tidak ada data customer untuk diekspor!")
                return
            
            filename = filedialog.asksaveasfilename(
                parent=self.window,
                title="Export Data Customer ke Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"data_customer_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            export_data = []
            for customer in self.original_customer_data:
                export_data.append({
                    'ID': customer['customer_id'],
                    'Nama Customer': customer['nama_customer'],
                    'Alamat': customer.get('alamat_customer', ''),
                    'Tanggal Dibuat': customer.get('created_at', '')
                })
            
            df = pd.DataFrame(export_data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data Customer')
                
                workbook = writer.book
                worksheet = writer.sheets['Data Customer']
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo(
                "Export Berhasil",
                f"Data customer berhasil diekspor ke:\n{filename}\n\n" +
                f"📊 Total: {len(export_data)} customer"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export data:\n{str(e)}")
    
    def show_error_details(self, errors, duplicate_list, success_count, total_count):
        """Show detailed error modal"""
        error_window = tk.Toplevel(self.window)
        error_window.title("📊 Detail Error Upload")
        error_window.geometry("900x600")
        error_window.configure(bg='#ecf0f1')
        error_window.transient(self.window)
        error_window.grab_set()
        
        error_window.update_idletasks()
        x = (error_window.winfo_screenwidth() // 2) - (900 // 2)
        y = (error_window.winfo_screenheight() // 2) - (600 // 2)
        error_window.geometry(f"900x600+{x}+{y}")
        
        header = tk.Label(
            error_window,
            text="📊 DETAIL HASIL UPLOAD",
            font=('Arial', self.scaled_font(16), 'bold'),
            bg='#e74c3c',
            fg='white',
            pady=15
        )
        header.pack(fill='x')
        
        summary_frame = tk.Frame(error_window, bg='#ffffff', relief='solid', bd=1)
        summary_frame.pack(fill='x', padx=20, pady=(20, 10))
        
        summary_text = f"""📈 RINGKASAN UPLOAD:
        
✅ Berhasil: {success_count} dari {total_count} customer
❌ Gagal: {len(errors)} customer  
⚠️ Duplikat dilewati: {len(duplicate_list)} customer
        
📋 Total diproses: {total_count} baris data"""
        
        summary_label = tk.Label(
            summary_frame,
            text=summary_text,
            font=('Arial', self.scaled_font(12)),
            fg='#2c3e50',
            bg='#ffffff',
            justify='left',
            padx=20,
            pady=15
        )
        summary_label.pack(fill='x')
        
        notebook = ttk.Notebook(error_window)
        notebook.pack(fill='both', expand=True, padx=20, pady=10)
        
        if errors:
            error_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(error_frame, text=f'❌ Error Processing ({len(errors)})')
            
            error_label = tk.Label(
                error_frame,
                text="Customer yang gagal diproses karena error teknis:",
                font=('Arial', self.scaled_font(12), 'bold'),
                fg='#e74c3c',
                bg='#ecf0f1'
            )
            error_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            error_main_frame = tk.Frame(error_frame, bg='#ecf0f1')
            error_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            error_tree = ttk.Treeview(error_main_frame, 
                                    columns=('No', 'Nama Customer', 'Error'), 
                                    show='headings', height=12)
            
            error_tree.heading('No', text='No')
            error_tree.heading('Nama Customer', text='Nama Customer')
            error_tree.heading('Error', text='Detail Error')
            
            error_tree.column('No', width=50)
            error_tree.column('Nama Customer', width=300)
            error_tree.column('Error', width=400)
            
            error_v_scroll = ttk.Scrollbar(error_main_frame, orient='vertical', command=error_tree.yview)
            error_tree.configure(yscrollcommand=error_v_scroll.set)
            
            error_tree.pack(side='left', fill='both', expand=True)
            error_v_scroll.pack(side='right', fill='y')
            
            for i, error_info in enumerate(errors, 1):
                error_tree.insert('', tk.END, values=(
                    i,
                    error_info.get('nama_customer', 'N/A'),
                    error_info.get('error', 'Unknown error')
                ))
        
        if duplicate_list:
            duplicate_frame = tk.Frame(notebook, bg='#ecf0f1')
            notebook.add(duplicate_frame, text=f'⚠️ Duplikat Dilewati ({len(duplicate_list)})')
            
            duplicate_label = tk.Label(
                duplicate_frame,
                text="Customer yang sudah ada di database (dilewati):",
                font=('Arial', self.scaled_font(12), 'bold'),
                fg='#f39c12',
                bg='#ecf0f1'
            )
            duplicate_label.pack(anchor='w', padx=10, pady=(10, 5))
            
            duplicate_main_frame = tk.Frame(duplicate_frame, bg='#ecf0f1')
            duplicate_main_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            duplicate_tree = ttk.Treeview(duplicate_main_frame, 
                                        columns=('No', 'Nama Customer', 'Alamat'), 
                                        show='headings', height=12)
            
            duplicate_tree.heading('No', text='No')
            duplicate_tree.heading('Nama Customer', text='Nama Customer')
            duplicate_tree.heading('Alamat', text='Alamat')
            
            duplicate_tree.column('No', width=50)
            duplicate_tree.column('Nama Customer', width=300)
            duplicate_tree.column('Alamat', width=400)
            
            duplicate_v_scroll = ttk.Scrollbar(duplicate_main_frame, orient='vertical', command=duplicate_tree.yview)
            duplicate_tree.configure(yscrollcommand=duplicate_v_scroll.set)
            
            duplicate_tree.pack(side='left', fill='both', expand=True)
            duplicate_v_scroll.pack(side='right', fill='y')
            
            for i, duplicate_info in enumerate(duplicate_list, 1):
                duplicate_tree.insert('', tk.END, values=(
                    i,
                    duplicate_info.get('nama_customer', 'N/A'),
                    duplicate_info.get('alamat', 'N/A')
                ))
        
        tips_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(tips_frame, text='💡 Tips Upload')
        
        tips_canvas = tk.Canvas(tips_frame, bg='#ecf0f1')
        tips_scrollbar = ttk.Scrollbar(tips_frame, orient="vertical", command=tips_canvas.yview)
        tips_scrollable_frame = tk.Frame(tips_canvas, bg='#ecf0f1')
        
        tips_scrollable_frame.bind(
            "<Configure>",
            lambda e: tips_canvas.configure(scrollregion=tips_canvas.bbox("all"))
        )
        
        tips_canvas.create_window((0, 0), window=tips_scrollable_frame, anchor="nw")
        tips_canvas.configure(yscrollcommand=tips_scrollbar.set)
        
        tips_text = """📋 TIPS UNTUK UPLOAD CUSTOMER YANG SUKSES:

1. 🔍 Format Excel yang Benar:
   • Gunakan template yang sudah disediakan
   • Kolom 'Nama' wajib diisi
   • Kolom 'Alamat' bersifat opsional

2. 📊 Data Quality:
   • Pastikan nama customer tidak kosong
   • Hindari duplikasi nama customer
   • Gunakan format teks yang konsisten

3. 🚨 Mengatasi Error:
   • Perbaiki baris yang error sesuai detail di tab Error
   • Pastikan tidak ada karakter khusus yang merusak format
   • Upload ulang file yang sudah diperbaiki

4. 💾 Backup Data:
   • Simpan file Excel asli sebagai backup
   • Export data yang sudah ada sebelum upload besar

5. 🔧 Troubleshooting Common Errors:
   • "Nama customer harus diisi" → Pastikan kolom nama tidak kosong
   • "Database error" → Periksa koneksi database
   • "File tidak bisa dibaca" → Pastikan file Excel tidak sedang dibuka

6. 📝 Best Practices:
   • Gunakan nama customer yang unik dan deskriptif
   • Isi alamat lengkap untuk memudahkan identifikasi
   • Periksa data di preview sebelum upload
   • Upload dalam batch kecil untuk file besar"""
        
        tips_label = tk.Label(
            tips_scrollable_frame,
            text=tips_text,
            font=('Arial', self.scaled_font(11)),
            fg='#2c3e50',
            bg='#ecf0f1',
            justify='left',
            padx=20,
            pady=20,
            wraplength=800
        )
        tips_label.pack(fill='both', expand=True)
        
        tips_canvas.pack(side="left", fill="both", expand=True)
        tips_scrollbar.pack(side="right", fill="y")
        
        close_btn = tk.Button(
            error_window,
            text="✅ Tutup",
            font=('Arial', self.scaled_font(14), 'bold'),
            bg='#27ae60',
            fg='white',
            padx=40,
            pady=15,
            command=error_window.destroy
        )
        close_btn.pack(pady=20)
    
    def center_window(self):
        """Center window with boundary checks"""
        self.window.update_idletasks()
        
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        window_width = min(int(screen_width * 0.8), 1400)
        window_height = min(int(screen_height * 0.85), 850)
        
        x = parent_x + (parent_width // 2) - (window_width // 2)
        y = parent_y + (parent_height // 2) - (window_height // 2) - 50
        
        if x + window_width > screen_width:
            x = screen_width - window_width - 20
        if x < 0:
            x = 20
        if y + window_height > screen_height:
            y = screen_height - window_height - 50
        if y < 0:
            y = 20
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.window.lift()
        self.window.focus_force()
    
    def browse_file(self):
        """Browse for Excel file"""
        file_types = [
            ('Excel files', '*.xlsx *.xls'),
            ('All files', '*.*')
        ]
        
        filename = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=file_types,
            parent=self.window
        )
        
        if filename:
            self.file_path_var.set(filename)
            self.preview_excel_file(filename)
    
    def preview_excel_file(self, filename):
        """Preview Excel file content"""
        try:
            self.status_label.config(text="📄 Membaca file Excel...", fg='#3498db')
            
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            if not os.path.exists(filename):
                self.status_label.config(text="❌ File tidak ditemukan!", fg='#e74c3c')
                self.upload_btn.config(state='disabled')
                return
            
            df = None
            try:
                df = pd.read_excel(filename, engine='openpyxl')
            except:
                try:
                    df = pd.read_excel(filename, engine='xlrd')
                except Exception as e:
                    self.status_label.config(
                        text=f"❌ Error membaca Excel: {str(e)}\n\nPastikan file Excel tidak sedang dibuka!", 
                        fg='#e74c3c'
                    )
                    self.upload_btn.config(state='disabled')
                    return
            
            if df is None or df.empty:
                self.status_label.config(text="❌ File Excel kosong atau tidak valid!", fg='#e74c3c')
                self.upload_btn.config(state='disabled')
                return
            
            df.columns = df.columns.astype(str).str.strip()
            
            nama_col = None
            alamat_col = None
            
            for col in df.columns:
                col_lower = col.lower().strip()
                if col_lower in ['nama', 'name', 'nama customer', 'customer']:
                    nama_col = col
                elif col_lower in ['alamat', 'address', 'alamat customer']:
                    alamat_col = col
            
            if nama_col is None:
                available_cols = ', '.join(df.columns.tolist())
                self.status_label.config(
                    text=f"❌ Kolom 'Nama' tidak ditemukan!\n\nKolom tersedia: {available_cols}", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            if alamat_col is None:
                if len(df.columns) > 1:
                    alamat_col = df.columns[1]
                else:
                    alamat_col = 'Alamat'
                    df[alamat_col] = ''
            
            existing_customers = {c['nama_customer'].upper() for c in self.db.get_all_customers()}
            
            valid_rows = df[df[nama_col].notna() & (df[nama_col].astype(str).str.strip() != '')]
            
            if len(valid_rows) == 0:
                self.status_label.config(
                    text="❌ Tidak ada data valid! Pastikan kolom 'Nama' tidak kosong.", 
                    fg='#e74c3c'
                )
                self.upload_btn.config(state='disabled')
                return
            
            preview_data = valid_rows.head(100)
            row_count = len(valid_rows)
            duplicate_count = 0
            
            for _, row in preview_data.iterrows():
                nama = str(row[nama_col]).strip() if pd.notna(row[nama_col]) else ''
                alamat = str(row[alamat_col]).strip() if pd.notna(row[alamat_col]) else ''
                
                if nama:
                    status = "⚠️" if nama.upper() in existing_customers else "✅"
                    if nama.upper() in existing_customers:
                        duplicate_count += 1
                    
                    self.preview_tree.insert('', tk.END, values=(status, nama, alamat))
            
            self.nama_column = nama_col
            self.alamat_column = alamat_col
            
            status_msg = f"✅ File berhasil dibaca: {row_count} baris data valid\n"
            status_msg += f"📋 Kolom: {nama_col}, {alamat_col}\n"
            if duplicate_count > 0:
                status_msg += f"⚠️ {duplicate_count} customer sudah ada (akan dilewati)\n"
            status_msg += f"💡 ✅ = Baru, ⚠️ = Sudah ada"
            
            self.status_label.config(text=status_msg, fg='#27ae60')
            self.upload_btn.config(state='normal')
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Preview error: {error_detail}")
            
            self.status_label.config(text=f"❌ Error membaca file: {str(e)}", fg='#e74c3c')
            self.upload_btn.config(state='disabled')
    
    def upload_excel_data(self):
        """Upload Excel data to database"""
        filename = self.file_path_var.get()
        if not filename:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu!")
            return
        
        try:
            print(f"🔄 Starting customer upload from file: {filename}")
            
            df = pd.read_excel(filename)
            df.columns = df.columns.astype(str).str.strip()
            
            nama_col = getattr(self, 'nama_column', 'Nama')
            alamat_col = getattr(self, 'alamat_column', 'Alamat')
            
            if nama_col not in df.columns:
                messagebox.showerror("Error", f"Kolom '{nama_col}' tidak ditemukan!")
                return
            
            valid_rows = df[df[nama_col].notna() & (df[nama_col].astype(str).str.strip() != '')]
            
            if len(valid_rows) == 0:
                messagebox.showerror("Error", "Tidak ada data valid untuk diupload!")
                return
            
            if not messagebox.askyesno(
                "Konfirmasi Upload", 
                f"Upload {len(valid_rows)} customer ke database?\n\n" +
                f"Data yang sudah ada tidak akan terduplikasi."
            ):
                return
            
            existing_customers = {c['nama_customer'].upper(): c['customer_id'] for c in self.db.get_all_customers()}
            
            success_count = 0
            errors = []
            duplicate_list = []
            
            for idx, (_, row) in enumerate(valid_rows.iterrows()):
                try:
                    nama = str(row[nama_col]).strip()
                    alamat = str(row[alamat_col]).strip() if alamat_col in df.columns and pd.notna(row[alamat_col]) else ''
                    
                    if nama:
                        if nama.upper() in existing_customers:
                            duplicate_list.append({
                                'nama_customer': nama,
                                'alamat': alamat
                            })
                            continue
                        
                        customer_id = self.db.create_customer(nama, alamat)
                        success_count += 1
                        
                        existing_customers[nama.upper()] = customer_id
                
                except Exception as e:
                    error_detail = str(e)
                    errors.append({
                        'nama_customer': nama,
                        'error': error_detail,
                        'row_number': idx + 2
                    })
            
            total_processed = len(valid_rows)
            
            if errors or duplicate_list:
                self.show_error_details(errors, duplicate_list, success_count, total_processed)
            else:
                messagebox.showinfo(
                    "Upload Berhasil! 🎉", 
                    f"Semua data berhasil diupload!\n\n" +
                    f"✅ Total berhasil: {success_count} customer\n" +
                    f"📊 Total diproses: {total_processed} baris data"
                )
            
            self.load_customers()
            if self.refresh_callback:
                self.refresh_callback()
            
            self.notebook.select(2)
            
            self.file_path_var.set("")
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            self.upload_btn.config(state='disabled')
            self.status_label.config(text="")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Upload error: {error_detail}")
            messagebox.showerror("Error", f"Gagal upload data:\n{str(e)}")
    
    def download_template(self):
        """Download Excel template"""
        try:
            print("📥 Creating customer template...")
            
            template_data = {
                'Nama': [
                    'PT. ANEKA PLASTIK INDONESIA',
                    'CV. MAJU BERSAMA',
                    'PT. TEKNOLOGI MODERN',
                    'UD. SUMBER REZEKI',
                    'PT. GLOBAL TRADING'
                ],
                'Alamat': [
                    'Jl. Industri Raya No. 123, Jakarta Timur 13450',
                    'Jl. Kemerdekaan No. 45, Surabaya 60119',
                    'Jl. Teknologi No. 67, Bandung 40132',
                    'Jl. Pasar Baru No. 89, Medan 20111',
                    'Jl. Pelabuhan No. 12, Makassar 90111'
                ]
            }
            
            df = pd.DataFrame(template_data)
            print("✅ Template data created")
            
            filename = filedialog.asksaveasfilename(
                parent=self.window,
                title="Simpan Template Excel Customer",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="template_customer_lengkap.xlsx"
            )
            
            if filename:
                print(f"💾 Saving template to: {filename}")
                
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Data Customer')
                    
                    instructions_data = {
                        'Kolom': ['Nama', 'Alamat'],
                        'Keterangan': [
                            'Nama customer yang unik dan lengkap (WAJIB)',
                            'Alamat lengkap customer termasuk kota dan kode pos (opsional)'
                        ],
                        'Contoh': [
                            'PT. ANEKA PLASTIK INDONESIA',
                            'Jl. Industri Raya No. 123, Jakarta Timur 13450'
                        ]
                    }
                    
                    instructions_df = pd.DataFrame(instructions_data)
                    instructions_df.to_excel(writer, index=False, sheet_name='Petunjuk')
                    
                    workbook = writer.book
                    data_worksheet = writer.sheets['Data Customer']
                    instructions_worksheet = writer.sheets['Petunjuk']
                    
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for worksheet, df_data in [(data_worksheet, df), (instructions_worksheet, instructions_df)]:
                        for col_num, column_title in enumerate(df_data.columns, 1):
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            cell.border = thin_border
                        
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 3, 60)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        for row in worksheet.iter_rows(min_row=2, max_row=len(df_data)+1):
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical="center")
                
                print("✅ Template saved successfully")
                messagebox.showinfo(
                    "Sukses", 
                    f"Template lengkap berhasil disimpan:\n{filename}\n\n" +
                    "Template berisi:\n" +
                    "• Sheet 'Data Customer' - contoh data customer\n" +
                    "• Sheet 'Petunjuk' - penjelasan format\n" +
                    "• Contoh nama dan alamat yang lengkap\n\n" +
                    "Pastikan nama customer unik untuk menghindari duplikasi!"
                )
            else:
                print("❌ Save cancelled by user")
        
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Template download error: {error_detail}")
            
            messagebox.showerror("Error", f"Gagal membuat template:\n\nError: {str(e)}")
    
    def add_customer(self):
        """Add new customer manually"""
        name = self.name_entry.get().strip()
        address = self.address_entry.get(1.0, tk.END).strip()
        
        if not name:
            messagebox.showerror("Error", "Nama customer harus diisi!")
            self.name_entry.focus()
            return
        
        try:
            customer_id = self.db.create_customer(name, address)
            messagebox.showinfo("Sukses", f"Customer berhasil ditambahkan dengan ID: {customer_id}")
            self.clear_form()
            
            self.load_customers()
            if self.refresh_callback:
                self.refresh_callback()
            
            self.notebook.select(2)
            
        except ValueError as ve:
            messagebox.showerror("Error Validasi", f"Data tidak valid!\nError: {str(ve)}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menambahkan customer:\n{str(e)}")
    
    def clear_form(self):
        """Clear form fields"""
        self.name_entry.delete(0, tk.END)
        self.address_entry.delete(1.0, tk.END)
        self.name_entry.focus()
    
    def load_customers(self):
        """Load customers into PaginatedTreeView"""
        try:
            print("Loading customers from database...")
            
            customers = self.db.get_all_customers()
            self.original_customer_data = customers
            print(f"Found {len(customers)} customers in database")
            
            formatted_data = []
            
            for customer in customers:
                created_date = customer.get('created_at', '')[:10] if customer.get('created_at') else '-'
                
                formatted_data.append({
                    'iid': str(customer['customer_id']),
                    'values': (
                        customer['customer_id'],
                        customer['nama_customer'],
                        customer['alamat_customer'] or '-',
                        created_date
                    )
                })
            
            self.tree.set_data(formatted_data)
            
            print("Customer list loaded successfully with pagination")
            
        except Exception as e:
            print(f"Error loading customers: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal memuat daftar customer: {str(e)}")
    
    def on_tab_changed(self, event):
        """Handle tab change event"""
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        
        if "Daftar Customer" in tab_text:
            self.load_customers()
            print("Tab changed to Customer List - data refreshed")


# Example usage and testing functions
def test_customer_window():
    """Test function for CustomerWindow"""
    try:
        from src.models.database import AppDatabase
        
        db = AppDatabase("test_customer.db")
        
        root = tk.Tk()
        root.title("Test Customer Window")
        root.geometry("400x300")
        
        def open_customer_window():
            CustomerWindow(root, db)
        
        test_btn = tk.Button(
            root,
            text="Open Customer Window",
            font=('Arial', 14),
            command=open_customer_window,
            padx=20,
            pady=10
        )
        test_btn.pack(expand=True)
        
        root.mainloop()
        
    except Exception as e:
        print(f"Test error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    test_customer_window()