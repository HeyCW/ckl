import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import date, timedelta
from src.models.database import AppDatabase
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class LiftingWindow:
    def __init__(self, parent, db: AppDatabase):
        self.parent = parent
        self.db = db
        self.create_window()
    
    def create_window(self):
        self.window = tk.Toplevel(self.parent)
        self.window.title("📦 Lifting Report (Biaya POL & POD)")
        self.window.configure(bg="#ecf0f1")
        
        # === Ukuran & posisi tengah ===
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        scale = max(0.8, min(screen_width / 1920, 1.0))
        width = int(1400 * scale)
        height = int(800 * scale)
        x_pos = int((screen_width / 2) - (width / 2))
        y_pos = int((screen_height / 2) - (height / 2))
        self.window.geometry(f"{width}x{height}+{x_pos}+{y_pos}")
        self.window.minsize(900, 600)
        
        # === Font ===
        font_small = ("Arial", int(11 * scale))
        font_normal = ("Arial", int(12 * scale))
        font_bold = ("Arial", int(13 * scale), "bold")
        font_title = ("Arial", int(16 * scale), "bold")
        
        # === Header ===
        title = tk.Label(
            self.window,
            text="📦 LAPORAN LIFTING (BIAYA POL & POD)",
            font=font_title,
            bg="#27ae60",
            fg="white",
            pady=int(8 * scale)
        )
        title.pack(fill="x")
        
        # === Filter Frame ===
        filter_frame = tk.Frame(self.window, bg="#ffffff", relief="solid", bd=1)
        filter_frame.pack(fill="x", padx=20, pady=(10, 0))
        
        tk.Label(
            filter_frame,
            text="🔍 Filter Tanggal (berdasarkan ETD SUB kapal):",
            font=font_bold,
            fg="#2c3e50",
            bg="#ffffff"
        ).pack(anchor="w", padx=10, pady=(10, 5))
        
        controls = tk.Frame(filter_frame, bg="#ffffff")
        controls.pack(fill="x", padx=10, pady=(0, 10))
        
        # === Date pickers ===
        tk.Label(controls, text="Dari:", bg="#ffffff", font=font_normal).pack(side="left", padx=(0, 5))
        
        self.start_date = DateEntry(
            controls,
            date_pattern='dd/MM/yyyy',  # ✅ FORMAT INDONESIA
            width=int(14 * scale),
            font=font_normal,
            borderwidth=2,
            background="#27ae60",
            foreground="white",
            selectmode="day",
            showweeknumbers=False,
            showothermonthdays=False,
            state="readonly",
            locale='id_ID'
        )
        
        try:
            self.start_date._top_cal.configure(showdropdowns=True)
        except Exception:
            pass
        
        self.start_date.set_date(date.today() - timedelta(days=30))
        self.start_date.pack(side="left", padx=(0, 20))
        
        tk.Label(controls, text="Sampai:", bg="#ffffff", font=font_normal).pack(side="left", padx=(0, 5))
        
        self.end_date = DateEntry(
            controls,
            date_pattern='dd/MM/yyyy',  # ✅ FORMAT INDONESIA
            width=int(14 * scale),
            font=font_normal,
            borderwidth=2,
            background="#27ae60",
            foreground="white",
            selectmode="day",
            showweeknumbers=False,
            showothermonthdays=False,
            state="readonly",
            locale='id_ID'
        )
        
        try:
            self.end_date._top_cal.configure(showdropdowns=True)
        except Exception:
            pass
        
        self.end_date.set_date(date.today())
        self.end_date.pack(side="left", padx=(0, 20))
        
        # === Tombol Filter & Clear ===
        tk.Button(
            controls, text="🔍 Filter", bg="#3498db", fg="white", font=font_bold,
            padx=int(12 * scale), pady=int(5 * scale), command=self.filter_data
        ).pack(side="left", padx=8)
        
        tk.Button(
            controls, text="❌ Clear", bg="#e67e22", fg="white", font=font_bold,
            padx=int(12 * scale), pady=int(5 * scale), command=self.clear_filter
        ).pack(side="left", padx=8)
        
        # ✅ TOMBOL EXPORT EXCEL
        tk.Button(
            controls, text="📊 Export Excel", bg="#27ae60", fg="white", font=font_bold,
            padx=int(12 * scale), pady=int(5 * scale), command=self.export_to_excel
        ).pack(side="left", padx=8)
        
        # === Table ===
        table_frame = tk.Frame(self.window, bg="#ecf0f1")
        table_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal")
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical")
        
        columns = [
            "ETD SUB", "NO JOA",
            "THC, LOLO, SEAL, DOC, CLEANING POL",
            "Freight LSS POL",
            "TRUCKING POL",
            "OPS POL",
            "BI. LAIN POL",
            "PPH POL", "PPN POL", "TOTAL BIAYA POL",
            "THC, LOLO, RELOKASI POD",
            "TRUCKING, DOORING POD",
            "FORKLIF POD",
            "BURUH POD",
            "BI. LAIN POD",
            "TOTAL BIAYA POD",
            "TOTAL BIAYA", "NILAI INVOICE", "PROFIT"
        ]
        
        self.tree = ttk.Treeview(
            table_frame, show="headings", columns=columns,
            xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set
        )
        self.tree.pack(fill="both", expand=True)
        
        x_scroll.config(command=self.tree.xview)
        y_scroll.config(command=self.tree.yview)
        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        
        # === Style Table ===
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview.Heading", background="#27ae60", foreground="white", font=font_bold)
        style.configure("Treeview", font=font_small, rowheight=int(26 * scale))
        style.map("Treeview", background=[("selected", "#3498db")])
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=int(120 * scale), anchor="center")
        
        self.tree.tag_configure("evenrow", background="#f8f9fa")
        self.tree.tag_configure("oddrow", background="#eaf2ef")
        
        # === Footer ===
        footer_frame = tk.Frame(self.window, bg="#27ae60")
        footer_frame.pack(fill="x", pady=(5, 0))
        
        self.total_label = tk.Label(
            footer_frame,
            text="TOTAL PROFIT: Rp 0",
            font=font_bold,
            fg="white",
            bg="#27ae60",
            pady=int(5 * scale)
        )
        self.total_label.pack(side="right", padx=20)
        
        self.load_data_from_db()
    
    # ============================================================
    # HELPER METHOD - FORMAT DATE
    # ============================================================
    
    def format_date_indonesian(self, date_string):
        """Convert date from YYYY-MM-DD to DD/MM/YYYY (Indonesian format)"""
        try:
            if not date_string or date_string in ['-', '', 'None', None]:
                return '-'
            
            date_str = str(date_string)
            
            # If already in DD/MM/YYYY format, return as is
            if '/' in date_str and len(date_str.split('/')[0]) <= 2:
                return date_str
            
            # Convert from YYYY-MM-DD to DD/MM/YYYY
            if '-' in date_str:
                parts = date_str.split(' ')[0].split('-')
                if len(parts) == 3 and len(parts[0]) == 4:
                    return f"{parts[2]}/{parts[1]}/{parts[0]}"
            
            return date_str
            
        except Exception as e:
            print(f"Error formatting date: {e}")
            return str(date_string) if date_string else '-'
    
    # ============================================================
    # LOAD DATA
    # ============================================================
    
    def load_data_from_db(self, start_date=None, end_date=None):
        try:
            query = """
                WITH invoice_per_container AS (
                    SELECT container_id, SUM(total_harga) AS total_invoice
                    FROM detail_container
                    GROUP BY container_id
                ),
                delivery_per_container AS (
                    SELECT
                        cdc.container_id,
                        -- POL (Surabaya) - THC Group (THC + LOLO + SEAL + DOC FEE + CLEANING)
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND (
                            cdc.description LIKE '%THC%' OR
                            cdc.description LIKE '%Lolo%' OR
                            cdc.description LIKE '%Seal%' OR
                            cdc.description LIKE '%DOC%' OR
                            cdc.description LIKE '%Doc%' OR
                            cdc.description LIKE '%Cleaning%'
                        ) THEN cdc.cost ELSE 0 END) AS thc_pol,
                        -- POL (Surabaya) - FREIGHT Group (FREIGHT + LSS)
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND (
                            cdc.description LIKE '%Freight%' OR
                            cdc.description LIKE '%Freigth%' OR
                            cdc.description LIKE '%LSS%'
                        ) THEN cdc.cost ELSE 0 END) AS freight_pol,
                        -- POL (Surabaya) - TRUCKING
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND (
                            cdc.description LIKE '%Trucking%' OR
                            cdc.description LIKE '%Truck%'
                        ) THEN cdc.cost ELSE 0 END) AS trucking_pol,
                        -- POL (Surabaya) - OPS
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND cdc.description LIKE '%OPS%' THEN cdc.cost ELSE 0 END) AS ops_pol,
                        -- POL (Surabaya) - PPH, PPN
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND cdc.description LIKE '%PPH%' THEN cdc.cost ELSE 0 END) AS pph_pol,
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND cdc.description LIKE '%PPN%' THEN cdc.cost ELSE 0 END) AS ppn_pol,
                        -- POL (Surabaya) - BI. LAIN (biaya lain-lain termasuk Asuransi)
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' AND NOT (
                            cdc.description LIKE '%THC%' OR
                            cdc.description LIKE '%Lolo%' OR
                            cdc.description LIKE '%Seal%' OR
                            cdc.description LIKE '%DOC%' OR
                            cdc.description LIKE '%Doc%' OR
                            cdc.description LIKE '%Cleaning%' OR
                            cdc.description LIKE '%Freight%' OR
                            cdc.description LIKE '%Freigth%' OR
                            cdc.description LIKE '%LSS%' OR
                            cdc.description LIKE '%Trucking%' OR
                            cdc.description LIKE '%Truck%' OR
                            cdc.description LIKE '%OPS%' OR
                            cdc.description LIKE '%PPH%' OR
                            cdc.description LIKE '%PPN%'
                        ) THEN cdc.cost ELSE 0 END) AS bi_lain_pol,
                        -- Total POL
                        SUM(CASE WHEN cdc.delivery = 'Surabaya' THEN cdc.cost ELSE 0 END) AS total_biaya_pol,
                        -- POD (Non-Surabaya) - THC Group (THC + LOLO + RELOKASI)
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' AND (
                            cdc.description LIKE '%THC%' OR
                            cdc.description LIKE '%Lolo%' OR
                            cdc.description LIKE '%Relokasi%' OR
                            cdc.description LIKE '%Rekolasi%'
                        ) THEN cdc.cost ELSE 0 END) AS thc_pod,
                        -- POD (Non-Surabaya) - TRUCKING, DOORING
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' AND (
                            cdc.description LIKE '%Trucking%' OR
                            cdc.description LIKE '%Truck%' OR
                            cdc.description LIKE '%Dooring%' OR
                            cdc.description LIKE '%Door%'
                        ) THEN cdc.cost ELSE 0 END) AS trucking_dooring_pod,
                        -- POD (Non-Surabaya) - FORKLIFT
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' AND (
                            cdc.description LIKE '%Forklift%' OR
                            cdc.description LIKE '%Forklif%'
                        ) THEN cdc.cost ELSE 0 END) AS forklift_pod,
                        -- POD (Non-Surabaya) - BURUH
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' AND (
                            cdc.description LIKE '%Buruh%'
                        ) THEN cdc.cost ELSE 0 END) AS buruh_pod,
                        -- POD (Non-Surabaya) - BI. LAIN (biaya lain-lain yang tidak termasuk kategori di atas)
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' AND NOT (
                            cdc.description LIKE '%THC%' OR
                            cdc.description LIKE '%Lolo%' OR
                            cdc.description LIKE '%Relokasi%' OR
                            cdc.description LIKE '%Rekolasi%' OR
                            cdc.description LIKE '%Trucking%' OR
                            cdc.description LIKE '%Truck%' OR
                            cdc.description LIKE '%Dooring%' OR
                            cdc.description LIKE '%Door%' OR
                            cdc.description LIKE '%Forklift%' OR
                            cdc.description LIKE '%Forklif%' OR
                            cdc.description LIKE '%Buruh%'
                        ) THEN cdc.cost ELSE 0 END) AS bi_lain_pod,
                        -- Total POD
                        SUM(CASE WHEN cdc.delivery <> 'Surabaya' THEN cdc.cost ELSE 0 END) AS total_biaya_pod,
                        -- Grand Total
                        SUM(cdc.cost) AS total_biaya
                    FROM container_delivery_costs cdc
                    GROUP BY cdc.container_id
                )
                SELECT
                    k.etd_sub AS 'ETD SUB',
                    c.ref_joa AS 'NO JOA',
                    SUM(COALESCE(d.thc_pol,0)) AS 'THC, LOLO, SEAL, DOC, CLEANING POL',
                    SUM(COALESCE(d.freight_pol,0)) AS 'Freight LSS POL',
                    SUM(COALESCE(d.trucking_pol,0)) AS 'TRUCKING POL',
                    SUM(COALESCE(d.ops_pol,0)) AS 'OPS POL',
                    SUM(COALESCE(d.bi_lain_pol,0)) AS 'BI. LAIN POL',
                    SUM(COALESCE(d.pph_pol,0)) AS 'PPH POL',
                    SUM(COALESCE(d.ppn_pol,0)) AS 'PPN POL',
                    SUM(COALESCE(d.total_biaya_pol,0)) AS 'TOTAL BIAYA POL',
                    SUM(COALESCE(d.thc_pod,0)) AS 'THC, LOLO, RELOKASI POD',
                    SUM(COALESCE(d.trucking_dooring_pod,0)) AS 'TRUCKING, DOORING POD',
                    SUM(COALESCE(d.forklift_pod,0)) AS 'FORKLIF POD',
                    SUM(COALESCE(d.buruh_pod,0)) AS 'BURUH POD',
                    SUM(COALESCE(d.bi_lain_pod,0)) AS 'BI. LAIN POD',
                    SUM(COALESCE(d.total_biaya_pod,0)) AS 'TOTAL BIAYA POD',
                    SUM(COALESCE(d.total_biaya,0)) AS 'TOTAL BIAYA',
                    SUM(COALESCE(inv.total_invoice,0)) AS 'NILAI INVOICE'
                FROM containers c
                LEFT JOIN kapals k ON c.kapal_id = k.kapal_id
                LEFT JOIN delivery_per_container d ON c.container_id = d.container_id
                LEFT JOIN invoice_per_container inv ON c.container_id = inv.container_id
                WHERE (? IS NULL OR date(k.etd_sub) >= date(?))
                  AND (? IS NULL OR date(k.etd_sub) <= date(?))
                GROUP BY c.ref_joa, k.etd_sub
                ORDER BY date(k.etd_sub) DESC;
            """
            
            params = (start_date, start_date, end_date, end_date)
            rows = [dict(r) for r in self.db.execute(query, params)]
            
            # Bersihkan tabel
            self.tree.delete(*self.tree.get_children())
            
            total_profit_all = 0
            
            for i, row in enumerate(rows):
                total_biaya = row.get("TOTAL BIAYA", 0) or 0
                nilai_invoice = row.get("NILAI INVOICE", 0) or 0
                profit = nilai_invoice - total_biaya
                row["PROFIT"] = profit
                total_profit_all += profit
                
                # ✅ FORMAT ETD SUB ke format Indonesia
                etd_sub = row.get("ETD SUB", '-')
                if etd_sub and etd_sub != '-':
                    row["ETD SUB"] = self.format_date_indonesian(etd_sub)
                
                formatted = [
                    f"{round(val):,}".replace(",", ".") if isinstance(val, (int, float)) else val
                    for val in row.values()
                ]
                
                self.tree.insert(
                    "", "end",
                    values=formatted,
                    tags=("evenrow" if i % 2 == 0 else "oddrow",)
                )
            
            self.total_label.config(
                text=f"TOTAL PROFIT: Rp {round(total_profit_all):,}".replace(",", ".")
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memuat data dari database:\n{e}")
    
    def filter_data(self):
        start_date = self.start_date.get_date().strftime("%Y-%m-%d")
        end_date = self.end_date.get_date().strftime("%Y-%m-%d")
        print("DEBUG Filter Date:", start_date, end_date)
        self.load_data_from_db(start_date, end_date)
    
    def clear_filter(self):
        self.start_date.set_date(date.today() - timedelta(days=30))
        self.end_date.set_date(date.today())
        self.load_data_from_db()
    
    # ============================================================
    # EXPORT TO EXCEL
    # ============================================================
    
    def export_to_excel(self):
        """Export data tabel ke Excel dengan formatting"""
        try:
            # Check apakah ada data
            if not self.tree.get_children():
                messagebox.showwarning("Peringatan", "Tidak ada data untuk diekspor!")
                return
            
            # Ask save location
            start_date_str = self.start_date.get_date().strftime("%Y%m%d")
            end_date_str = self.end_date.get_date().strftime("%Y%m%d")
            default_filename = f"Lifting_Report_{start_date_str}_to_{end_date_str}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                title="Simpan Laporan Lifting"
            )
            
            if not file_path:
                return  # User cancelled
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lifting Report"
            
            # Page setup for printing
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Styles
            header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            title_font = Font(name='Arial', size=14, bold=True)
            title_alignment = Alignment(horizontal='center', vertical='center')
            
            normal_font = Font(name='Arial', size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            current_row = 1
            ws.merge_cells(f'A{current_row}:S{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "📦 LAPORAN LIFTING (BIAYA POL & POD)"
            title_cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
            title_cell.alignment = title_alignment
            title_cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # Filter info
            ws.merge_cells(f'A{current_row}:S{current_row}')
            filter_cell = ws[f'A{current_row}']
            start_date_display = self.start_date.get_date().strftime("%d/%m/%Y")
            end_date_display = self.end_date.get_date().strftime("%d/%m/%Y")
            filter_cell.value = f"Periode: {start_date_display} s/d {end_date_display}"
            filter_cell.font = Font(name='Arial', size=10, italic=True)
            filter_cell.alignment = center_align
            ws.row_dimensions[current_row].height = 18
            current_row += 2
            
            # Headers
            columns = [
                "ETD SUB", "NO JOA",
                "THC, LOLO, SEAL, DOC, CLEANING POL",
                "Freight LSS POL",
                "TRUCKING POL",
                "OPS POL",
                "BI. LAIN POL",
                "PPH POL", "PPN POL", "TOTAL BIAYA POL",
                "THC, LOLO, RELOKASI POD",
                "TRUCKING, DOORING POD",
                "FORKLIF POD",
                "BURUH POD",
                "BI. LAIN POD",
                "TOTAL BIAYA POD",
                "TOTAL BIAYA", "NILAI INVOICE", "PROFIT"
            ]
            
            for col_idx, header in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.row_dimensions[current_row].height = 20
            
            current_row += 1
            
            # Data rows
            total_profit = 0
            
            for item_id in self.tree.get_children():
                values = self.tree.item(item_id)['values']
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    
                    # Parse nilai (hapus format ribuan)
                    if isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit():
                        cell.value = int(value.replace('.', ''))
                        cell.number_format = '#,##0'
                        cell.alignment = right_align
                    else:
                        cell.value = value
                        if col_idx <= 2:  # ETD SUB dan NO JOA
                            cell.alignment = center_align
                        else:
                            cell.alignment = right_align
                    
                    cell.font = normal_font
                    cell.border = thin_border
                
                # Track profit for total
                try:
                    profit_str = str(values[-1]).replace('.', '')
                    total_profit += int(profit_str)
                except:
                    pass
                
                ws.row_dimensions[current_row].height = 18
                current_row += 1
            
            # Total row
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                
                if col_idx == 1:
                    cell.value = "TOTAL"
                    cell.alignment = center_align
                elif col_idx == len(columns):  # Profit column
                    cell.value = total_profit
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                else:
                    cell.alignment = right_align
            
            ws.row_dimensions[current_row].height = 22
            
            # Column widths
            column_widths = {
                1: 12,   # ETD SUB
                2: 15,   # NO JOA
                3: 20,   # THC, LOLO, SEAL, DOC, CLEANING POL
                4: 15,   # Freight LSS POL
                5: 13,   # TRUCKING POL
                6: 11,   # OPS POL
                7: 14,   # BI. LAIN POL
                8: 11,   # PPH POL
                9: 11,   # PPN POL
                10: 16,  # TOTAL BIAYA POL
                11: 20,  # THC, LOLO, RELOKASI POD
                12: 17,  # TRUCKING, DOORING POD
                13: 13,  # FORKLIF POD
                14: 12,  # BURUH POD
                15: 13,  # BI. LAIN POD
                16: 16,  # TOTAL BIAYA POD
                17: 14,  # TOTAL BIAYA
                18: 15,  # NILAI INVOICE
                19: 13   # PROFIT
            }
            
            for col_idx, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Save file
            wb.save(file_path)
            
            messagebox.showinfo(
                "Sukses",
                f"✅ Laporan Lifting berhasil diekspor!\n\n"
                f"📁 Lokasi: {file_path}\n\n"
                f"📊 Total data: {len(self.tree.get_children())} baris\n"
                f"💰 Total Profit: Rp {total_profit:,.0f}".replace(',', '.')
            )
            
            # Ask to open file
            if messagebox.askyesno("Buka File?", "Apakah Anda ingin membuka file Excel sekarang?"):
                try:
                    import os
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS/Linux
                        os.system(f'open "{file_path}"')
                except Exception as e:
                    messagebox.showwarning("Info", f"File berhasil disimpan, tapi gagal membuka otomatis.\nSilakan buka manual: {file_path}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal mengekspor ke Excel:\n{str(e)}")