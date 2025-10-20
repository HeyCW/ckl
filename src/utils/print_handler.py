import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from src.widget.pdf_packing_list_generator import PDFPackingListGenerator
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from PIL import Image as PILImage, ImageTk
import io
import base64
import traceback

class PrintHandler:
    def __init__(self, db):
        self.pdf_generator = PDFPackingListGenerator(db)
        self.db = db
   
    def print_container_invoice(self, container_id):
        """Generate and export container invoice to Excel optimized for printing"""
        try:
            # Get container details
            container = self.db.get_container_by_id(container_id)
            if not container:
                messagebox.showerror("Error", "Container tidak ditemukan!")
                return
           
            # Get barang in container with pricing
            container_barang = self.db.get_barang_in_container_with_colli_and_pricing(container_id)
           
            if not container_barang:
                messagebox.showwarning("Peringatan", "Container kosong, tidak ada yang akan diprint!")
                return
           
            # Generate Excel invoice
            self._generate_excel_invoice_optimized(container, container_barang, container_id)
           
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat invoice: {str(e)}")
               
   
    def _generate_excel_invoice_optimized(self, container, barang_list, container_id):
        """Generate Excel invoice document optimized for A4 printing with profit calculation and tax as separate rows"""
        try:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoice Packing List"
        
            # SET PAGE LAYOUT FOR A4 PRINT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Set margins
            ws.page_margins = PageMargins(
                left=0.5, right=0.5, top=0.75, bottom=0.75, 
                header=0.3, footer=0.3
            )
            
            ws.print_options.horizontalCentered = True
        
            # Define optimized styles - ALL TIMES NEW ROMAN SIZE 12
            header_font = Font(name='Times New Roman', size=12, bold=True)
            company_font = Font(name='Times New Roman', size=12, bold=True)
            normal_font = Font(name='Times New Roman', size=12)
            small_font = Font(name='Times New Roman', size=12)
            table_header_font = Font(name='Times New Roman', size=12, bold=True)
            profit_header_font = Font(name='Times New Roman', size=12, bold=True)
            profit_font = Font(name='Times New Roman', size=12)
            tax_font = Font(name='Times New Roman', size=12, italic=True)
        
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
            # Safe way to get container values
            def safe_get(key, default='-'):
                try:
                    if hasattr(container, 'get'):
                        return container.get(key, default) or default
                    else:
                        return container[key] if key in container and container[key] else default
                except:
                    return default
        
            # Header section - COMPACT
            current_row = 1
        
            # Title
            ws.merge_cells(f'A{current_row}:L{current_row}')
            ws[f'A{current_row}'] = "INVOICE PACKING LIST"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].alignment = center_align
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
            # Company info - UPDATED
            ws.merge_cells(f'A{current_row}:L{current_row}')
            ws[f'A{current_row}'] = "PT. CAHAYA KARUNIA LOGISTIK | Jl Teluk Bone Selatan No 05 | Phone: 031-60166017"
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].alignment = center_align
            ws.row_dimensions[current_row].height = 18
            current_row += 2
        
            # Container information - 3 columns layout
            container_info = [
                ("Container No", safe_get('container')),
                ("Feeder", safe_get('feeder')),
                ("Destination", safe_get('destination')),
                ("Party", safe_get('party')),
                ("ETD Sub", safe_get('etd_sub')),
                ("CLS", safe_get('cls')),
                ("Open", safe_get('open')),
                ("Full", safe_get('full')),
                ("Seal", safe_get('seal')),
                ("Ref JOA", safe_get('ref_joa'))
            ]

            info_start_row = current_row
            
            for i, (label, value) in enumerate(container_info):
                col_group = i // 4
                row_in_group = i % 4
                
                row = info_start_row + row_in_group
                col_start = col_group * 4 + 1
                
                ws[f'{get_column_letter(col_start)}{row}'] = f"{label}:"
                ws[f'{get_column_letter(col_start)}{row}'].font = small_font
                ws[f'{get_column_letter(col_start + 1)}{row}'] = str(value)
                ws[f'{get_column_letter(col_start + 1)}{row}'].font = small_font
                
                ws.row_dimensions[row].height = 18
            
            current_row = info_start_row + 4 + 1
        
            # Table headers
            headers = ['Tgl', 'Pengirim', 'Penerima', 'Nama Barang', 'Kubikasi', 'M3', 'Ton', 'Col', 'Satuan', 'Door', 'Unit Price', 'Price']
        
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border
                ws.row_dimensions[current_row].height = 20
        
            current_row += 1
        
            # Get tax information by receiver and container
            print("\n=== GETTING TAX INFORMATION BY RECEIVER ===")
            tax_summary = {}
            try:
                # Get all unique receivers first
                receivers = set()
                for barang_row in barang_list:
                    barang = dict(barang_row)
                    receiver = barang.get('receiver_name', '')
                    if receiver:
                        receivers.add(receiver)
                
                # Get aggregated tax info for each receiver in this container
                for receiver in receivers:
                    tax_query = """
                        SELECT 
                            bt.penerima,
                            AVG(bt.ppn_rate) as avg_ppn_rate,
                            AVG(bt.pph23_rate) as avg_pph_rate,
                            SUM(bt.ppn_amount) as total_ppn_amount,
                            SUM(bt.pph23_amount) as total_pph_amount,
                            SUM(bt.total_nilai_barang) as total_nilai
                        FROM barang_tax bt
                        WHERE bt.container_id = ? AND bt.penerima = ?
                        GROUP BY bt.penerima
                    """
                    result = self.db.execute(tax_query, (container_id, receiver))
                    
                    if result and len(result) > 0:
                        row = result[0]
                        tax_summary[receiver] = {
                            'ppn_rate': (row['avg_ppn_rate'] or 0) * 100,
                            'pph_rate': (row['avg_pph_rate'] or 0) * 100,
                            'ppn_amount': row['total_ppn_amount'] or 0,
                            'pph_amount': row['total_pph_amount'] or 0,
                            'has_tax': True
                        }
                        print(f"  {receiver}: PPN {tax_summary[receiver]['ppn_rate']:.1f}% = Rp {tax_summary[receiver]['ppn_amount']:,.0f}, PPH {tax_summary[receiver]['pph_rate']:.1f}% = Rp {tax_summary[receiver]['pph_amount']:,.0f}")
                    else:
                        tax_summary[receiver] = {
                            'ppn_rate': 0,
                            'pph_rate': 0,
                            'ppn_amount': 0,
                            'pph_amount': 0,
                            'has_tax': False
                        }
                        print(f"  {receiver}: No tax records found")
                        
            except Exception as e:
                print(f"Error getting tax info: {e}")
                for barang_row in barang_list:
                    barang = dict(barang_row)
                    receiver = barang.get('receiver_name', '')
                    if receiver and receiver not in tax_summary:
                        tax_summary[receiver] = {'ppn_rate': 0, 'pph_rate': 0, 'ppn_amount': 0, 'pph_amount': 0, 'has_tax': False}
            
            print("\n=== PROCESSING ITEMS ===")
            print(f"Total items: {len(barang_list)}")
            print("=" * 60)
        
            # Table data with GROUPING
            total_m3 = 0
            total_ton = 0
            total_colli = 0
            total_nilai = 0
            
            previous_date = None
            previous_pengirim = None
            previous_penerima = None
            
            # Variables for tax row insertion
            total_ppn_all = 0
            total_pph_all = 0
            
            # Track receiver groups for tax calculation
            # PENTING: Hanya hitung barang dengan pajak = 1
            receiver_groups = []
            current_receiver = None
            current_receiver_subtotal = 0
            current_receiver_start = 0
            
            # Function to insert tax rows for a receiver
            def insert_tax_rows_for_receiver(receiver, subtotal, current_row):
                nonlocal total_ppn_all, total_pph_all
                
                if receiver not in tax_summary:
                    return current_row
                    
                tax_data = tax_summary[receiver]
                if not tax_data['has_tax'] or (tax_data['ppn_rate'] == 0 and tax_data['pph_rate'] == 0):
                    return current_row
                
                # Skip tax calculation if subtotal is 0 or negative
                if subtotal <= 0:
                    print(f"  ⚠ Skipping tax for {receiver}: Subtotal is {subtotal}")
                    return current_row
                
                # Calculate tax based on subtotal (only from items with pajak=1)
                ppn_amount = subtotal * (tax_data['ppn_rate'] / 100)
                pph_amount = subtotal * (tax_data['pph_rate'] / 100)
                
                print(f"  ✓ Inserting tax for '{receiver}': Taxable Subtotal Rp {subtotal:,.0f}")
                print(f"    - PPN {tax_data['ppn_rate']:.1f}% = Rp {ppn_amount:,.0f}")
                print(f"    - PPH {tax_data['pph_rate']:.1f}% = Rp {pph_amount:,.0f}")
                
                # Add PPN row
                if tax_data['ppn_rate'] > 0:
                    ppn_row_data = ['', '', '', f"PPN {tax_data['ppn_rate']:.1f}%", '', '', '', '', '', '', '', ppn_amount]
                    
                    for col, value in enumerate(ppn_row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = tax_font
                        ws.row_dimensions[current_row].height = 18
                        
                        if col == 1:
                            cell.alignment = center_align
                        elif col == 12:
                            cell.alignment = right_align
                            cell.number_format = '#,##0'
                        else:
                            cell.alignment = left_align
                    
                    total_ppn_all += ppn_amount
                    current_row += 1
                
                # Add PPH row - WITH MINUS SIGN
                if tax_data['pph_rate'] > 0:
                    pph_row_data = ['', '', '', f"PPH {23}", '', '', '', '', '', '', '', -pph_amount]  # NEGATIVE VALUE
                    
                    for col, value in enumerate(pph_row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = tax_font
                        ws.row_dimensions[current_row].height = 18
                        
                        if col == 1:
                            cell.alignment = center_align
                        elif col == 12:
                            cell.alignment = right_align
                            cell.number_format = '#,##0'
                        else:
                            cell.alignment = left_align
                    
                    total_pph_all += pph_amount
                    current_row += 1
                
                return current_row
        
            # Process each item
            for i, barang_row in enumerate(barang_list, 1):
                barang = dict(barang_row)

                try:
                    def safe_barang_get(key, default='-'):
                        try:
                            value = barang.get(key, default)
                            return value if value not in [None, '', 'NULL', 'null'] else default
                        except Exception:
                            return default
                
                    # Format date - use 'tanggal' column (DATE)
                    tanggal = safe_barang_get('tanggal', datetime.now())
                    
                    # Parse string to datetime if needed
                    if isinstance(tanggal, str):
                        try:
                            tanggal = datetime.strptime(tanggal, '%Y-%m-%d')
                        except:
                            try:
                                tanggal = datetime.strptime(tanggal, '%Y-%m-%d %H:%M:%S')
                            except:
                                tanggal = datetime.now()
                    
                    formatted_date = tanggal.strftime('%d-%b') if isinstance(tanggal, datetime) else '-'
                    
                    pengirim = str(safe_barang_get('sender_name', '-'))
                    penerima = str(safe_barang_get('receiver_name', '-'))
                    nama_barang = str(safe_barang_get('nama_barang', '-'))
                    
                    # FIX: Door type to uppercase
                    door = str(safe_barang_get('door_type', safe_barang_get('alamat_tujuan', '-')))[:10].upper()

                    # Format dimensions
                    p = safe_barang_get('panjang_barang', '-')
                    l = safe_barang_get('lebar_barang', '-')
                    t = safe_barang_get('tinggi_barang', '-')
                    
                    if str(p) != '-':
                        p = str(p).replace(',', '')
                    if str(l) != '-':
                        l = str(l).replace(',', '')
                    if str(t) != '-':
                        t = str(t).replace(',', '')
                    
                    kubikasi = f"{p}×{l}×{t}"
                
                    m3 = safe_barang_get('m3_barang', 0)
                    ton = safe_barang_get('ton_barang', 0)
                    colli = safe_barang_get('colli_amount', 0)
                    satuan = str(safe_barang_get('satuan', safe_barang_get('unit', 'pcs')))
                    unit_price = safe_barang_get('harga_per_unit', safe_barang_get('unit_price', 0))
                    total_harga = safe_barang_get('total_harga', 0)
                    
                    # ===== CRITICAL: Get pajak flag =====
                    pajak_flag = safe_barang_get('pajak', 0)
                    try:
                        pajak_flag = int(pajak_flag) if pajak_flag not in [None, '', '-'] else 0
                    except:
                        pajak_flag = 0
                
                    # Format values
                    m3_val = float(m3) if m3 not in [None, '', '-'] else 0
                    ton_val = float(ton) if ton not in [None, '', '-'] else 0
                    colli_val = int(colli) if colli not in [None, '', '-'] else 0
                    unit_price_val = float(unit_price) if unit_price not in [None, '', '-'] else 0
                    harga_val = float(total_harga) if total_harga not in [None, '', '-'] else 0
                    
                    # M3 dan Ton dikali dengan colli
                    m3_total = m3_val * colli_val if colli_val > 0 else m3_val
                    ton_total = ton_val * colli_val if colli_val > 0 else ton_val
                    
                    # Add to totals
                    try:
                        total_m3 += m3_total
                        total_ton += ton_total
                        total_colli += int(colli) if colli not in [None, '', '-'] else 0
                        total_nilai += float(total_harga) if total_harga not in [None, '', '-'] else 0
                    except (ValueError, TypeError):
                        pass
                    
                    # ===== TRACK RECEIVER CHANGES - HANYA UNTUK BARANG DENGAN PAJAK = 1 =====
                    if current_receiver is None:
                        # First item
                        current_receiver = penerima
                        current_receiver_subtotal = harga_val if pajak_flag == 1 else 0
                        current_receiver_start = i
                        
                        if pajak_flag == 1:
                            print(f"  [{i}] 🏁 Starting group: '{penerima}' | Taxable item: Rp {harga_val:,.0f}")
                        else:
                            print(f"  [{i}] 🏁 Starting group: '{penerima}' | Non-taxable item (pajak=0)")
                        
                    elif current_receiver != penerima:
                        # Receiver changed
                        print(f"\n  [{i}] ═══ RECEIVER CHANGED ═══")
                        print(f"      From: '{current_receiver}' → To: '{penerima}'")
                        print(f"      Taxable subtotal for '{current_receiver}': Rp {current_receiver_subtotal:,.0f}")
                        
                        # Save completed group
                        receiver_groups.append((current_receiver, current_receiver_subtotal, current_receiver_start, i-1))
                        
                        # Insert tax rows for previous receiver
                        current_row = insert_tax_rows_for_receiver(current_receiver, current_receiver_subtotal, current_row)
                        
                        # Start new receiver group
                        current_receiver = penerima
                        current_receiver_subtotal = harga_val if pajak_flag == 1 else 0
                        current_receiver_start = i
                        
                        if pajak_flag == 1:
                            print(f"      Starting new group: '{penerima}' | Taxable item: Rp {harga_val:,.0f}")
                        else:
                            print(f"      Starting new group: '{penerima}' | Non-taxable item (pajak=0)")
                        
                    else:
                        # Same receiver - ONLY add to subtotal if pajak = 1
                        if pajak_flag == 1:
                            current_receiver_subtotal += harga_val
                            print(f"  [{i}] ✓ Same receiver '{penerima}' | Taxable +Rp {harga_val:,.0f} | Subtotal: Rp {current_receiver_subtotal:,.0f}")
                        else:
                            print(f"  [{i}] ⊘ Same receiver '{penerima}' | Non-taxable Rp {harga_val:,.0f} (pajak=0) | Subtotal: Rp {current_receiver_subtotal:,.0f}")
                    
                    # GROUPING LOGIC for display
                    display_date = formatted_date
                    display_pengirim = pengirim
                    display_penerima = penerima
                    
                    if (formatted_date == previous_date and 
                        pengirim == previous_pengirim and 
                        penerima == previous_penerima):
                        display_date = ""
                        display_pengirim = ""
                        display_penerima = ""
                    else:
                        previous_date = formatted_date
                        previous_pengirim = pengirim
                        previous_penerima = penerima
                
                    # Fill row data
                    row_data = [
                        display_date, display_pengirim, display_penerima, nama_barang,
                        kubikasi, m3_total, ton_total, colli_val, satuan,
                        door, unit_price_val, harga_val
                    ]
                
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = small_font
                        ws.row_dimensions[current_row].height = 25
                    
                        if col == 1:
                            cell.alignment = center_align
                        elif col in [6, 7, 8, 11, 12]:
                            cell.alignment = right_align
                            if col in [11, 12]:
                                cell.number_format = '#,##0'
                            elif col in [6, 7]:
                                cell.number_format = '0.0000'
                        else:
                            cell.alignment = left_align
                
                    current_row += 1
                    
                except Exception as e:
                    print(f"Error processing barang {i}: {e}")
                    continue
            
            # Insert tax for the last receiver group
            if current_receiver is not None:
                print(f"\n  ═══ PROCESSING LAST GROUP ═══")
                print(f"      Receiver: '{current_receiver}'")
                print(f"      Final taxable subtotal: Rp {current_receiver_subtotal:,.0f}")
                
                receiver_groups.append((current_receiver, current_receiver_subtotal, current_receiver_start, len(barang_list)))
                current_row = insert_tax_rows_for_receiver(current_receiver, current_receiver_subtotal, current_row)
        
            # Summary of receiver groups
            print(f"\n=== RECEIVER GROUPS SUMMARY ===")
            for receiver, subtotal, start, end in receiver_groups:
                print(f"  {receiver}: Items {start}-{end} | Taxable Subtotal: Rp {subtotal:,.0f}")
            print(f"  Total PPN: Rp {total_ppn_all:,.0f}")
            print(f"  Total PPH: Rp {total_pph_all:,.0f}")
            print("=" * 60)
        
            # Total row
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.font = table_header_font
                ws.row_dimensions[current_row].height = 20
            
                if col == 1:
                    cell.value = "TOTAL"
                    cell.alignment = center_align
                elif col == 6:
                    cell.value = total_m3
                    cell.number_format = '0.0000'
                    cell.alignment = right_align
                elif col == 7:
                    cell.value = total_ton
                    cell.number_format = '0.0000'
                    cell.alignment = right_align
                elif col == 8:
                    cell.value = total_colli
                    cell.alignment = right_align
                elif col == 12:
                    # PPH23 dikurangkan, PPN ditambahkan
                    cell.value = total_nilai + total_ppn_all - total_pph_all
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
        
            current_row += 3
        
            # PROFIT CALCULATION SECTION
            try:
                costs_surabaya = self._get_container_costs(container_id, 'Surabaya')
                costs_destinasi = self._get_container_costs(container_id, container.get('destination', 'Samarinda'))
            except Exception as e:
                print(f"Error getting costs from database: {e}")
                costs_surabaya = None
                costs_destinasi = None

            if costs_surabaya or costs_destinasi:
                if not costs_surabaya:
                    costs_surabaya = {}
                if not costs_destinasi:
                    costs_destinasi = {}
                
                def safe_sum_costs(costs_dict):
                    if not costs_dict:
                        return 0
                    total = 0
                    for value in costs_dict.values():
                        try:
                            if isinstance(value, (tuple, list)):
                                total += float(value[0]) if len(value) > 0 else 0
                            else:
                                total += float(value) if value else 0
                        except (ValueError, TypeError):
                            continue
                    return total
                
                total_biaya_surabaya = safe_sum_costs(costs_surabaya)
                total_biaya_samarinda = safe_sum_costs(costs_destinasi)
                total_biaya = total_biaya_surabaya + total_biaya_samarinda
                # PPH23 dikurangkan dari total nilai, PPN ditambahkan
                profit_lcl = (total_nilai + total_ppn_all - total_pph_all) - total_biaya
                
                # BIAYA SURABAYA SECTION
                if costs_surabaya:
                    ws[f'A{current_row}'] = "Biaya Surabaya"
                    ws[f'L{current_row}'] = "Cost (Rp)"
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'].font = profit_header_font
                    ws[f'A{current_row}'].alignment = left_align
                    ws[f'A{current_row}'].border = thin_border
                    ws[f'L{current_row}'].font = profit_header_font
                    ws[f'L{current_row}'].alignment = right_align
                    ws[f'L{current_row}'].border = thin_border
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1
                    
                    for cost_name, cost_value in costs_surabaya.items():
                        ws[f'A{current_row}'] = cost_name
                        ws[f'A{current_row}'].font = profit_font
                        ws[f'A{current_row}'].alignment = left_align
                        
                        try:
                            if isinstance(cost_value, (tuple, list)) and len(cost_value) >= 2:
                                cost_desc = str(cost_value[1])
                                cost_amount = float(cost_value[0])
                            else:
                                cost_desc = ""
                                cost_amount = float(cost_value) if cost_value else 0
                        except (ValueError, TypeError):
                            cost_desc = ""
                            cost_amount = 0

                        ws[f'B{current_row}'] = cost_desc
                        ws[f'B{current_row}'].font = profit_font
                        ws[f'B{current_row}'].alignment = left_align

                        ws[f'L{current_row}'] = cost_amount
                        ws[f'L{current_row}'].font = profit_font
                        ws[f'L{current_row}'].alignment = right_align
                        ws[f'L{current_row}'].number_format = '#,##0'
                        ws.row_dimensions[current_row].height = 18
                        current_row += 1

                    ws[f'L{current_row}'] = total_biaya_surabaya
                    ws[f'L{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
                    ws[f'L{current_row}'].alignment = right_align
                    ws[f'L{current_row}'].number_format = '#,##0'
                    ws[f'L{current_row}'].border = thin_border
                    ws.row_dimensions[current_row].height = 20
                    current_row += 2
                
                # BIAYA DESTINASI SECTION
                if costs_destinasi:
                    destination_name = container.get('destination', 'Destinasi')
                    ws[f'A{current_row}'] = f"Biaya {destination_name}"
                    ws[f'L{current_row}'] = "Cost (Rp)"
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'].font = profit_header_font
                    ws[f'A{current_row}'].alignment = left_align
                    ws[f'A{current_row}'].border = thin_border
                    ws[f'L{current_row}'].font = profit_header_font
                    ws[f'L{current_row}'].alignment = right_align
                    ws[f'L{current_row}'].border = thin_border
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1
                    
                    for cost_name, cost_value in costs_destinasi.items():
                        ws[f'A{current_row}'] = cost_name
                        ws[f'A{current_row}'].font = profit_font
                        ws[f'A{current_row}'].alignment = left_align
                        
                        try:
                            if isinstance(cost_value, (tuple, list)) and len(cost_value) >= 2:
                                cost_desc = str(cost_value[1])
                                cost_amount = float(cost_value[0])
                            else:
                                cost_desc = ""
                                cost_amount = float(cost_value) if cost_value else 0
                        except (ValueError, TypeError):
                            cost_desc = ""
                            cost_amount = 0

                        ws[f'B{current_row}'] = cost_desc
                        ws[f'B{current_row}'].font = profit_font
                        ws[f'B{current_row}'].alignment = left_align

                        ws[f'L{current_row}'] = cost_amount
                        ws[f'L{current_row}'].font = profit_font
                        ws[f'L{current_row}'].alignment = right_align
                        ws[f'L{current_row}'].number_format = '#,##0'
                        ws.row_dimensions[current_row].height = 18
                        current_row += 1
                    
                    ws[f'L{current_row}'] = total_biaya_samarinda
                    ws[f'L{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
                    ws[f'L{current_row}'].alignment = right_align
                    ws[f'L{current_row}'].number_format = '#,##0'
                    ws[f'L{current_row}'].border = thin_border
                    ws.row_dimensions[current_row].height = 20
                    current_row += 2
                
                # PROFIT CALCULATION
                if costs_surabaya or costs_destinasi:
                    ws[f'A{current_row}'] = "PROFIT LCL"
                    ws[f'L{current_row}'] = profit_lcl
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
                    ws[f'A{current_row}'].alignment = left_align
                    ws[f'A{current_row}'].border = thin_border
                    ws[f'L{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
                    ws[f'L{current_row}'].alignment = right_align
                    ws[f'L{current_row}'].number_format = '#,##0'
                    ws[f'L{current_row}'].border = thin_border
                    ws.row_dimensions[current_row].height = 20
                    current_row += 2
        
            # AUTO-ADJUST Column widths
            estimated_widths = [
                10,  # Tgl
                25,  # Pengirim (diperbesar)
                25,  # Penerima  
                35,  # Nama Barang (diperbesar)
                18,  # Kubikasi (diperbesar)
                10,  # M3
                10,  # Ton
                8,   # Col
                10,  # Satuan
                12,  # Door
                14,  # Unit Price
                14   # Price
            ]
            
            for i, width in enumerate(estimated_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
        
            # Set print area
            last_row = current_row
            ws.print_area = f'A1:L{last_row}'
        
            # Add page breaks if needed
            if len(barang_list) > 35:
                header_rows = 11
                for i in range(35, len(barang_list), 35):
                    break_row = header_rows + i
                    ws.row_breaks.append(break_row)
        
            # Save file
            filename = f"Invoice_Container_{container_id}_TaxFiltered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self._save_excel_file(wb, filename, "INVOICE CONTAINER WITH TAX FILTERING")
        
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat Excel invoice: {str(e)}")
            print(f"Full error: {traceback.format_exc()}")
            
                           
    def _get_container_costs(self, container_id, location):
        """Get container costs from database by location (SURABAYA or SAMARINDA)"""
        try:
            result = self.db.execute("""
                SELECT description, cost_description, cost 
                FROM container_delivery_costs 
                WHERE container_id = ? AND delivery = ?
                ORDER BY id
            """, (container_id, location))
            
            costs_dict = {}
            for row in result:
                key = row['description']
                value = (float(row['cost']), row['cost_description'])
                costs_dict[key] = value
            
            return costs_dict if costs_dict else None
            
        except Exception as e:
            print(f"Error getting container costs: {e}")
            return None
  
                           
    def _get_container_costs(self, container_id, location):
        """Get container costs from database by location (SURABAYA or SAMARINDA)"""
        try:
            result = self.db.execute("""
                SELECT description, cost_description, cost 
                FROM container_delivery_costs 
                WHERE container_id = ? AND delivery = ?
                ORDER BY id
            """, (container_id, location))
            
            costs_dict = {}
            for row in result:
                key = row['description']
                value = (float(row['cost']), row['cost_description'])
                costs_dict[key] = value
            
            return costs_dict if costs_dict else None
            
        except Exception as e:
            print(f"Error getting container costs: {e}")
            return None

    def print_customer_packing_list(self, container_id, filter_criteria=None):
        """Generate and export packing list to Excel optimized for printing"""
        try:
            container = self.db.get_container_by_id(container_id)
            if not container:
                messagebox.showerror("Error", "Container tidak ditemukan!")
                return
           
            container_barang = self.db.get_barang_in_container_with_colli_and_pricing(container_id)
           
            if not container_barang:
                messagebox.showwarning("Peringatan", "Container kosong, tidak ada yang akan diprint!")
                return
           
            self._generate_excel_packing_list_optimized(container, container_barang, container_id, filter_criteria)
           
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat packing list: {str(e)}")
        
    def _save_excel_file(self, workbook, filename, doc_type):
        """Save Excel workbook to file with print preview info"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename,
                title=f"Save {doc_type}"
            )
           
            if file_path:
                workbook.save(file_path)
                messagebox.showinfo(
                    "Sukses",
                    f"📊 {doc_type} berhasil disimpan!\n\n"
                    f"📁 Lokasi: {file_path}\n\n"
                    f"🖨️ OPTIMIZED untuk print A4 Landscape\n"
                    f"💡 Buka Excel → Print Preview untuk melihat hasil print\n"
                    f"⚙️ Pastikan setting printer: A4, Landscape, Fit to 1 page wide"
                )
               
                if messagebox.askyesno("Buka File?", f"Apakah Anda ingin membuka file Excel sekarang?"):
                    try:
                        if os.name == 'nt':
                            os.startfile(file_path)
                        elif os.name == 'posix':
                            os.system(f'open "{file_path}"')
                    except Exception as e:
                        messagebox.showwarning("Info", f"File berhasil disimpan, tapi gagal membuka otomatis.\nSilakan buka manual: {file_path}")
       
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan file Excel: {str(e)}")

    def show_sender_receiver_selection_dialog(self, container_id):
        """Show dialog to select sender-receiver combinations for packing list"""
        try:
            container_barang = self.db.get_barang_in_container_with_colli_and_pricing(container_id)
            
            if not container_barang:
                messagebox.showwarning("Peringatan", "Container kosong!")
                return
            
            combinations = set()
            for barang in container_barang:
                if hasattr(barang, 'keys'):
                    sender = barang['sender_name'] if 'sender_name' in barang.keys() and barang['sender_name'] else '-'
                    receiver = barang['receiver_name'] if 'receiver_name' in barang.keys() and barang['receiver_name'] else '-'
                else:
                    sender = barang.get('sender_name', '-')
                    receiver = barang.get('receiver_name', '-')
                combinations.add((sender, receiver))
            
            combinations = sorted(list(combinations))
            
            if not combinations:
                messagebox.showwarning("Peringatan", "Tidak ada data pengirim-penerima!")
                return
            
            if len(combinations) == 1:
                sender, receiver = combinations[0]
                filter_criteria = {'sender_name': sender, 'receiver_name': receiver}
                self.print_customer_packing_list(container_id, filter_criteria)
                return
            
            dialog = tk.Toplevel()
            dialog.title("Pilih Pengirim - Penerima")
            dialog.geometry("500x400")
            dialog.resizable(False, False)
            dialog.grab_set()
            dialog.transient()
            
            header_frame = ttk.Frame(dialog)
            header_frame.pack(fill='x', padx=10, pady=5)
            
            ttk.Label(header_frame, text="Pilih kombinasi Pengirim - Penerima untuk Packing List:", 
                    font=('Arial', 10, 'bold')).pack(anchor='w')
            
            list_frame = ttk.Frame(dialog)
            list_frame.pack(fill='both', expand=True, padx=10, pady=5)
            
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side='right', fill='y')
            
            listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, 
                                font=('Arial', 9), height=15)
            listbox.pack(side='left', fill='both', expand=True)
            scrollbar.config(command=listbox.yview)
            
            for i, (sender, receiver) in enumerate(combinations):
                display_text = f"{i+1:2d}. {sender} → {receiver}"
                listbox.insert(tk.END, display_text)
            
            if combinations:
                listbox.selection_set(0)
            
            button_frame = ttk.Frame(dialog)
            button_frame.pack(fill='x', padx=10, pady=10)
            
            selected_combination = None
            
            def on_print():
                nonlocal selected_combination
                selection = listbox.curselection()
                if selection:
                    idx = selection[0]
                    selected_combination = combinations[idx]
                    dialog.destroy()
            
            def on_print_all():
                nonlocal selected_combination
                selected_combination = "ALL"
                dialog.destroy()
            
            def on_cancel():
                dialog.destroy()
            
            ttk.Button(button_frame, text="Print Packing List Terpilih", 
                    command=on_print).pack(side='left', padx=5)
            ttk.Button(button_frame, text="Print Semua (Terpisah)", 
                    command=on_print_all).pack(side='left', padx=5)
            ttk.Button(button_frame, text="Batal", 
                    command=on_cancel).pack(side='right', padx=5)
            
            listbox.bind('<Double-1>', lambda e: on_print())
            
            dialog.wait_window()
            
            if selected_combination:
                if selected_combination == "ALL":
                    for sender, receiver in combinations:
                        filter_criteria = {'sender_name': sender, 'receiver_name': receiver}
                        self.print_customer_packing_list(container_id, filter_criteria)
                else:
                    sender, receiver = selected_combination
                    filter_criteria = {'sender_name': sender, 'receiver_name': receiver}
                    self.print_customer_packing_list(container_id, filter_criteria)
            
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menampilkan dialog pemilihan: {str(e)}")

    def _generate_excel_packing_list_optimized(self, container, barang_list, container_id, filter_criteria=None):
        """Generate Excel packing list document optimized for A4 printing with grouping logic"""
        try:
            if filter_criteria:
                filtered_barang = []
                sender_filter = filter_criteria.get('sender_name', '')
                receiver_filter = filter_criteria.get('receiver_name', '')
                
                for barang in barang_list:
                    if hasattr(barang, 'keys'):
                        sender = barang['sender_name'] if 'sender_name' in barang.keys() and barang['sender_name'] else ''
                        receiver = barang['receiver_name'] if 'receiver_name' in barang.keys() and barang['receiver_name'] else ''
                    else:
                        sender = barang.get('sender_name', '')
                        receiver = barang.get('receiver_name', '')
                    
                    if sender == sender_filter and receiver == receiver_filter:
                        filtered_barang.append(barang)
                
                barang_list = filtered_barang
                
                if not barang_list:
                    messagebox.showwarning("Peringatan", "Tidak ada barang untuk kombinasi yang dipilih!")
                    return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Packing List"
        
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            ws.page_margins = PageMargins(
                left=0.5, right=0.5, top=0.75, bottom=0.75, 
                header=0.3, footer=0.3
            )
            
            ws.print_options.horizontalCentered = True
        
            header_font = Font(name='Arial', size=14, bold=True)
            company_font = Font(name='Arial', size=10, bold=True)
            normal_font = Font(name='Arial', size=8)
            small_font = Font(name='Arial', size=7)
            table_header_font = Font(name='Arial', size=8, bold=True)
        
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
            def safe_get(key, default='-'):
                try:
                    if hasattr(container, 'get'):
                        return container.get(key, default) or default
                    else:
                        return container[key] if key in container and container[key] else default
                except:
                    return default
        
            current_row = 1
        
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "PACKING LIST"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].alignment = center_align
            current_row += 1
        
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "CV. CAHAYA KARUNIA | Jl. Teluk Raya Selatan No. 6 Surabaya | Phone: 031-60166017"
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].alignment = center_align
            current_row += 2
        
            container_info = [
                ("Container No", safe_get('container')),
                ("Feeder", safe_get('feeder')),
                ("Destination", safe_get('destination')),
                ("Party", safe_get('party')),
                ("ETD Sub", safe_get('etd_sub')),
                ("CLS", safe_get('cls'))
            ]
            
            if filter_criteria:
                sender = filter_criteria.get('sender_name', '')
                receiver = filter_criteria.get('receiver_name', '')
                container_info.extend([
                    ("Pengirim", sender),
                    ("Penerima", receiver)
                ])

            info_start_row = current_row
            
            for i, (label, value) in enumerate(container_info):
                col_group = i // 4
                row_in_group = i % 4
                
                row = info_start_row + row_in_group
                col_start = col_group * 3 + 1
                
                ws[f'{get_column_letter(col_start)}{row}'] = f"{label}:"
                ws[f'{get_column_letter(col_start)}{row}'].font = small_font
                ws[f'{get_column_letter(col_start + 1)}{row}'] = str(value)
                ws[f'{get_column_letter(col_start + 1)}{row}'].font = small_font
                
                ws.row_dimensions[row].height = 12
            
            current_row = info_start_row + 4 + 1
        
            headers = ['No', 'Tanggal', 'Pengirim', 'Penerima', 'Nama Barang', 'Jenis Barang', 'Dimensi (cm)', 'M3', 'Ton', 'Colli', 'Satuan', 'Door']
        
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border
                ws.row_dimensions[current_row].height = 12
        
            current_row += 1
        
            def get_sort_key(barang_row):
                try:
                    if hasattr(barang_row, 'keys'):
                        barang = {key: barang_row[key] for key in barang_row.keys()}
                    else:
                        barang = dict(barang_row)
                    
                    def safe_barang_get(key, default='-'):
                        try:
                            value = barang.get(key, default)
                            return value if value not in [None, '', 'NULL', 'null'] else default
                        except Exception:
                            return default
                    
                    tanggal_raw = safe_barang_get('tanggal_barang', datetime.now())
                    
                    if isinstance(tanggal_raw, str):
                        try:
                            tanggal = datetime.strptime(tanggal_raw, '%Y-%m-%d')
                        except:
                            try:
                                tanggal = datetime.strptime(tanggal_raw, '%Y-%m-%d %H:%M:%S')
                            except:
                                tanggal = datetime.now()
                    elif isinstance(tanggal_raw, datetime):
                        tanggal = tanggal_raw
                    else:
                        tanggal = datetime.now()
                    
                    pengirim = str(safe_barang_get('sender_name', '')).strip()
                    penerima = str(safe_barang_get('receiver_name', '')).strip()
                    
                    if not pengirim:
                        pengirim = 'ZZZ_NO_SENDER'
                    if not penerima:
                        penerima = 'ZZZ_NO_RECEIVER'
                    
                    return (tanggal, pengirim, penerima)
                    
                except Exception as e:
                    print(f"ERROR in get_sort_key: {e}")
                    return (datetime.now(), 'ZZZ_ERROR', 'ZZZ_ERROR')
            
            try:
                barang_list = sorted(barang_list, key=get_sort_key)
            except Exception as e:
                print(f"SORTING FAILED: {e}")
            
            total_m3 = 0
            total_ton = 0
            total_colli = 0
            
            previous_date = None
            previous_pengirim = None
            previous_penerima = None
        
            for i, barang_row in enumerate(barang_list, 1):
                if hasattr(barang_row, 'keys'):
                    barang = {key: barang_row[key] for key in barang_row.keys()}
                else:
                    barang = dict(barang_row)
                
                try:
                    def safe_barang_get(key, default='-'):
                        try:
                            value = barang.get(key, default)
                            return value if value not in [None, '', 'NULL', 'null'] else default
                        except Exception:
                            return default
                
                    tanggal = safe_barang_get('tanggal_barang', datetime.now())
                    if isinstance(tanggal, str):
                        try:
                            tanggal = datetime.strptime(tanggal, '%Y-%m-%d')
                        except:
                            tanggal = datetime.now()
                    formatted_date = tanggal.strftime('%d-%b-%Y') if isinstance(tanggal, datetime) else '-'
                    
                    pengirim = str(safe_barang_get('sender_name', '-'))
                    penerima = str(safe_barang_get('receiver_name', '-'))
                    
                    nama_barang = str(safe_barang_get('nama_barang', '-'))
                    jenis_barang = str(safe_barang_get('jenis_barang', '-'))
                    
                    p = safe_barang_get('panjang_barang', '-')
                    l = safe_barang_get('lebar_barang', '-')
                    t = safe_barang_get('tinggi_barang', '-')
                    dimensi = f"{p}×{l}×{t}"
                
                    m3 = safe_barang_get('m3_barang', 0)
                    ton = safe_barang_get('ton_barang', 0)
                    colli = safe_barang_get('colli_amount', 0)
                    satuan = str(safe_barang_get('satuan', safe_barang_get('unit', 'pcs')))
                    door = str(safe_barang_get('door_type', safe_barang_get('alamat_tujuan', '-')))
                
                    try:
                        total_m3 += float(m3) if m3 not in [None, '', '-'] else 0
                        total_ton += float(ton) if ton not in [None, '', '-'] else 0
                        total_colli += int(colli) if colli not in [None, '', '-'] else 0
                    except (ValueError, TypeError):
                        pass
                
                    m3_val = float(m3) if m3 not in [None, '', '-'] else 0
                    ton_val = float(ton) if ton not in [None, '', '-'] else 0
                    colli_val = int(colli) if colli not in [None, '', '-'] else 0
                    
                    display_date = formatted_date
                    display_pengirim = pengirim
                    display_penerima = penerima
                    
                    if (formatted_date == previous_date and 
                        pengirim == previous_pengirim and 
                        penerima == previous_penerima):
                        display_date = ""
                        display_pengirim = ""
                        display_penerima = ""
                    else:
                        previous_date = formatted_date
                        previous_pengirim = pengirim
                        previous_penerima = penerima
                    
                    row_data = [
                        i, display_date, display_pengirim, display_penerima, nama_barang, jenis_barang,
                        dimensi, m3_val, ton_val, colli_val, satuan, door
                    ]
                
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = small_font
                        cell.border = thin_border
                        ws.row_dimensions[current_row].height = 10
                    
                        if col == 1:
                            cell.alignment = center_align
                        elif col in [8, 9, 10]:
                            cell.alignment = right_align
                            if col in [8, 9]:
                                cell.number_format = '0.00'
                        else:
                            cell.alignment = left_align
                
                    current_row += 1
                
                except Exception as e:
                    print(f"Error processing barang {i}: {e}")
                    continue
        
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.font = table_header_font
            
                if col == 1:
                    cell.value = "TOTAL"
                    cell.alignment = center_align
                elif col == 8:
                    cell.value = total_m3
                    cell.number_format = '0.00'
                    cell.alignment = right_align
                elif col == 9:
                    cell.value = total_ton
                    cell.number_format = '0.00'
                    cell.alignment = right_align
                elif col == 10:
                    cell.value = total_colli
                    cell.alignment = right_align
        
            current_row += 2
        
            summary_text = f"Total Items: {len(barang_list)} | Total M3: {total_m3:.2f} | Total Ton: {total_ton:.2f} | Total Colli: {total_colli} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            if filter_criteria:
                summary_text = f"Pengirim: {filter_criteria.get('sender_name', '')} → Penerima: {filter_criteria.get('receiver_name', '')} | " + summary_text
            
            ws[f'A{current_row}'] = summary_text
            ws[f'A{current_row}'].font = small_font
        
            estimated_widths = [
                5,   # No
                12,  # Tanggal
                20,  # Pengirim
                20,  # Penerima
                25,  # Nama Barang
                15,  # Jenis Barang
                15,  # Dimensi
                8,   # M3
                8,   # Ton
                6,   # Colli
                8,   # Satuan
                15   # Door
            ]
            
            for i, width in enumerate(estimated_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
        
            last_row = current_row
            ws.print_area = f'A1:L{last_row}'
        
            filter_suffix = ""
            if filter_criteria:
                sender = filter_criteria.get('sender_name', 'Unknown')[:10]
                receiver = filter_criteria.get('receiver_name', 'Unknown')[:10]
                filter_suffix = f"_{sender}_to_{receiver}"
            
            filename = f"PackingList_Container_{container_id}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self._save_excel_file(wb, filename, "PACKING LIST")
        
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat Excel packing list: {str(e)}")
            print(f"Full error: {traceback.format_exc()}")
            
    def print_customer_packing_list_pdf(self, container_ids, filter_criteria=None):
        """
        Generate PDF packing list - SAME LOGIC as Invoice (by JOA and RECEIVER)
        
        Args:
            container_ids: Can be single int or list of ints
        """
        try:
            # Convert single ID to list
            if isinstance(container_ids, (int, str)):
                container_ids = [int(container_ids)]
            elif not isinstance(container_ids, list):
                container_ids = list(container_ids)
            
            print(f"\n{'='*60}")
            print(f"[PACKING LIST PDF] Processing {len(container_ids)} container(s)")
            print(f"{'='*60}\n")
            
            # Collect data grouped by JOA and RECEIVER ONLY (same as invoice)
            joa_receiver_groups = {}
            container_info = {}
            all_barang = []   # inisialisasi sebelum loop

            for container_id in container_ids:
                try:
                    container = self.db.get_container_by_id(container_id)
                    if not container:
                        print(f"[ERROR] Container {container_id} tidak ditemukan!")
                        continue
                    
                    container_info[container_id] = container
                    
                    ref_joa = container.get('ref_joa', '') or f'NO_JOA_{container_id}'
                    
                    if ref_joa not in joa_receiver_groups:
                        joa_receiver_groups[ref_joa] = {}
                    
                    container_barang = self.db.get_barang_in_container_with_colli_and_pricing(container_id)
                    if not container_barang:
                        print(f"[WARNING] Container {container_id} kosong!")
                        continue

                    # Convert sqlite3.Row → dict + inject container_id
                    container_barang = [{**dict(b), "container_id": container_id} for b in container_barang]

                    # ➕ Kumpulin ke all_barang
                    all_barang.extend(container_barang)

                    # Extract receivers only
                    for barang in container_barang:
                        receiver = barang.get('receiver_name', '-') if barang.get('receiver_name') else '-'
                        if receiver not in joa_receiver_groups[ref_joa]:
                            joa_receiver_groups[ref_joa][receiver] = set()
                        joa_receiver_groups[ref_joa][receiver].add(container_id)

                except Exception as e:
                    print(f"[ERROR] Error collecting data: {e}")
                    continue
                
                
            if not joa_receiver_groups:
                messagebox.showwarning("Peringatan", "Tidak ada data yang dapat diproses!")
                return
            
            # Convert to dialog format
            joa_groups = {}
            for ref_joa, receivers in joa_receiver_groups.items():
                joa_groups[ref_joa] = {}
                for receiver, container_set in receivers.items():
                    container_list = sorted(list(container_set))
                    joa_groups[ref_joa][("ALL_SENDERS", receiver)] = container_list
            
            # Show selection dialog - USE CUSTOM PACKING LIST DIALOG
            selected = self.show_joa_grouped_packing_list_dialog(
                joa_groups, 
                container_info,
                len(container_ids) > 1
            )
            
            if not selected:
                print("[INFO] User cancelled")
                return
            
            # Merge by receiver only
            merged_by_receiver = {}
            
            for key, container_ids_for_pdf in selected.items():
                ref_joa, sender, receiver = key
                merge_key = (ref_joa, receiver)
                
                if merge_key not in merged_by_receiver:
                    merged_by_receiver[merge_key] = set()
                merged_by_receiver[merge_key].update(container_ids_for_pdf)
            
            print(f"\n{'='*60}")
            print(f"[MERGED] {len(merged_by_receiver)} Packing List PDFs will be created")
            print(f"{'='*60}\n")
            
            total_pdfs = 0
            
            for merge_key, container_ids_set in merged_by_receiver.items():
                ref_joa, receiver = merge_key
                container_ids_for_pdf = sorted(list(container_ids_set))
                
                # Filter: get ALL senders for this receiver
                filter_criteria = {'receiver_name': receiver}
                
                try:
                    print(f"[PDF] COMBINED PACKING LIST: JOA={ref_joa} | Receiver={receiver} | Containers: {container_ids_for_pdf}")
                    # Generate combined packing list PDF
                    self.pdf_generator.generate_combined_pdf_packing_list(
                        container_ids_for_pdf, ref_joa, filter_criteria, all_barang
                    )
                    
                except Exception as pdf_error:
                    print(f"[ERROR] Failed PDF: {pdf_error}")
                    import traceback
                    print(traceback.format_exc())
            
            if total_pdfs > 0:
                messagebox.showinfo("Selesai", 
                    f"✅ Berhasil membuat {total_pdfs} file Packing List PDF!\n"
                    f"💡 Setiap PDF berisi semua pengirim untuk penerima yang sama")
            
        except Exception as e:
            import traceback
            print(f"[ERROR] {e}")
            print(traceback.print_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")

       
    def print_container_invoice_pdf(self, container_ids):
        """
        Generate and export container invoice to PDF with receiver selection
        Groups by REF JOA and RECEIVER - ONE PDF per receiver (combines all senders)
        
        Args:
            container_ids: Can be single int or list of ints
        """
        try:
            # Convert single ID to list
            if isinstance(container_ids, (int, str)):
                container_ids = [int(container_ids)]
            elif not isinstance(container_ids, list):
                container_ids = list(container_ids)
            
            print(f"\n{'='*60}")
            print(f"[INVOICE PDF] Processing {len(container_ids)} container(s)")
            print(f"{'='*60}\n")
            
            # Collect data grouped by JOA and RECEIVER ONLY (same as packing list)
            joa_receiver_groups = {}
            container_info = {}
            all_barang = []   # inisialisasi sebelum loop

            for container_id in container_ids:
                try:
                    container = self.db.get_container_by_id(container_id)
                    if not container:
                        print(f"[ERROR] Container {container_id} tidak ditemukan!")
                        continue
                    
                    container_info[container_id] = container
                    
                    ref_joa = container.get('ref_joa', '') or f'NO_JOA_{container_id}'
                    
                    if ref_joa not in joa_receiver_groups:
                        joa_receiver_groups[ref_joa] = {}
                    
                    container_barang = self.db.get_barang_in_container_with_colli_and_pricing(container_id)
                    if not container_barang:
                        print(f"[WARNING] Container {container_id} kosong!")
                        continue

                    # Convert sqlite3.Row → dict + inject container_id
                    container_barang = [{**dict(b), "container_id": container_id} for b in container_barang]

                    # ➕ Kumpulin ke all_barang
                    all_barang.extend(container_barang)

                    # Extract receivers only
                    for barang in container_barang:
                        receiver = barang.get('receiver_name', '-') if barang.get('receiver_name') else '-'
                        if receiver not in joa_receiver_groups[ref_joa]:
                            joa_receiver_groups[ref_joa][receiver] = set()
                        joa_receiver_groups[ref_joa][receiver].add(container_id)

                except Exception as e:
                    print(f"[ERROR] Error collecting data: {e}")
                    continue
                
                
            if not joa_receiver_groups:
                messagebox.showwarning("Peringatan", "Tidak ada data yang dapat diproses!")
                return
            
            # Convert to dialog format
            joa_groups = {}
            for ref_joa, receivers in joa_receiver_groups.items():
                joa_groups[ref_joa] = {}
                for receiver, container_set in receivers.items():
                    container_list = sorted(list(container_set))
                    joa_groups[ref_joa][("ALL_SENDERS", receiver)] = container_list
            
            # Show selection dialog
            selected = self.show_joa_grouped_invoice_dialog(
                joa_groups, 
                container_info,
                len(container_ids) > 1
            )
            
            if not selected:
                print("[INFO] User cancelled")
                return
            
            # Merge by receiver only
            merged_by_receiver = {}
            
            for key, container_ids_for_pdf in selected.items():
                ref_joa, sender, receiver = key
                merge_key = (ref_joa, receiver)
                
                if merge_key not in merged_by_receiver:
                    merged_by_receiver[merge_key] = set()
                merged_by_receiver[merge_key].update(container_ids_for_pdf)
            
            print(f"\n{'='*60}")
            print(f"[MERGED] {len(merged_by_receiver)} Invoice PDFs will be created")
            print(f"{'='*60}\n")
            
            total_pdfs = 0
            
            for merge_key, container_ids_set in merged_by_receiver.items():
                ref_joa, receiver = merge_key
                container_ids_for_pdf = sorted(list(container_ids_set))
                
                # Filter: get ALL senders for this receiver
                filter_criteria = {'receiver_name': receiver}
                
                try:
                    print(f"[PDF] COMBINED INVOICE: JOA={ref_joa} | Receiver={receiver} | Containers: {container_ids_for_pdf}")
                    # Generate combined invoice PDF with all_barang
                    self.pdf_generator.generate_combined_pdf_invoice_with_tax(
                        container_ids_for_pdf, ref_joa, filter_criteria, all_barang
                    )
                    total_pdfs += 1
                    
                except Exception as pdf_error:
                    print(f"[ERROR] Failed PDF: {pdf_error}")
                    import traceback
                    print(traceback.format_exc())
            
            if total_pdfs > 0:
                messagebox.showinfo("Selesai", 
                    f"✅ Berhasil membuat {total_pdfs} file Invoice PDF!\n"
                    f"💡 Setiap PDF berisi semua pengirim untuk penerima yang sama")
            
        except Exception as e:
            import traceback
            print(f"[ERROR] {e}")
            print(traceback.format_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")
        
    def show_joa_grouped_invoice_dialog(self, joa_groups, container_info, is_batch):
        """
        Show dialog with JOA grouping for invoice selection - GROUPED BY RECEIVER ONLY
        
        Args:
            joa_groups: dict {ref_joa: {(sender, receiver): [container_ids]}}
            container_info: dict {container_id: container_data}
            is_batch: bool
        
        Returns:
            dict {(ref_joa, sender, receiver): [container_ids]} or None
        """
        try:
            dialog = tk.Toplevel()
            
            if is_batch:
                title = f"Pilih Penerima - {len(container_info)} Containers (Grouped by JOA)"
            else:
                container_id = list(container_info.keys())[0]
                title = f"Pilih Penerima - Container ID {container_id}"
            
            dialog.title(title)
            dialog.geometry("900x700")
            dialog.resizable(True, True)
            dialog.grab_set()
            dialog.transient()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 450
            y = (dialog.winfo_screenheight() // 2) - 350
            dialog.geometry(f"900x700+{x}+{y}")
            
            selected_data = {}
            dialog_result = {"cancelled": True}
            
            # Main frame
            main_frame = ttk.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = ttk.Label(main_frame, text=title, font=('Arial', 14, 'bold'))
            title_label.pack(pady=(0, 10))
            
            # Instruction
            instruction = "✨ Containers dengan REF JOA yang sama akan digabung dalam satu PDF\n"
            instruction += "📋 Pilih PENERIMA untuk dibuat Invoice PDF:"
            
            ttk.Label(main_frame, text=instruction, font=('Arial', 10), 
                    justify='center').pack(pady=(0, 10))
            
            # Scrollable frame
            scroll_frame = ttk.Frame(main_frame)
            scroll_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            canvas = tk.Canvas(scroll_frame)
            scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind("<Configure>", 
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Create checkboxes grouped by JOA - GROUP BY RECEIVER ONLY
            checkbox_data = []  # [(var, ref_joa, receiver, all_combinations_for_receiver)]
            
            sorted_joas = sorted(joa_groups.keys())
            
            for joa_idx, ref_joa in enumerate(sorted_joas):
                # JOA Header
                joa_frame = ttk.LabelFrame(scrollable_frame, 
                                        text=f"📌 REF JOA: {ref_joa}",
                                        padding=10)
                joa_frame.pack(fill='x', pady=10, padx=10)
                
                combinations = joa_groups[ref_joa]
                
                # GROUP BY RECEIVER - merge all senders for same receiver
                receiver_groups = {}  # {receiver: [(sender, receiver, container_ids), ...]}
                
                for combination, container_ids in combinations.items():
                    sender, receiver = combination
                    if receiver not in receiver_groups:
                        receiver_groups[receiver] = []
                    receiver_groups[receiver].append((sender, receiver, container_ids))
                
                sorted_receivers = sorted(receiver_groups.keys())
                
                for receiver in sorted_receivers:
                    all_combinations_for_receiver = receiver_groups[receiver]
                    
                    var = tk.BooleanVar()
                    
                    # Combination frame
                    combo_frame = ttk.Frame(joa_frame)
                    combo_frame.pack(fill='x', pady=5, padx=5)
                    
                    # Main checkbox - SHOW RECEIVER ONLY
                    cb_text = f"{receiver}"
                    # Use tk.Checkbutton instead of ttk for font support
                    checkbox = tk.Checkbutton(combo_frame, text=cb_text, variable=var, 
                                            font=('Arial', 10, 'bold'),
                                            bg='white', activebackground='white',
                                            relief='flat', borderwidth=0)
                    checkbox.pack(anchor='w')
                    
                    # Show all senders for this receiver
                    all_senders = [s for s, r, _ in all_combinations_for_receiver]
                    if len(all_senders) > 0:
                        sender_text = f"   📤 Dari: {', '.join(sorted(set(all_senders)))}"
                        sender_label = ttk.Label(combo_frame, text=sender_text, 
                                            font=('Arial', 8), foreground='blue')
                        sender_label.pack(anchor='w', padx=(20, 0))
                    
                    # Container details
                    all_container_ids = []
                    for _, _, cids in all_combinations_for_receiver:
                        all_container_ids.extend(cids)
                    all_container_ids = list(set(all_container_ids))  # unique
                    
                    container_names = []
                    for cid in all_container_ids:
                        if cid in container_info:
                            container_names.append(
                                container_info[cid].get('container', f'ID {cid}')
                            )
                    
                    if len(all_container_ids) > 1:
                        info_text = f"   📦 {len(all_container_ids)} Containers akan DIGABUNG: {', '.join(container_names)}"
                        color = 'green'
                    else:
                        info_text = f"   📦 Container: {container_names[0]}"
                        color = 'gray'
                    
                    info_label = ttk.Label(combo_frame, text=info_text, 
                                        font=('Arial', 8), foreground=color)
                    info_label.pack(anchor='w', padx=(20, 0))
                    
                    # Store: var, ref_joa, receiver, all combinations for this receiver
                    checkbox_data.append((var, ref_joa, receiver, all_combinations_for_receiver))
            
            # Select buttons
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(fill='x', pady=(0, 10))
            
            def select_all():
                for var, _, _, _ in checkbox_data:
                    var.set(True)
            
            def deselect_all():
                for var, _, _, _ in checkbox_data:
                    var.set(False)
            
            ttk.Button(select_frame, text="Pilih Semua", 
                    command=select_all).pack(side='left', padx=(0, 10))
            ttk.Button(select_frame, text="Hapus Semua", 
                    command=deselect_all).pack(side='left')
            
            # Info label
            info_text = "💡 Tip: Container dengan JOA sama akan otomatis digabung dalam 1 PDF per penerima"
            ttk.Label(main_frame, text=info_text, font=('Arial', 8), 
                    foreground='blue').pack(pady=(0, 10))
            
            # Summary label
            summary_label = ttk.Label(main_frame, text="", font=('Arial', 10, 'bold'))
            summary_label.pack(pady=(0, 10))
            
            def update_summary():
                selected_count = sum(var.get() for var, _, _, _ in checkbox_data)
                total_pdfs = sum(1 for var, _, _, _ in checkbox_data if var.get())
                
                # Count total unique containers
                all_container_ids = set()
                for var, _, _, all_combos in checkbox_data:
                    if var.get():
                        for _, _, cids in all_combos:
                            all_container_ids.update(cids)
                
                summary_text = f"📊 Terpilih: {selected_count} penerima\n"
                summary_text += f"📄 {total_pdfs} PDF akan dibuat dari {len(all_container_ids)} container(s)"
                summary_label.config(text=summary_text)
            
            # Bind checkbox changes
            for var, _, _, _ in checkbox_data:
                var.trace('w', lambda *args: update_summary())
            
            update_summary()
            
            # Action buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', pady=(10, 0))
            
            def on_create():
                selected_count = sum(var.get() for var, _, _, _ in checkbox_data)
                if selected_count == 0:
                    messagebox.showwarning("Peringatan", "Pilih minimal satu penerima!")
                    return
                
                selected_data.clear()
                # For each selected receiver, include ALL combinations (all senders)
                for var, ref_joa, receiver, all_combinations_for_receiver in checkbox_data:
                    if var.get():
                        # Each combination (sender, receiver, container_ids) needs to be added
                        for sender, recv, container_ids in all_combinations_for_receiver:
                            key = (ref_joa, sender, receiver)
                            if key not in selected_data:
                                selected_data[key] = []
                            selected_data[key].extend(container_ids)
                        
                        # Remove duplicates in container_ids
                        for key in selected_data:
                            selected_data[key] = list(set(selected_data[key]))
                
                dialog_result["cancelled"] = False
                dialog.destroy()
            
            def on_cancel():
                dialog_result["cancelled"] = True
                dialog.destroy()
            
            ttk.Button(button_frame, text="✅ Buat Invoice PDF", 
                    command=on_create).pack(side='left', padx=(0, 10))
            ttk.Button(button_frame, text="❌ Batal", 
                    command=on_cancel).pack(side='right')
            
            # Keyboard bindings
            dialog.bind("<Return>", lambda e: on_create())
            dialog.bind("<Escape>", lambda e: on_cancel())
            
            # Mouse wheel scrolling
            def on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
            canvas.bind_all("<MouseWheel>", on_mousewheel)
            
            print(f"[DIALOG] Showing receiver-grouped selection: {len(checkbox_data)} receivers in {len(sorted_joas)} JOA groups")
            dialog.wait_window()
            
            canvas.unbind_all("<MouseWheel>")
            
            if dialog_result["cancelled"]:
                return None
            
            return selected_data
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Dialog error: {e}")
            print(traceback.format_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")
            return None
    
    def show_joa_grouped_packing_list_dialog(self, joa_groups, container_info, is_batch):
        """
        Show dialog with JOA grouping for PACKING LIST selection - GROUPED BY RECEIVER ONLY
        Same structure as invoice dialog but with Packing List specific text
        
        Args:
            joa_groups: dict {ref_joa: {(sender, receiver): [container_ids]}}
            container_info: dict {container_id: container_data}
            is_batch: bool
        
        Returns:
            dict {(ref_joa, sender, receiver): [container_ids]} or None
        """
        try:
            dialog = tk.Toplevel()
            
            if is_batch:
                title = f"Pilih Penerima - {len(container_info)} Containers (Grouped by JOA)"
            else:
                container_id = list(container_info.keys())[0]
                title = f"Pilih Penerima - Container ID {container_id}"
            
            dialog.title(title)
            dialog.geometry("900x700")
            dialog.resizable(True, True)
            dialog.grab_set()
            dialog.transient()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 450
            y = (dialog.winfo_screenheight() // 2) - 350
            dialog.geometry(f"900x700+{x}+{y}")
            
            selected_data = {}
            dialog_result = {"cancelled": True}
            
            # Main frame
            main_frame = ttk.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = ttk.Label(main_frame, text=title, font=('Arial', 14, 'bold'))
            title_label.pack(pady=(0, 10))
            
            # Instruction - CHANGED TO PACKING LIST
            instruction = "✨ Containers dengan REF JOA yang sama akan digabung dalam satu PDF\n"
            instruction += "📋 Pilih PENERIMA untuk dibuat Packing List PDF:"
            
            ttk.Label(main_frame, text=instruction, font=('Arial', 10), 
                    justify='center').pack(pady=(0, 10))
            
            # Scrollable frame
            scroll_frame = ttk.Frame(main_frame)
            scroll_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            canvas = tk.Canvas(scroll_frame)
            scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind("<Configure>", 
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Create checkboxes grouped by JOA - GROUP BY RECEIVER ONLY
            checkbox_data = []
            
            sorted_joas = sorted(joa_groups.keys())
            
            for joa_idx, ref_joa in enumerate(sorted_joas):
                # JOA Header
                joa_frame = ttk.LabelFrame(scrollable_frame, 
                                        text=f"📌 REF JOA: {ref_joa}",
                                        padding=10)
                joa_frame.pack(fill='x', pady=10, padx=10)
                
                combinations = joa_groups[ref_joa]
                
                # GROUP BY RECEIVER - merge all senders for same receiver
                receiver_groups = {}
                
                for combination, container_ids in combinations.items():
                    sender, receiver = combination
                    if receiver not in receiver_groups:
                        receiver_groups[receiver] = []
                    receiver_groups[receiver].append((sender, receiver, container_ids))
                
                sorted_receivers = sorted(receiver_groups.keys())
                
                for receiver in sorted_receivers:
                    all_combinations_for_receiver = receiver_groups[receiver]
                    
                    var = tk.BooleanVar()
                    
                    # Combination frame
                    combo_frame = ttk.Frame(joa_frame)
                    combo_frame.pack(fill='x', pady=5, padx=5)
                    
                    # Main checkbox - SHOW RECEIVER ONLY
                    cb_text = f"{receiver}"
                    checkbox = tk.Checkbutton(combo_frame, text=cb_text, variable=var, 
                                            font=('Arial', 10, 'bold'),
                                            bg='white', activebackground='white',
                                            relief='flat', borderwidth=0)
                    checkbox.pack(anchor='w')
                    
                    # Show all senders for this receiver
                    all_senders = [s for s, r, _ in all_combinations_for_receiver]
                    if len(all_senders) > 0:
                        sender_text = f"   📤 Dari: {', '.join(sorted(set(all_senders)))}"
                        sender_label = ttk.Label(combo_frame, text=sender_text, 
                                            font=('Arial', 8), foreground='blue')
                        sender_label.pack(anchor='w', padx=(20, 0))
                    
                    # Container details
                    all_container_ids = []
                    for _, _, cids in all_combinations_for_receiver:
                        all_container_ids.extend(cids)
                    all_container_ids = list(set(all_container_ids))
                    
                    container_names = []
                    for cid in all_container_ids:
                        if cid in container_info:
                            container_names.append(
                                container_info[cid].get('container', f'ID {cid}')
                            )
                    
                    if len(all_container_ids) > 1:
                        info_text = f"   📦 {len(all_container_ids)} Containers akan DIGABUNG: {', '.join(container_names)}"
                        color = 'green'
                    else:
                        info_text = f"   📦 Container: {container_names[0]}"
                        color = 'gray'
                    
                    info_label = ttk.Label(combo_frame, text=info_text, 
                                        font=('Arial', 8), foreground=color)
                    info_label.pack(anchor='w', padx=(20, 0))
                    
                    checkbox_data.append((var, ref_joa, receiver, all_combinations_for_receiver))
            
            # Select buttons
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(fill='x', pady=(0, 10))
            
            def select_all():
                for var, _, _, _ in checkbox_data:
                    var.set(True)
            
            def deselect_all():
                for var, _, _, _ in checkbox_data:
                    var.set(False)
            
            ttk.Button(select_frame, text="Pilih Semua", 
                    command=select_all).pack(side='left', padx=(0, 10))
            ttk.Button(select_frame, text="Hapus Semua", 
                    command=deselect_all).pack(side='left')
            
            # Info label
            info_text = "💡 Tip: Container dengan JOA sama akan otomatis digabung dalam 1 PDF per penerima"
            ttk.Label(main_frame, text=info_text, font=('Arial', 8), 
                    foreground='blue').pack(pady=(0, 10))
            
            # Summary label
            summary_label = ttk.Label(main_frame, text="", font=('Arial', 10, 'bold'))
            summary_label.pack(pady=(0, 10))
            
            def update_summary():
                selected_count = sum(var.get() for var, _, _, _ in checkbox_data)
                total_pdfs = sum(1 for var, _, _, _ in checkbox_data if var.get())
                
                all_container_ids = set()
                for var, _, _, all_combos in checkbox_data:
                    if var.get():
                        for _, _, cids in all_combos:
                            all_container_ids.update(cids)
                
                summary_text = f"📊 Terpilih: {selected_count} penerima\n"
                summary_text += f"📄 {total_pdfs} PDF akan dibuat dari {len(all_container_ids)} container(s)"
                summary_label.config(text=summary_text)
            
            # Bind checkbox changes
            for var, _, _, _ in checkbox_data:
                var.trace('w', lambda *args: update_summary())
            
            update_summary()
            
            # Action buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', pady=(10, 0))
            
            def on_create():
                selected_count = sum(var.get() for var, _, _, _ in checkbox_data)
                if selected_count == 0:
                    messagebox.showwarning("Peringatan", "Pilih minimal satu penerima!")
                    return
                
                selected_data.clear()
                for var, ref_joa, receiver, all_combinations_for_receiver in checkbox_data:
                    if var.get():
                        for sender, recv, container_ids in all_combinations_for_receiver:
                            key = (ref_joa, sender, receiver)
                            if key not in selected_data:
                                selected_data[key] = []
                            selected_data[key].extend(container_ids)
                        
                        for key in selected_data:
                            selected_data[key] = list(set(selected_data[key]))
                
                dialog_result["cancelled"] = False
                dialog.destroy()
            
            def on_cancel():
                dialog_result["cancelled"] = True
                dialog.destroy()
            
            # CHANGED BUTTON TEXT TO PACKING LIST
            ttk.Button(button_frame, text="✅ Buat Packing List PDF", 
                    command=on_create).pack(side='left', padx=(0, 10))
            ttk.Button(button_frame, text="❌ Batal", 
                    command=on_cancel).pack(side='right')
            
            # Keyboard bindings
            dialog.bind("<Return>", lambda e: on_create())
            dialog.bind("<Escape>", lambda e: on_cancel())
            
            # Mouse wheel scrolling
            def on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
            canvas.bind_all("<MouseWheel>", on_mousewheel)
            
            print(f"[DIALOG] Showing packing list receiver-grouped selection: {len(checkbox_data)} receivers in {len(sorted_joas)} JOA groups")
            dialog.wait_window()
            
            canvas.unbind_all("<MouseWheel>")
            
            if dialog_result["cancelled"]:
                return None
            
            return selected_data
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Dialog error: {e}")
            print(traceback.format_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")
            return None
    
    def show_unified_invoice_selection_dialog(self, all_combinations, container_info, is_batch):
        """
        Show unified dialog for selecting sender-receiver combinations across multiple containers
        
        Args:
            all_combinations: dict {(sender, receiver): [container_ids]}
            container_info: dict {container_id: container_data}
            is_batch: bool, whether processing multiple containers
        
        Returns:
            dict {(sender, receiver): [container_ids]} or None if cancelled
        """
        try:
            dialog = tk.Toplevel()
            
            if is_batch:
                title = f"Pilih Kombinasi - {len(container_info)} Containers"
            else:
                container_id = list(container_info.keys())[0]
                title = f"Pilih Penerima - Container ID {container_id}"
            
            dialog.title(title)
            dialog.geometry("800x600")
            dialog.resizable(False, False)
            dialog.grab_set()
            dialog.transient()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 400
            y = (dialog.winfo_screenheight() // 2) - 300
            dialog.geometry(f"800x600+{x}+{y}")
            
            selected_data = {}
            dialog_result = {"cancelled": True}
            
            # Main frame
            main_frame = ttk.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = ttk.Label(main_frame, text=title, font=('Arial', 14, 'bold'))
            title_label.pack(pady=(0, 10))
            
            # Instruction
            if is_batch:
                instruction = "Pilih kombinasi Pengirim → Penerima untuk dibuat Invoice PDF-nya:"
            else:
                instruction = "Pilih kombinasi yang ingin dibuat Invoice PDF-nya:"
            
            ttk.Label(main_frame, text=instruction, font=('Arial', 10)).pack(pady=(0, 10))
            
            # Scrollable frame
            scroll_frame = ttk.Frame(main_frame)
            scroll_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            canvas = tk.Canvas(scroll_frame)
            scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind("<Configure>", 
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Create checkboxes with container info
            checkbox_data = []  # [(var, combination, container_ids)]
            
            sorted_combinations = sorted(all_combinations.keys())
            
            for idx, combination in enumerate(sorted_combinations):
                sender, receiver = combination
                container_ids = all_combinations[combination]
                
                var = tk.BooleanVar()
                
                # Create frame for each combination
                combo_frame = ttk.Frame(scrollable_frame)
                combo_frame.pack(fill='x', pady=5, padx=10)
                
                # Main checkbox
                cb_text = f"{sender} → {receiver}"
                checkbox = ttk.Checkbutton(combo_frame, text=cb_text, variable=var)
                checkbox.pack(anchor='w')
                
                # Container info (if batch and multiple containers have this combo)
                if is_batch and len(container_ids) > 1:
                    container_names = []
                    for cid in container_ids:
                        if cid in container_info:
                            container_names.append(
                                container_info[cid].get('container', f'ID {cid}')
                            )
                    
                    info_text = f"   📦 Containers: {', '.join(container_names)}"
                    info_label = ttk.Label(combo_frame, text=info_text, 
                                        font=('Arial', 8), foreground='gray')
                    info_label.pack(anchor='w', padx=(20, 0))
                
                checkbox_data.append((var, combination, container_ids))
            
            # Select buttons
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(fill='x', pady=(0, 10))
            
            def select_all():
                for var, _, _ in checkbox_data:
                    var.set(True)
            
            def deselect_all():
                for var, _, _ in checkbox_data:
                    var.set(False)
            
            ttk.Button(select_frame, text="Pilih Semua", 
                    command=select_all).pack(side='left', padx=(0, 10))
            ttk.Button(select_frame, text="Hapus Semua", 
                    command=deselect_all).pack(side='left')
            
            # Info label
            if is_batch:
                info_text = f"💡 Tip: Setiap kombinasi akan dibuat PDF terpisah untuk setiap container"
            else:
                info_text = "💡 Tip: Pilih beberapa kombinasi untuk membuat Invoice PDF terpisah"
            
            ttk.Label(main_frame, text=info_text, font=('Arial', 8), 
                    foreground='blue').pack(pady=(0, 10))
            
            # Summary label
            summary_label = ttk.Label(main_frame, text="", font=('Arial', 9, 'bold'))
            summary_label.pack(pady=(0, 10))
            
            def update_summary():
                selected_count = sum(var.get() for var, _, _ in checkbox_data)
                total_pdfs = sum(
                    len(container_ids) for var, _, container_ids in checkbox_data 
                    if var.get()
                )
                summary_label.config(
                    text=f"📊 Terpilih: {selected_count} kombinasi → {total_pdfs} PDF akan dibuat"
                )
            
            # Bind checkbox changes to update summary
            for var, _, _ in checkbox_data:
                var.trace('w', lambda *args: update_summary())
            
            update_summary()
            
            # Action buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', pady=(10, 0))
            
            def on_create():
                selected_count = sum(var.get() for var, _, _ in checkbox_data)
                if selected_count == 0:
                    messagebox.showwarning("Peringatan", "Pilih minimal satu kombinasi!")
                    return
                
                selected_data.clear()
                for var, combination, container_ids in checkbox_data:
                    if var.get():
                        selected_data[combination] = container_ids
                
                dialog_result["cancelled"] = False
                dialog.destroy()
            
            def on_cancel():
                dialog_result["cancelled"] = True
                dialog.destroy()
            
            ttk.Button(button_frame, text="✅ Buat Invoice PDF", 
                    command=on_create).pack(side='left', padx=(0, 10))
            ttk.Button(button_frame, text="❌ Batal", 
                    command=on_cancel).pack(side='right')
            
            # Keyboard bindings
            dialog.bind("<Return>", lambda e: on_create())
            dialog.bind("<Escape>", lambda e: on_cancel())
            
            print(f"[DIALOG] Showing unified selection: {len(checkbox_data)} combinations")
            dialog.wait_window()
            
            if dialog_result["cancelled"]:
                return None
            
            return selected_data
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Dialog error: {e}")
            print(traceback.format_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")
            return None       
        
    def show_unified_invoice_selection_dialog(self, all_combinations, container_info, is_batch):
        """
        Show unified dialog for selecting sender-receiver combinations across multiple containers
        
        Args:
            all_combinations: dict {(sender, receiver): [container_ids]}
            container_info: dict {container_id: container_data}
            is_batch: bool, whether processing multiple containers
        
        Returns:
            dict {(sender, receiver): [container_ids]} or None if cancelled
        """
        try:
            dialog = tk.Toplevel()
            
            if is_batch:
                title = f"Pilih Kombinasi - {len(container_info)} Containers"
            else:
                container_id = list(container_info.keys())[0]
                title = f"Pilih Penerima - Container ID {container_id}"
            
            dialog.title(title)
            dialog.geometry("800x600")
            dialog.resizable(False, False)
            dialog.grab_set()
            dialog.transient()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 400
            y = (dialog.winfo_screenheight() // 2) - 300
            dialog.geometry(f"800x600+{x}+{y}")
            
            selected_data = {}
            dialog_result = {"cancelled": True}
            
            # Main frame
            main_frame = ttk.Frame(dialog)
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Title
            title_label = ttk.Label(main_frame, text=title, font=('Arial', 14, 'bold'))
            title_label.pack(pady=(0, 10))
            
            # Instruction
            if is_batch:
                instruction = "Pilih kombinasi Pengirim → Penerima untuk dibuat Invoice PDF-nya:"
            else:
                instruction = "Pilih kombinasi yang ingin dibuat Invoice PDF-nya:"
            
            ttk.Label(main_frame, text=instruction, font=('Arial', 10)).pack(pady=(0, 10))
            
            # Scrollable frame
            scroll_frame = ttk.Frame(main_frame)
            scroll_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            canvas = tk.Canvas(scroll_frame)
            scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind("<Configure>", 
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Create checkboxes with container info
            checkbox_data = []  # [(var, combination, container_ids)]
            
            sorted_combinations = sorted(all_combinations.keys())
            
            for idx, combination in enumerate(sorted_combinations):
                sender, receiver = combination
                container_ids = all_combinations[combination]
                
                var = tk.BooleanVar()
                
                # Create frame for each combination
                combo_frame = ttk.Frame(scrollable_frame)
                combo_frame.pack(fill='x', pady=5, padx=10)
                
                # Main checkbox
                cb_text = f"{sender} → {receiver}"
                checkbox = ttk.Checkbutton(combo_frame, text=cb_text, variable=var)
                checkbox.pack(anchor='w')
                
                # Container info (if batch and multiple containers have this combo)
                if is_batch and len(container_ids) > 1:
                    container_names = []
                    for cid in container_ids:
                        if cid in container_info:
                            container_names.append(
                                container_info[cid].get('container', f'ID {cid}')
                            )
                    
                    info_text = f"   📦 Containers: {', '.join(container_names)}"
                    info_label = ttk.Label(combo_frame, text=info_text, 
                                        font=('Arial', 8), foreground='gray')
                    info_label.pack(anchor='w', padx=(20, 0))
                
                checkbox_data.append((var, combination, container_ids))
            
            # Select buttons
            select_frame = ttk.Frame(main_frame)
            select_frame.pack(fill='x', pady=(0, 10))
            
            def select_all():
                for var, _, _ in checkbox_data:
                    var.set(True)
            
            def deselect_all():
                for var, _, _ in checkbox_data:
                    var.set(False)
            
            ttk.Button(select_frame, text="Pilih Semua", 
                    command=select_all).pack(side='left', padx=(0, 10))
            ttk.Button(select_frame, text="Hapus Semua", 
                    command=deselect_all).pack(side='left')
            
            # Info label
            if is_batch:
                info_text = f"💡 Tip: Setiap kombinasi akan dibuat PDF terpisah untuk setiap container"
            else:
                info_text = "💡 Tip: Pilih beberapa kombinasi untuk membuat Invoice PDF terpisah"
            
            ttk.Label(main_frame, text=info_text, font=('Arial', 8), 
                    foreground='blue').pack(pady=(0, 10))
            
            # Summary label
            summary_label = ttk.Label(main_frame, text="", font=('Arial', 9, 'bold'))
            summary_label.pack(pady=(0, 10))
            
            def update_summary():
                selected_count = sum(var.get() for var, _, _ in checkbox_data)
                total_pdfs = sum(
                    len(container_ids) for var, _, container_ids in checkbox_data 
                    if var.get()
                )
                summary_label.config(
                    text=f"📊 Terpilih: {selected_count} kombinasi → {total_pdfs} PDF akan dibuat"
                )
            
            # Bind checkbox changes to update summary
            for var, _, _ in checkbox_data:
                var.trace('w', lambda *args: update_summary())
            
            update_summary()
            
            # Action buttons
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', pady=(10, 0))
            
            def on_create():
                selected_count = sum(var.get() for var, _, _ in checkbox_data)
                if selected_count == 0:
                    messagebox.showwarning("Peringatan", "Pilih minimal satu kombinasi!")
                    return
                
                selected_data.clear()
                for var, combination, container_ids in checkbox_data:
                    if var.get():
                        selected_data[combination] = container_ids
                
                dialog_result["cancelled"] = False
                dialog.destroy()
            
            def on_cancel():
                dialog_result["cancelled"] = True
                dialog.destroy()
            
            ttk.Button(button_frame, text="✅ Buat Invoice PDF", 
                    command=on_create).pack(side='left', padx=(0, 10))
            ttk.Button(button_frame, text="❌ Batal", 
                    command=on_cancel).pack(side='right')
            
            # Keyboard bindings
            dialog.bind("<Return>", lambda e: on_create())
            dialog.bind("<Escape>", lambda e: on_cancel())
            
            print(f"[DIALOG] Showing unified selection: {len(checkbox_data)} combinations")
            dialog.wait_window()
            
            if dialog_result["cancelled"]:
                return None
            
            return selected_data
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Dialog error: {e}")
            print(traceback.format_exc())
            messagebox.showerror("Error", f"Error: {str(e)}")
            return None